from __future__ import annotations

import argparse
import json
import logging
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
import httpx
from openpyxl import load_workbook
import os

from supabase_rest import SupabaseRestClient


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_YEAR = 2026

CLIENT_NAMES = {
    "ECOSMART",
    "CLICKIE",
    "JUST4U",
    "TIRESIAS",
    "BAMBUTECH",
    "GBS LOGISTICS",
}

MONTHS_ES = {
    "ENERO": 1,
    "FEBRERO": 2,
    "MARZO": 3,
    "ABRIL": 4,
    "MAYO": 5,
    "JUNIO": 6,
    "JULIO": 7,
    "AGOSTO": 8,
    "SEPTIEMBRE": 9,
    "SETIEMBRE": 9,
    "OCTUBRE": 10,
    "NOVIEMBRE": 11,
    "DICIEMBRE": 12,
}


@dataclass
class ClientBlock:
    nombre: str
    slug: str
    ghl_location_id: str | None = None
    env_location_key: str | None = None
    fecha_inicio: str | None = None
    fecha_termino: str | None = None
    tipo_contrato: str = "plazo_fijo"
    duracion_meses: float | None = None
    meta: float | None = None
    reuniones_actuales: float | None = None
    pais_prospeccion: str | None = None
    pago_mensual: float | None = None
    pago_variable: float | None = None
    estado_contrato: str = "activo"
    notas: list[str] = field(default_factory=list)
    sdrs: list[dict[str, Any]] = field(default_factory=list)


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_value = re.sub(r"[^a-zA-Z0-9]+", "_", ascii_value).strip("_")
    return ascii_value.lower()


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text:
        return None
    upper = text.upper()
    if "NO HAY" in upper or "YA NO HAY" in upper:
        return None
    digits = re.sub(r"[^0-9,.-]", "", text).replace(".", "").replace(",", ".")
    if not digits:
        return None
    try:
        return float(digits)
    except ValueError:
        return None


def parse_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = clean_text(value)
    if not text:
        return None
    upper = text.upper()
    if "TERMIN" in upper and "CONTRATO" in upper:
        return None

    match = re.search(r"(\d{1,2})\s*(?:DE)?\s*([A-ZÁÉÍÓÚÑ]+)", upper)
    if not match:
        return None
    day = int(match.group(1))
    month_name = unicodedata.normalize("NFKD", match.group(2)).encode("ascii", "ignore").decode("ascii")
    month = MONTHS_ES.get(month_name)
    if not month:
        return None
    return date(DEFAULT_YEAR, month, day).isoformat()


def load_env() -> None:
    env_path = ROOT / ".env"
    env_txt_path = ROOT / ".env.txt"
    if env_path.exists():
        load_dotenv(env_path)
    elif env_txt_path.exists():
        load_dotenv(env_txt_path)


def find_excel(path_arg: str | None) -> Path:
    if path_arg:
        path = Path(path_arg)
        return path if path.is_absolute() else ROOT / path
    matches = list(ROOT.glob("*.xlsx"))
    if not matches:
        raise FileNotFoundError("No se encontro ningun archivo .xlsx en el proyecto.")
    return matches[0]


def is_client_row(row: list[Any]) -> bool:
    value = clean_text(row[3] if len(row) > 3 else None)
    return bool(value and value.upper() in CLIENT_NAMES)


def row_text(row: list[Any]) -> str:
    return " | ".join(str(v).strip() for v in row if v is not None and str(v).strip())


def infer_estado(*texts: str | None) -> str:
    joined = " ".join(t.upper() for t in texts if t)
    if "YA TERMIN" in joined or "TERMINO SU CONTRATO" in joined or "TERMINÓ SU CONTRATO" in joined:
        return "terminado"
    if "ACTIVO" in joined:
        return "activo"
    return "activo"


def looks_like_ghl_id(value: str | None) -> bool:
    if not value:
        return False
    if value.upper() == "X":
        return True
    return bool(re.fullmatch(r"[A-Za-z0-9]{12,40}", value))


def is_global_section_row(row: list[Any]) -> bool:
    texts = {clean_text(cell).upper() for cell in row if clean_text(cell)}
    return bool(texts & {"PAGO SDR", "COSTOS FIJOS", "COSTOS VARIABLES"})


def env_key_for_client(nombre: str) -> str:
    normalized = slugify(nombre).upper()
    aliases = {"GBS_LOGISTICS": "GBS"}
    suffix = aliases.get(normalized, normalized)
    return f"GHL_LOCATION_{suffix}"


def apply_contract_overrides(block: ClientBlock) -> None:
    if block.slug in {"bambutech", "gbs"}:
        block.tipo_contrato = "plazo_fijo"
        block.duracion_meses = 5
        block.notas.append("Contrato por 5 meses.")
    if block.slug == "clickie":
        block.tipo_contrato = "indefinido"
        block.duracion_meses = None
        block.fecha_inicio = None
        block.fecha_termino = None
        block.notas.append("Contrato indefinido.")


def parse_workbook(path: Path) -> dict[str, Any]:
    load_env()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    sdr_names: set[str] = set()
    for row in rows:
        if isinstance(row[1] if len(row) > 1 else None, int) and clean_text(row[2] if len(row) > 2 else None):
            sdr_names.add(clean_text(row[2]))

    blocks: list[ClientBlock] = []
    i = 0
    while i < len(rows):
        row = rows[i]
        if not is_client_row(row):
            i += 1
            continue

        name = clean_text(row[3])
        block = ClientBlock(nombre=name, slug=slugify(name))
        block.env_location_key = env_key_for_client(name)
        block.fecha_inicio = parse_date(row[4] if len(row) > 4 else None)
        block.fecha_termino = parse_date(row[5] if len(row) > 5 else None)
        block.meta = parse_number(row[6] if len(row) > 6 else None)
        block.reuniones_actuales = parse_number(row[7] if len(row) > 7 else None)
        block.pais_prospeccion = clean_text(row[8] if len(row) > 8 else None)
        block.pago_mensual = parse_number(row[9] if len(row) > 9 else None)
        block.pago_variable = parse_number(row[10] if len(row) > 10 else None)
        block.estado_contrato = infer_estado(row_text(row))
        if clean_text(row[11] if len(row) > 11 else None):
            block.notas.append(clean_text(row[11]))
        for cell in row[4:12]:
            text = clean_text(cell)
            if text and parse_date(text) is None and text.upper() not in {"YA NO HAY", "NO HAY"}:
                if len(text) > 25:
                    block.notas.append(text)

        i += 1
        while i < len(rows) and not is_client_row(rows[i]):
            current = rows[i]
            if is_global_section_row(current):
                break
            label = clean_text(current[2] if len(current) > 2 else None)
            value = clean_text(current[3] if len(current) > 3 else None)
            if label and label.upper().startswith("ID "):
                block.ghl_location_id = value
                row_pais = clean_text(current[8] if len(current) > 8 else None)
                row_pago_mensual = current[9] if len(current) > 9 else None
                row_pago_variable = current[10] if len(current) > 10 else None
                if row_pais and row_pais.upper() != "PAIS DE PROSPECCION":
                    block.pais_prospeccion = row_pais
                if parse_number(row_pago_mensual) is not None:
                    block.pago_mensual = parse_number(row_pago_mensual)
                if parse_number(row_pago_variable) is not None:
                    block.pago_variable = parse_number(row_pago_variable)
                for cell in current[4:12]:
                    text = clean_text(cell)
                    if text and len(text) > 25 and text != row_pais:
                        block.notas.append(text)
            elif label and value and (label in sdr_names or looks_like_ghl_id(value)):
                sdr_names.add(label)
                block.sdrs.append(
                    {
                        "nombre": label,
                        "slug": slugify(label),
                        "ghl_user_id": None if value.upper() == "X" else value,
                        "notas": "GHL user id pendiente" if value.upper() == "X" else None,
                    }
                )
            elif any(v is not None for v in current):
                text = row_text(current)
                if text and not text.upper().startswith("FECHA INICIO"):
                    block.notas.append(text)
            i += 1

        if not block.ghl_location_id:
            block.ghl_location_id = clean_text(os.getenv(block.env_location_key or ""))
        apply_contract_overrides(block)
        blocks.append(block)

    fixed_costs = []
    payment_rules = []
    for row in rows:
        label = clean_text(row[2] if len(row) > 2 else None)
        value = row[3] if len(row) > 3 else None
        if label in {"APOLLO.IO", "SNOV.IO", "GHL", "CORREOS"}:
            fixed_costs.append(
                {
                    "nombre": label,
                    "slug": slugify(label),
                    "monto": parse_number(value),
                    "moneda": "CLP",
                    "frecuencia": "mensual",
                }
            )

        text = row_text(row)
        upper = text.upper()
        if "100.000 SIEMPRE" in upper:
            payment_rules.append(
                {
                    "slug": "base_4_reuniones_validas_cliente",
                    "nombre": "Base por cliente con 4 reuniones validas",
                    "tipo": "base_condicional",
                    "monto": 100000,
                    "moneda": "CLP",
                    "condicion": text,
                }
            )
        elif "10.000 POR REUNION" in upper:
            payment_rules.append(
                {
                    "slug": "variable_por_reunion_valida",
                    "nombre": "Variable por reunion valida",
                    "tipo": "variable_reunion",
                    "monto": 10000,
                    "moneda": "CLP",
                    "condicion": text,
                }
            )
        elif "15.000 SI EN LA SEMANA" in upper:
            payment_rules.append(
                {
                    "slug": "bono_semanal_3_reuniones_validas",
                    "nombre": "Bono semanal por 3 reuniones validas",
                    "tipo": "bono_semanal",
                    "monto": 15000,
                    "moneda": "CLP",
                    "condicion": text,
                }
            )
        elif "5 A 7 REUNIONES" in upper or "7 A 9 REUNIONS" in upper or "10 + VALIDAS" in upper:
            amount = parse_number(row[5] if len(row) > 5 else None)
            condition = clean_text(row[4] if len(row) > 4 else None)
            if amount and condition:
                payment_rules.append(
                    {
                        "slug": f"bono_mensual_{slugify(condition)}",
                        "nombre": f"Bono mensual {condition}",
                        "tipo": "bono_mensual",
                        "monto": amount,
                        "moneda": "CLP",
                        "condicion": condition,
                    }
                )

    return {
        "sdrs": sorted(
            ({"nombre": name, "slug": slugify(name), "activo": True} for name in sdr_names),
            key=lambda x: x["slug"],
        ),
        "clientes": blocks,
        "costos_fijos": fixed_costs,
        "sdr_pago_reglas": payment_rules,
    }


def as_payload(parsed: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    now_stats: dict[str, list[dict[str, Any]]] = {
        "sdrs": parsed["sdrs"],
        "clientes": [],
        "sdr_cliente": [],
        "cliente_metas": [],
        "cliente_contratos": [],
        "cliente_costos": [],
        "costos_fijos": parsed["costos_fijos"],
        "sdr_pago_reglas": parsed["sdr_pago_reglas"],
    }

    for block in parsed["clientes"]:
        notes = "\n".join(dict.fromkeys(n for n in block.notas if n))
        now_stats["clientes"].append(
            {
                "nombre": block.nombre,
                "slug": block.slug,
                "ghl_location_id": block.ghl_location_id,
                "env_location_key": block.env_location_key,
                "pais_prospeccion": block.pais_prospeccion,
                "estado_contrato": block.estado_contrato,
                "notas": notes or None,
            }
        )
        now_stats["cliente_metas"].append(
            {
                "cliente_slug": block.slug,
                "periodo": "contrato",
                "reuniones_validas_meta": block.meta,
                "reuniones_validas_actuales": block.reuniones_actuales,
                "notas": notes or None,
            }
        )
        now_stats["cliente_contratos"].append(
            {
                "cliente_slug": block.slug,
                "fecha_inicio": block.fecha_inicio,
                "fecha_termino": block.fecha_termino,
                "tipo_contrato": block.tipo_contrato,
                "duracion_meses": block.duracion_meses,
                "estado": block.estado_contrato,
                "notas": notes or None,
            }
        )
        now_stats["cliente_costos"].append(
            {
                "cliente_slug": block.slug,
                "pago_mensual": block.pago_mensual,
                "pago_mensual_moneda": "CLP",
                "pago_variable": block.pago_variable,
                "pago_variable_moneda": "CLP",
                "notas": notes or None,
            }
        )
        for sdr in block.sdrs:
            now_stats["sdr_cliente"].append(
                {
                    "cliente_slug": block.slug,
                    "sdr_slug": sdr["slug"],
                    "ghl_user_id": sdr["ghl_user_id"],
                    "activo": True,
                    "notas": sdr["notas"],
                }
            )
    return now_stats


def upsert_table(client: Any, table: str, rows: list[dict[str, Any]], conflict: str) -> None:
    if not rows:
        return
    logging.info("Upsert %s: %s filas", table, len(rows))
    client.upsert(table, rows, conflict)



def import_to_supabase(payload: dict[str, list[dict[str, Any]]], source_file: Path) -> None:
    load_env()
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SECRET_KEY")
    if not url or not key:
        raise RuntimeError("Faltan SUPABASE_URL o SUPABASE_SECRET_KEY en .env/.env.txt")

    client = SupabaseRestClient(url, key)
    upsert_table(client, "sdrs", payload["sdrs"], "slug")
    upsert_table(client, "clientes", payload["clientes"], "slug")

    upsert_table(client, "sdr_cliente", payload["sdr_cliente"], "cliente_slug,sdr_slug")
    upsert_table(client, "cliente_metas", payload["cliente_metas"], "cliente_slug,periodo")
    upsert_table(client, "cliente_contratos", payload["cliente_contratos"], "cliente_slug")
    upsert_table(client, "cliente_costos", payload["cliente_costos"], "cliente_slug")
    upsert_table(client, "costos_fijos", payload["costos_fijos"], "slug")
    upsert_table(client, "sdr_pago_reglas", payload["sdr_pago_reglas"], "slug")

    stats = {key: len(value) for key, value in payload.items()}
    client.insert("import_runs", {"source_file": source_file.name, "status": "success", "stats": stats, "errors": []})
    logging.info("Importacion completada: %s", stats)


def main() -> None:
    parser = argparse.ArgumentParser(description="Carga configuracion SDR desde Excel a Supabase.")
    parser.add_argument("--excel", help="Ruta al archivo Excel. Por defecto usa el primer .xlsx del proyecto.")
    parser.add_argument("--dry-run", action="store_true", help="Parsea y muestra resumen sin escribir en Supabase.")
    parser.add_argument("--json", action="store_true", help="Imprime payload completo en JSON.")
    args = parser.parse_args()

    setup_logging()
    excel_path = find_excel(args.excel)
    parsed = parse_workbook(excel_path)
    payload = as_payload(parsed)

    summary = {key: len(value) for key, value in payload.items()}
    logging.info("Excel parseado: %s", excel_path.name)
    logging.info("Resumen: %s", summary)

    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))

    if args.dry_run:
        return

    import_to_supabase(payload, excel_path)


if __name__ == "__main__":
    main()
