"""
Conprospección OS — v2.0
Flujo rediseñado: mínima info → análisis → estructura → firma
"""
import streamlit as st
import json
import re
import shutil
import base64
import sys
import os
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

# Cargar .env desde la raíz del proyecto; si no existe, busca en mvp_setup/
try:
    from dotenv import load_dotenv
    _root = Path(__file__).resolve().parent.parent
    for _cand in [_root / ".env", Path(__file__).parent / ".env"]:
        if _cand.exists():
            load_dotenv(_cand, override=True)
            break
except Exception:
    pass

from config import APP_NAME, APP_VERSION, CLIENTES_DIR, ETAPAS
from modules.estructura import crear_estructura_carpetas
from modules.archivos import guardar_archivo, listar_logos, listar_archivos_por_carpeta
from modules.firma import generar_firma_html, generar_firma_texto
from modules.templates import generar_todos_los_archivos
from modules.estado import cargar_estado, guardar_estado, actualizar_campo, calcular_progreso, emoji_estado

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Conprospección OS",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded",
)

PAISES_LA = [
    "Argentina", "Bolivia", "Brasil", "Chile", "Colombia",
    "Costa Rica", "Ecuador", "El Salvador", "España",
    "Estados Unidos", "Guatemala", "Honduras", "México",
    "Nicaragua", "Panamá", "Paraguay", "Perú",
    "Puerto Rico", "República Dominicana", "Uruguay",
    "Venezuela", "Otro",
]

OBJETIVOS = [
    "🤝 Conseguir reunión comercial calificada",
    "🔍 Conocer y mapear el mercado objetivo",
    "💡 Generar interés / awareness de marca",
    "📞 Conseguir llamada de discovery",
    "📝 Agendar demo del producto o servicio",
    "🔗 Conectar y hacer networking en LinkedIn",
    "🌱 Expandir en nuevas industrias o segmentos",
    "🏆 Reactivar clientes perdidos o inactivos",
    "📊 Validar propuesta de valor en nuevo mercado",
    "🎯 Generar pipeline calificado para el equipo de ventas",
]

# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
/* Sidebar oscura */
[data-testid="stSidebar"] {
  background: #0f172a !important;
}
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] div {
  color: #cbd5e1 !important;
}
[data-testid="stSidebar"] hr {
  border-color: #1e293b !important;
}
/* Botones sidebar */
[data-testid="stSidebar"] .stButton > button {
  background: transparent !important;
  border: none !important;
  color: #94a3b8 !important;
  text-align: left !important;
  padding: 8px 14px !important;
  border-radius: 8px !important;
  font-size: 13px !important;
  width: 100%;
  transition: all 0.15s;
}
[data-testid="stSidebar"] .stButton > button:hover {
  background: #1e293b !important;
  color: #f1f5f9 !important;
}
/* Botón activo */
[data-testid="stSidebar"] .btn-activo > button {
  background: #1e3a5f !important;
  color: #93c5fd !important;
  font-weight: 600 !important;
}
/* Steps */
.step-card {
  background: white;
  border: 1.5px solid #e2e8f0;
  border-radius: 12px;
  padding: 20px 24px;
  margin-bottom: 12px;
  transition: all 0.2s;
}
.step-card.activo {
  border-color: #3b82f6;
  box-shadow: 0 0 0 3px rgba(59,130,246,0.1);
}
.step-card.completo {
  border-color: #10b981;
  background: #f0fdf4;
}
.step-card.pendiente {
  opacity: 0.55;
}
/* Pill objetivo */
.pill {
  display: inline-block;
  padding: 6px 16px;
  border-radius: 999px;
  font-size: 13px;
  font-weight: 600;
  background: #eff6ff;
  color: #1d4ed8;
  margin: 4px;
  border: 1.5px solid #bfdbfe;
  cursor: pointer;
}
/* Metric cards */
.metric-card {
  background: white;
  border: 1px solid #e2e8f0;
  border-radius: 12px;
  padding: 16px 20px;
  text-align: center;
}
.metric-num {
  font-size: 28px;
  font-weight: 800;
  color: #1a1a2e;
  line-height: 1.2;
}
.metric-label {
  font-size: 12px;
  color: #64748b;
  margin-top: 4px;
}
/* Etapa badge */
.etapa-badge {
  background: #dbeafe;
  color: #1e40af;
  font-size: 11px;
  font-weight: 700;
  padding: 3px 10px;
  border-radius: 999px;
  display: inline-block;
}
/* File row */
.file-tag {
  display: inline-block;
  background: #f1f5f9;
  border: 1px solid #e2e8f0;
  border-radius: 6px;
  padding: 4px 10px;
  font-size: 12px;
  margin: 2px;
  color: #334155;
}
/* Análisis placeholder */
.placeholder-box {
  background: linear-gradient(135deg, #f8fafc 0%, #f1f5f9 100%);
  border: 2px dashed #cbd5e1;
  border-radius: 12px;
  padding: 32px;
  text-align: center;
  color: #64748b;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────
def init_state():
    defaults = {
        "page": "inicio",
        "cliente_id": None,
        "step": 1,
        "firma_html": None,
        # Wizard nuevo cliente
        "wiz_step": 1,
        "wiz_nombre": "",
        "wiz_web": "",
        "wiz_paises": [],
        "wiz_objetivo": "",
        "wiz_archivos": [],   # lista de {"nombre": str, "bytes": bytes}
        "wiz_analisis": "",
        "wiz_descripcion": "",
        # Cards de análisis
        "wiz_resumen_servicio": "",
        "wiz_propuesta_valor": "",
        "wiz_problema": "",
        "wiz_icp_tipo_cliente": "",
        "wiz_diferenciacion": "",
        # Campos importantes del cliente
        "wiz_industrias_no_prospectar": "",
        "wiz_clientes_actuales": "",
        "wiz_competidores": "",
        "wiz_ticket_promedio": "",
        "wiz_canales_actuales": "",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def normalizar(nombre: str) -> str:
    n = nombre.strip().upper()
    for a, b in [("Á","A"),("É","E"),("Í","I"),("Ó","O"),("Ú","U"),("Ñ","N"),("á","A"),("é","E"),("í","I"),("ó","O"),("ú","U"),("ñ","N")]:
        n = n.replace(a, b)
    n = re.sub(r"[^A-Z0-9\s\-_]", "", n)
    n = re.sub(r"\s+", "_", n.strip())
    return n

def ir(page, cliente_id=None, step=1):
    st.session_state.page = page
    if cliente_id is not None:
        st.session_state.cliente_id = cliente_id
    st.session_state.step = step

def ir_a_tab(indice: int):
    """Navega programáticamente al tab indicado (0-based) via JS."""
    import streamlit.components.v1 as _components
    _components.html(f"""
    <script>
      (function() {{
        // Espera a que los tabs estén renderizados y hace click en el indicado
        function clickTab() {{
          var tabs = window.parent.document.querySelectorAll('[data-baseweb="tab"]');
          if (tabs && tabs.length > {indice}) {{
            tabs[{indice}].click();
          }} else {{
            setTimeout(clickTab, 100);
          }}
        }}
        setTimeout(clickTab, 80);
      }})();
    </script>
    """, height=0)

def c_dir(cid: str) -> Path:
    return CLIENTES_DIR / cid

def get_clientes() -> list:
    if not CLIENTES_DIR.exists():
        return []
    result = []
    for d in sorted(CLIENTES_DIR.iterdir()):
        if d.is_dir():
            est = cargar_estado(d)
            result.append({
                "id": d.name,
                "nombre": est.get("nombre_cliente", d.name),
                "web": est.get("sitio_web", ""),
                "estado": est.get("estado_general", "creado"),
                "etapa": est.get("etapa_actual", 1),
                "etapa_nombre": est.get("etapa_nombre", "Inicio"),
                "paises": est.get("pais_objetivo", ""),
                "objetivo": est.get("objetivo_comercial", ""),
                "dir": d,
            })
    return result

def imagen_b64(path: str) -> str:
    try:
        data = open(path, "rb").read()
        ext = Path(path).suffix.lower().replace(".", "")
        mime = {"svg": "image/svg+xml", "jpg": "image/jpeg", "jpeg": "image/jpeg"}.get(ext, f"image/{ext}")
        return f"data:{mime};base64,{base64.b64encode(data).decode()}"
    except:
        return ""


# ─────────────────────────────────────────────
# CLAUDE API
# ─────────────────────────────────────────────
def get_claude_client():
    api_key = os.environ.get("ANTHROPIC_API_KEY", "").strip()
    if not api_key or api_key.startswith("sk-ant-PEGA"):
        return None
    try:
        import anthropic
        return anthropic.Anthropic(api_key=api_key)
    except Exception:
        return None


def build_icp_system_prompt(est: dict, path: Path) -> str:
    nombre = est.get("nombre_cliente", "")
    web = est.get("sitio_web", "")
    paises = est.get("pais_objetivo", "")
    objetivo = est.get("objetivo_comercial", "")
    resumen = est.get("resumen_servicio", "")
    propuesta = est.get("propuesta_valor", "")
    problema = est.get("problema_que_resuelve", "")
    icp_inicial = est.get("icp_tipo_cliente", "")
    diferenciacion = est.get("diferenciacion", "")
    industrias_no = est.get("industrias_no_prospectar", "")
    clientes_actuales = est.get("clientes_actuales", "")
    competidores = est.get("competidores", "")
    ticket = est.get("ticket_promedio", "")
    canales = est.get("canales_actuales", "")

    # Leer TODOS los documentos subidos (PDF, DOCX, TXT, MD, etc.)
    docs_content = leer_todos_los_documentos(path, max_por_doc=4000)

    return f"""Eres un experto senior en prospección B2B y definición de ICP (Ideal Customer Profile) para la agencia Conprospección. Tu trabajo es EXTRAER y estructurar el ICP a partir de los documentos que el cliente ya proporcionó, complementando con el análisis web.

CLIENTE: {nombre}
Web: {web}
Países objetivo: {paises}
Objetivo comercial: {objetivo}

ANÁLISIS DEL CLIENTE:
- Resumen del servicio: {resumen or '(no completado)'}
- Propuesta de valor: {propuesta or '(no completado)'}
- Problema que resuelve: {problema or '(no completado)'}
- Diferenciación: {diferenciacion or '(no completado)'}
- Tipo de cliente ideal (preliminar): {icp_inicial or '(no completado)'}

INFORMACIÓN ADICIONAL:
- Clientes actuales del cliente: {clientes_actuales or '(no especificado)'}
- Competidores: {competidores or '(no especificado)'}
- Ticket/contrato promedio: {ticket or '(no especificado)'}
- Canales de venta actuales: {canales or '(no especificado)'}
- Industrias a NO prospectar: {industrias_no or '(no especificado)'}
{f'DOCUMENTOS DEL CLIENTE:{docs_content}' if docs_content else ''}

REGLA FUNDAMENTAL — PRIORIDAD TOTAL A LOS DOCUMENTOS DEL CLIENTE:
1. Si el cliente subió documentos (PDF, DOCX, presentación, brief), EXTRAE el ICP desde ahí. No lo inventes.
2. El ICP son los COMPRADORES/CLIENTES de {nombre}, NUNCA sus competidores del mismo rubro.
   Ejemplo: si {nombre} es freight forwarding/logística, sus clientes son importadores, exportadores,
   retailers, manufactureros — los que CONTRATAN logística — JAMÁS otras empresas de logística.
3. Si hay contradicción entre documentos y análisis web, los documentos del cliente tienen prioridad.
{f'DOCUMENTOS DEL CLIENTE (lee esto primero):{docs_content}' if docs_content else '⚠️ Sin documentos subidos aún — el cliente debe subir su brief o presentación.'}

INSTRUCCIONES:
- Extrae cargos, industrias, tamaños y criterios DIRECTAMENTE de los documentos cuando estén disponibles
- Cuando des listas de cargos o industrias, incluye SIEMPRE versión separada por coma lista para Apollo.io
- El ICP debe ser específico y accionable — nada genérico
- El mercado principal es Latinoamérica (Chile, México, Colombia, Perú, Argentina, etc.) salvo que se especifique otro
- Cuando te pidan empresas reales, busca nombres concretos que calcen con el ICP definido
- Si te falta información en los documentos, dilo explícitamente y pregunta"""


def llamar_claude(client, messages: list, system: str, max_tokens: int = 2000) -> str:
    import time as _time
    intentos = 0
    espera = 15  # segundos entre reintentos por rate limit
    while intentos < 4:
        try:
            resp = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=max_tokens,
                system=system,
                messages=messages,
            )
            return resp.content[0].text
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate_limit" in msg.lower() or "overloaded" in msg.lower():
                intentos += 1
                if intentos < 4:
                    _time.sleep(espera)
                    espera = min(espera * 2, 60)  # backoff: 15 → 30 → 60 seg
                    continue
            return f"❌ Error al llamar a Claude: {msg}"
    return "❌ Error: límite de tasa de la API superado. Espera 1 minuto y vuelve a intentar."


def grabar_y_transcribir(duracion: int = 20, idioma: str = "es-CL") -> str:
    """Graba audio desde el micrófono del servidor y transcribe con Google Speech."""
    try:
        import sounddevice as sd
        import scipy.io.wavfile as wav
        import speech_recognition as sr
        import tempfile, os
        samplerate = 16000
        audio = sd.rec(int(duracion * samplerate), samplerate=samplerate, channels=1, dtype='int16')
        sd.wait()
        with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as f:
            tmp_path = f.name
        wav.write(tmp_path, samplerate, audio)
        r = sr.Recognizer()
        with sr.AudioFile(tmp_path) as source:
            audio_data = r.record(source)
        os.unlink(tmp_path)
        return r.recognize_google(audio_data, language=idioma)
    except Exception as e:
        return f"ERROR: {e}"


def leer_documento(archivo: Path, max_chars: int = 4000) -> str:
    """Lee el contenido de un archivo (PDF, DOCX, TXT, MD, CSV, HTML)."""
    ext = archivo.suffix.lower()
    try:
        if ext in (".txt", ".md", ".csv"):
            return archivo.read_text(encoding="utf-8", errors="ignore")[:max_chars]
        elif ext == ".pdf":
            import pypdf
            reader = pypdf.PdfReader(str(archivo))
            texto = "\n".join(p.extract_text() or "" for p in reader.pages)
            return texto[:max_chars]
        elif ext in (".docx", ".doc"):
            import docx as _docx
            doc = _docx.Document(str(archivo))
            texto = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            return texto[:max_chars]
        elif ext in (".html", ".htm"):
            import re as _re
            raw = archivo.read_text(encoding="utf-8", errors="ignore")
            txt = _re.sub(r'<[^>]+>', ' ', raw)
            return _re.sub(r'\s+', ' ', txt).strip()[:max_chars]
        elif ext in (".xlsx", ".xls"):
            import pandas as _pd
            df = _pd.read_excel(str(archivo), nrows=200)
            return df.to_string(index=False)[:max_chars]
    except Exception as e:
        return f"[Error leyendo {archivo.name}: {e}]"
    return ""


def _sync_supabase_cliente(path: Path):
    """Sincroniza datos del cliente con Supabase si hay credenciales configuradas."""
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_KEY", "").strip()
    if not url or not key or url.startswith("https://TU"):
        return  # Sin configurar
    try:
        import requests as req
        est = cargar_estado(path)
        com_file = path / "07_BASE_DATOS" / "comercial.json"
        com = {}
        if com_file.exists():
            try: com = json.loads(com_file.read_text(encoding="utf-8"))
            except: pass
        payload = {
            "nombre": est.get("nombre_cliente", ""),
            "web": est.get("sitio_web", ""),
            "paises": est.get("pais_objetivo", ""),
            "objetivo": est.get("objetivo_comercial", ""),
            "industria": est.get("industria", ""),
            "prospector": est.get("nombre_prospector", ""),
            "moneda": com.get("moneda", "CLP"),
            "monto_fijo": com.get("monto_fijo", 0),
            "meses_contrato": com.get("meses_contrato", 5),
            "reuniones_garantizadas": com.get("reuniones_garantizadas", 0),
            "fecha_inicio": com.get("fecha_inicio_contrato", ""),
            "estado": est.get("estado_general", "creado"),
            "updated_at": datetime.now().isoformat(),
        }
        headers = {
            "apikey": key,
            "Authorization": f"Bearer {key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates",
        }
        req.post(f"{url}/rest/v1/clientes", json=payload, headers=headers, timeout=5)
    except Exception:
        pass  # Supabase no disponible, continuar sin error


def leer_todos_los_documentos(path: Path, max_por_doc: int = 4000) -> str:
    """Lee todos los documentos del cliente y los devuelve como string concatenado."""
    carpetas = [
        path / "00_INPUT_CLIENTE" / "documentos",
        path / "00_INPUT_CLIENTE" / "otros",
        path / "00_INPUT_CLIENTE" / "bases",
    ]
    exts_validas = {".pdf", ".docx", ".doc", ".txt", ".md", ".csv", ".html", ".htm", ".xlsx"}
    resultado = ""
    archivos_leidos = 0
    for carpeta in carpetas:
        if not carpeta.exists():
            continue
        for f in sorted(carpeta.iterdir()):
            if f.is_file() and f.suffix.lower() in exts_validas and archivos_leidos < 8:
                contenido = leer_documento(f, max_por_doc)
                if contenido.strip():
                    resultado += f"\n\n{'='*50}\n📄 DOCUMENTO: {f.name}\n{'='*50}\n{contenido}"
                    archivos_leidos += 1
    return resultado


def extraer_comercial_de_propuesta(client, texto: str, nombre_cliente: str) -> dict:
    """Llama a Claude para extraer datos comerciales desde el texto de una propuesta."""
    system = """Eres un asistente que extrae datos comerciales de propuestas de servicios de prospección B2B.
Devuelve SÓLO un JSON con exactamente estas claves (usa null si no se encuentra el dato):
{
  "moneda": "CLP" o "USD",
  "monto_setup": número (pago único de onboarding/setup al inicio del contrato, si lo hay),
  "monto_fijo": número (monto mensual fijo del contrato),
  "meses_contrato": número (duración total en meses),
  "semanas_setup": número (semanas de onboarding/setup antes de prospección activa),
  "reuniones_garantizadas": número (total de reuniones comprometidas en el contrato),
  "costo_reunion": número (precio por reunión adicional o variable, en USD),
  "notas_comerciales": "texto breve con condiciones especiales, bonos, garantías u observaciones relevantes"
}
Si hay rango de precios, toma el valor base/mínimo. Si el monto está en CLP y es >= 100000, moneda = CLP. Si está en USD, moneda = USD."""

    prompt = f"""Analiza esta propuesta comercial del cliente "{nombre_cliente}" y extrae los datos del contrato:

{texto[:6000]}

Responde SÓLO con el JSON, sin texto adicional."""

    try:
        raw = llamar_claude(client, [{"role": "user", "content": prompt}], system, max_tokens=600)
        txt = raw.strip()
        start = txt.find('{')
        end   = txt.rfind('}')
        if start >= 0 and end > start:
            return json.loads(txt[start:end+1])
    except Exception:
        pass
    return {}


def actualizar_resumen_global_clientes():
    """Genera/actualiza un MD global con la rentabilidad de todos los clientes."""
    import datetime as dt
    clientes = get_clientes()
    lineas = [
        "# Resumen comercial — Todos los clientes",
        f"_Actualizado: {dt.date.today().strftime('%d/%m/%Y')}_\n",
        "| Cliente | Moneda | Setup único | Fijo/mes | Meses | Reu. gar. | Costo/reu | Total contrato | Inicio prosp. | Fin contrato |",
        "|---------|--------|-------------|----------|-------|-----------|-----------|---------------|---------------|-------------|",
    ]
    for c in clientes:
        com_file = Path(c["dir"]) / "07_BASE_DATOS" / "comercial.json"
        com = {}
        if com_file.exists():
            try: com = json.loads(com_file.read_text(encoding="utf-8"))
            except: pass
        nombre   = c["nombre"]
        moneda   = com.get("moneda", "CLP")
        setup_p  = int(com.get("monto_setup", 0))
        fijo     = int(com.get("monto_fijo", 0))
        meses    = int(com.get("meses_contrato", 5))
        reu_gar  = int(com.get("reuniones_garantizadas", 0))
        costo_r  = int(com.get("costo_reunion", 0))
        total    = setup_p + fijo * meses + reu_gar * costo_r
        fp       = com.get("fecha_inicio_prospeccion", "—")
        fc_str   = com.get("fecha_inicio_contrato", "")
        fin      = "—"
        if fc_str:
            try:
                fin = (dt.date.fromisoformat(fc_str) + dt.timedelta(days=meses * 30.4)).strftime('%d/%m/%Y')
            except: pass
        try: fp_fmt = dt.date.fromisoformat(fp).strftime('%d/%m/%Y')
        except: fp_fmt = fp

        lineas.append(
            f"| {nombre} | {moneda} | {setup_p:,} | {fijo:,} | {meses} | {reu_gar} | {costo_r} USD | {moneda} {total:,} | {fp_fmt} | {fin} |"
        )

    resumen_path = CLIENTES_DIR.parent / "resumen_comercial_clientes.md"
    resumen_path.write_text("\n".join(lineas), encoding="utf-8")
    return resumen_path


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
def sidebar():
    with st.sidebar:
        st.markdown("""
        <div style="padding: 16px 8px 8px;">
          <div style="font-size:20px;font-weight:800;color:#f1f5f9;letter-spacing:-0.5px;">🚀 Conprospección OS</div>
          <div style="font-size:11px;color:#475569;margin-top:2px;">v2.0 MVP</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<hr style="border-color:#1e293b;margin:8px 0;">', unsafe_allow_html=True)

        # Nav principal
        is_inicio = st.session_state.page == "inicio"
        is_nuevo = st.session_state.page == "nuevo_cliente"

        col_inicio = "btn-activo" if is_inicio else ""
        with st.container():
            st.markdown(f'<div class="{col_inicio}">', unsafe_allow_html=True)
            if st.button("🏠  Inicio", key="nav_inicio", use_container_width=True):
                ir("inicio")
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        col_nuevo = "btn-activo" if is_nuevo else ""
        with st.container():
            st.markdown(f'<div class="{col_nuevo}">', unsafe_allow_html=True)
            if st.button("➕  Nuevo cliente", key="nav_nuevo", use_container_width=True):
                ir("nuevo_cliente")
                st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<hr style="border-color:#1e293b;margin:8px 0;">', unsafe_allow_html=True)
        st.markdown('<div style="font-size:10px;color:#334155;text-transform:uppercase;letter-spacing:1.5px;padding:0 8px;margin-bottom:4px;">Clientes</div>', unsafe_allow_html=True)

        clientes = get_clientes()
        if not clientes:
            st.markdown('<div style="font-size:12px;color:#475569;padding:4px 14px;">Sin clientes aún</div>', unsafe_allow_html=True)
        else:
            for c in clientes:
                es_activo = st.session_state.page == "cliente" and st.session_state.cliente_id == c["id"]
                cls = "btn-activo" if es_activo else ""
                with st.container():
                    st.markdown(f'<div class="{cls}">', unsafe_allow_html=True)
                    label = f"📁  {c['nombre']}"
                    if st.button(label, key=f"nav_{c['id']}", use_container_width=True):
                        ir("cliente", c["id"])
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

        # Info cliente activo en sidebar
        if st.session_state.page == "cliente" and st.session_state.cliente_id:
            cid = st.session_state.cliente_id
            est = cargar_estado(CLIENTES_DIR / cid)
            st.markdown('<hr style="border-color:#1e293b;margin:8px 0;">', unsafe_allow_html=True)
            prog = calcular_progreso(est)
            st.markdown(f"""
            <div style="padding:8px 12px;background:#0f2040;border-radius:8px;margin:4px;">
              <div style="font-size:11px;color:#64748b;">PROGRESO</div>
              <div style="font-size:18px;font-weight:800;color:#93c5fd;">{prog['porcentaje']}%</div>
              <div style="background:#1e293b;border-radius:4px;height:4px;margin-top:6px;">
                <div style="background:#3b82f6;height:4px;border-radius:4px;width:{prog['porcentaje']}%;"></div>
              </div>
              <div style="font-size:10px;color:#475569;margin-top:4px;">{prog['completados']}/{prog['total']} módulos</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown('<hr style="border-color:#1e293b;margin-top:auto;">', unsafe_allow_html=True)
        st.markdown('<div style="font-size:11px;color:#334155;padding:8px;text-align:center;">Conprospección © 2025</div>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# INICIO
# ─────────────────────────────────────────────
def page_inicio():
    clientes = get_clientes()

    # Header
    st.markdown("""
    <div style="margin-bottom:24px;">
      <h1 style="font-size:26px;font-weight:800;margin:0;color:#0f172a;">Centro de Operaciones</h1>
      <p style="color:#64748b;margin:4px 0 0;font-size:14px;">Gestiona cada cliente desde el onboarding hasta la operación.</p>
    </div>
    """, unsafe_allow_html=True)

    # Métricas — calcular facturación mensual
    _fijo_total = 0
    for _c in clientes:
        _cf = Path(_c["dir"]) / "07_BASE_DATOS" / "comercial.json"
        if _cf.exists():
            try: _fijo_total += int(json.loads(_cf.read_text(encoding="utf-8")).get("monto_fijo", 0))
            except: pass

    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        st.markdown(f'<div class="metric-card"><div class="metric-num">{len(clientes)}</div><div class="metric-label">Clientes totales</div></div>', unsafe_allow_html=True)
    with col2:
        en_setup = sum(1 for c in clientes if c["etapa"] <= 3)
        st.markdown(f'<div class="metric-card"><div class="metric-num">{en_setup}</div><div class="metric-label">En setup</div></div>', unsafe_allow_html=True)
    with col3:
        en_operacion = sum(1 for c in clientes if c["etapa"] >= 8)
        st.markdown(f'<div class="metric-card"><div class="metric-num">{en_operacion}</div><div class="metric-label">En operación</div></div>', unsafe_allow_html=True)
    with col4:
        listos = sum(1 for c in clientes if c["estado"] == "listo_para_operacion")
        st.markdown(f'<div class="metric-card"><div class="metric-num">{listos}</div><div class="metric-label">Listos para operar</div></div>', unsafe_allow_html=True)
    with col5:
        _fijo_fmt = f"${_fijo_total:,.0f}" if _fijo_total > 0 else "—"
        st.markdown(f'<div class="metric-card"><div class="metric-num" style="font-size:18px;">{_fijo_fmt}</div><div class="metric-label">Fijo mensual total</div></div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if not clientes:
        st.markdown("""
        <div style="background:#f8fafc;border:2px dashed #cbd5e1;border-radius:16px;padding:48px;text-align:center;">
          <div style="font-size:40px;margin-bottom:12px;">🚀</div>
          <div style="font-size:18px;font-weight:700;color:#1a1a2e;margin-bottom:6px;">Aún no tienes clientes</div>
          <div style="font-size:14px;color:#64748b;">Crea tu primer cliente para comenzar</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("➕  Crear primer cliente", type="primary", use_container_width=False):
            ir("nuevo_cliente")
            st.rerun()
        return

    # Lista de clientes
    st.markdown("### Clientes")

    if "confirmar_eliminar" not in st.session_state:
        st.session_state["confirmar_eliminar"] = None

    for c in clientes:
        est = cargar_estado(c["dir"])
        prog = calcular_progreso(est)
        pendiente = est.get("pendientes", ["—"])[0] if est.get("pendientes") else "—"
        confirmando = st.session_state["confirmar_eliminar"] == c["id"]

        with st.container():
            if confirmando:
                st.markdown(
                    f'<div style="background:#fff1f2;border:1.5px solid #fca5a5;border-radius:10px;'
                    f'padding:12px 16px;margin:4px 0;">'
                    f'<span style="font-size:13px;color:#991b1b;font-weight:700;">⚠️ ¿Eliminar "{c["nombre"]}" y toda su carpeta? Esta acción no se puede deshacer.</span>'
                    f'</div>', unsafe_allow_html=True)
                bc1, bc2, bc3 = st.columns([2, 1, 1])
                with bc2:
                    if st.button("🗑️ Sí, eliminar", key=f"confirm_del_{c['id']}", type="primary"):
                        dir_path = Path(c["dir"])
                        try:
                            shutil.rmtree(str(dir_path))
                            st.session_state["confirmar_eliminar"] = None
                            st.rerun()
                        except Exception as e:
                            # En Windows a veces hay archivos bloqueados — forzar archivo por archivo
                            import os
                            errores = []
                            for root, dirs, files in os.walk(str(dir_path), topdown=False):
                                for f in files:
                                    try: os.remove(os.path.join(root, f))
                                    except Exception as ef: errores.append(str(ef))
                                for d in dirs:
                                    try: os.rmdir(os.path.join(root, d))
                                    except: pass
                            try: os.rmdir(str(dir_path))
                            except: pass
                            if dir_path.exists():
                                st.error(f"No se pudo eliminar completamente: {errores[:2]}")
                            else:
                                st.session_state["confirmar_eliminar"] = None
                                st.rerun()
                with bc3:
                    if st.button("Cancelar", key=f"cancel_del_{c['id']}"):
                        st.session_state["confirmar_eliminar"] = None
                        st.rerun()
            else:
                col1, col2, col3, col4, col5 = st.columns([3, 3, 3, 1, 1])

                with col1:
                    _com_c = {}
                    _com_f = Path(c["dir"]) / "07_BASE_DATOS" / "comercial.json"
                    if _com_f.exists():
                        try: _com_c = json.loads(_com_f.read_text(encoding="utf-8"))
                        except: pass
                    _plan = _com_c.get("plan_contratado", "")
                    _plan_badge = ""
                    if _plan and _plan != "— Sin definir —":
                        _col = "#dbeafe" if _plan == "Starter" else "#f3e8ff"
                        _tc  = "#1e40af" if _plan == "Starter" else "#6b21a8"
                        _plan_badge = f'<span style="background:{_col};color:{_tc};font-size:10px;font-weight:700;padding:1px 7px;border-radius:999px;margin-left:6px;">{_plan}</span>'
                    st.markdown(f"""
                    <div style="padding:4px 0;">
                      <div style="font-weight:700;font-size:15px;color:#0f172a;">{c['nombre']}{_plan_badge}</div>
                      <div style="font-size:12px;color:#64748b;">🌐 {c['web'] or '—'}</div>
                      <div style="font-size:11px;color:#94a3b8;margin-top:2px;">📍 {c['paises'] or '—'}</div>
                    </div>
                    """, unsafe_allow_html=True)

                with col2:
                    etapa_n = est.get("etapa_actual", 1)
                    etapa_nom = ETAPAS[etapa_n - 1]["nombre"] if etapa_n <= len(ETAPAS) else "—"
                    etapa_ico = ETAPAS[etapa_n - 1]["icono"] if etapa_n <= len(ETAPAS) else "📌"
                    st.markdown(f"""
                    <div style="padding:4px 0;">
                      <span class="etapa-badge">{etapa_ico} {etapa_nom}</span>
                      <div style="margin-top:8px;">
                        <div style="background:#e2e8f0;border-radius:999px;height:6px;overflow:hidden;">
                          <div style="background:#3b82f6;height:6px;width:{prog['porcentaje']}%;border-radius:999px;"></div>
                        </div>
                        <div style="font-size:11px;color:#64748b;margin-top:3px;">{prog['porcentaje']}% completado</div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)

                with col3:
                    st.markdown(f"""
                    <div style="padding:4px 0;">
                      <div style="font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:0.5px;">SIGUIENTE PASO</div>
                      <div style="font-size:13px;color:#374151;margin-top:2px;">⏳ {pendiente}</div>
                    </div>
                    """, unsafe_allow_html=True)

                with col4:
                    if st.button("Abrir →", key=f"open_{c['id']}", type="primary"):
                        ir("cliente", c["id"])
                        st.rerun()

                with col5:
                    if st.button("🗑️", key=f"del_{c['id']}", help="Eliminar cliente"):
                        st.session_state["confirmar_eliminar"] = c["id"]
                        st.rerun()

            st.markdown('<hr style="margin:8px 0;border-color:#f1f5f9;">', unsafe_allow_html=True)

    # ── Tabla de rentabilidad ─────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    with st.expander("💰 Rentabilidad comparativa de clientes", expanded=False):
        import datetime as dt
        filas = []
        for c in clientes:
            com_file = Path(c["dir"]) / "07_BASE_DATOS" / "comercial.json"
            com_c = {}
            if com_file.exists():
                try: com_c = json.loads(com_file.read_text(encoding="utf-8"))
                except: pass
            moneda  = com_c.get("moneda", "—")
            setup_p = int(com_c.get("monto_setup", 0))
            fijo    = int(com_c.get("monto_fijo", 0))
            meses   = int(com_c.get("meses_contrato", 5))
            reu_gar = int(com_c.get("reuniones_garantizadas", 0))
            costo_r = int(com_c.get("costo_reunion", 0))
            total   = setup_p + fijo * meses + reu_gar * costo_r
            fp      = com_c.get("fecha_inicio_prospeccion", "")
            fin_str = "—"
            fp_fmt  = "—"
            if com_c.get("fecha_inicio_contrato"):
                try:
                    fin_str = (dt.date.fromisoformat(com_c["fecha_inicio_contrato"])
                               + dt.timedelta(days=meses * 30.4)).strftime('%d/%m/%Y')
                except: pass
            if fp:
                try: fp_fmt = dt.date.fromisoformat(fp).strftime('%d/%m/%Y')
                except: fp_fmt = fp
            notas = com_c.get("notas_comerciales", "")
            filas.append({
                "Cliente": c["nombre"],
                "Plan": com_c.get("plan_contratado", "—"),
                "Moneda": moneda,
                "Setup único": setup_p,
                "Fijo/mes": fijo,
                "Meses": meses,
                "Reu. gar.": reu_gar,
                "Costo/reu USD": costo_r,
                "Total contrato": total,
                "Inicio prosp.": fp_fmt,
                "Fin contrato": fin_str,
                "Notas": notas,
            })

        if filas:
            import pandas as _pd_inicio
            df_rent = _pd_inicio.DataFrame(filas)
            # Ordenar por total contrato descendente
            df_rent = df_rent.sort_values("Total contrato", ascending=False)
            st.dataframe(df_rent, use_container_width=True, hide_index=True)

            # Métricas rápidas
            r1, r2, r3 = st.columns(3)
            total_cartera = sum(f["Total contrato"] for f in filas)
            fijo_mensual  = sum(f["Fijo/mes"] for f in filas)
            reu_total     = sum(f["Reu. gar."] for f in filas)
            r1.metric("💼 Total cartera", f"${total_cartera:,.0f}")
            r2.metric("📆 Fijo mensual total", f"${fijo_mensual:,.0f}")
            r3.metric("🤝 Reuniones gar. totales", str(reu_total))

            if st.button("📥 Exportar resumen a MD", key="export_resumen_md"):
                try:
                    p = actualizar_resumen_global_clientes()
                    st.success(f"✅ Guardado en: {p}")
                except Exception as e:
                    st.error(str(e))
        else:
            st.info("Aún no hay datos comerciales cargados. Sube propuestas comerciales en cada cliente para poblar esta tabla.")


# ─────────────────────────────────────────────
# NUEVO CLIENTE — onboarding página única
# ─────────────────────────────────────────────

def _wiz_header():
    col_back, col_title = st.columns([1, 8])
    with col_back:
        if st.button("← Inicio", key="wiz_back"):
            for k in ["wiz_step","wiz_nombre","wiz_web","wiz_paises","wiz_objetivo","wiz_archivos","wiz_analisis","wiz_descripcion"]:
                st.session_state[k] = [] if k in ("wiz_paises","wiz_archivos") else ""
            st.session_state.wiz_step = 1
            ir("inicio")
            st.rerun()
    with col_title:
        st.markdown("### ➕ Nuevo cliente")

def _wiz_progress(paso: int):
    pasos_wiz = ["1 · Datos y archivos", "2 · Análisis web", "3 · Confirmar y crear"]
    cols = st.columns(3)
    for i, label in enumerate(pasos_wiz):
        n = i + 1
        if n < paso:
            bg, tc, border = "#d1fae5", "#065f46", "#10b981"
            ico = "✓"
        elif n == paso:
            bg, tc, border = "#dbeafe", "#1e40af", "#3b82f6"
            ico = str(n)
        else:
            bg, tc, border = "#f8fafc", "#94a3b8", "#e2e8f0"
            ico = str(n)
        with cols[i]:
            st.markdown(f"""
            <div style="background:{bg};border:2px solid {border};border-radius:10px;
                        padding:10px;text-align:center;">
              <div style="font-size:18px;font-weight:800;color:{tc};">{ico}</div>
              <div style="font-size:11px;font-weight:600;color:{tc};margin-top:2px;">{label}</div>
            </div>
            """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)


def _folder_badge(carpeta: str):
    """Muestra a qué carpeta va a parar algo."""
    st.markdown(
        f'<span style="font-size:11px;background:#f1f5f9;color:#64748b;'
        f'padding:2px 8px;border-radius:4px;font-family:monospace;">→ {carpeta}</span>',
        unsafe_allow_html=True
    )


def page_nuevo_cliente():
    _wiz_header()
    st.markdown("---")

    # ── Una sola página, no wizard ────────────────────────────────
    st.markdown("""
    <div style="margin-bottom:24px;">
      <h2 style="font-size:22px;font-weight:800;margin:0 0 4px;color:#0f172a;">
        Etapa 1 — Onboarding nuevo cliente
      </h2>
      <p style="font-size:14px;color:#64748b;margin:0;">
        Primero archivos, luego datos. Las carpetas y archivos se crean al final, con información real.
      </p>
    </div>
    """, unsafe_allow_html=True)

    # ─────────────────────────────────────────
    # SECCIÓN 1: ARCHIVOS
    # ─────────────────────────────────────────
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:12px;">
      <div style="background:#3b82f6;color:white;width:28px;height:28px;border-radius:50%;
                  display:flex;align-items:center;justify-content:center;font-weight:800;font-size:14px;flex-shrink:0;">1</div>
      <div>
        <div style="font-size:16px;font-weight:700;color:#0f172a;">Archivos del cliente</div>
        <div style="font-size:12px;color:#64748b;">Sube lo que tengas: docs, logos, bases, presentaciones. Se clasifican automáticamente.</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_up, col_dest = st.columns([3, 2])
    with col_up:
        archivos_up = st.file_uploader(
            "Arrastra archivos aquí",
            accept_multiple_files=True,
            type=["pdf","docx","txt","md","csv","xlsx","png","jpg","jpeg","webp","html","pptx","doc","svg"],
            label_visibility="collapsed",
        )
        if archivos_up:
            for f in archivos_up:
                nombres = [a["nombre"] for a in st.session_state.wiz_archivos]
                if f.name not in nombres:
                    st.session_state.wiz_archivos.append({"nombre": f.name, "bytes": f.getvalue()})

    with col_dest:
        st.markdown("""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:14px 16px;">
          <div style="font-size:11px;font-weight:700;color:#475569;text-transform:uppercase;
                      letter-spacing:0.5px;margin-bottom:10px;">Destino por tipo</div>
          <div style="font-size:12px;color:#374151;line-height:2;">
            🎨 logo*.* → <code>00_INPUT_CLIENTE/logos/</code><br>
            📄 PDF, DOCX → <code>documentos/</code><br>
            📊 CSV, XLSX → <code>bases/</code><br>
            🖼️ PNG, JPG → <code>imagenes/</code><br>
            📝 Minutas → <code>minutas/</code>
          </div>
        </div>
        """, unsafe_allow_html=True)

    # Mostrar archivos cargados
    if st.session_state.wiz_archivos:
        st.markdown(f"<br>**{len(st.session_state.wiz_archivos)} archivo(s) listos:**", unsafe_allow_html=True)
        from modules.archivos import clasificar_archivo
        cat_icons = {"logo":"🎨","documento":"📄","imagen":"🖼️","base":"📊","minuta":"📝","html":"🌐","otro":"📦"}
        cols_f = st.columns(3)
        for i, af in enumerate(st.session_state.wiz_archivos):
            subcarpeta, cat = clasificar_archivo(af["nombre"])
            ico = cat_icons.get(cat, "📦")
            with cols_f[i % 3]:
                st.markdown(
                    f'<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;'
                    f'padding:6px 10px;margin:3px 0;font-size:12px;">'
                    f'{ico} <b>{af["nombre"]}</b><br>'
                    f'<span style="color:#16a34a;font-size:10px;">→ {subcarpeta}</span></div>',
                    unsafe_allow_html=True
                )
        if st.button("🗑 Limpiar archivos", key="limpiar_archivos"):
            st.session_state.wiz_archivos = []
            st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<hr style="border-color:#f1f5f9;">', unsafe_allow_html=True)

    # ─────────────────────────────────────────
    # SECCIÓN 2: DATOS BASE
    # ─────────────────────────────────────────
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;margin-top:8px;">
      <div style="background:#3b82f6;color:white;width:28px;height:28px;border-radius:50%;
                  display:flex;align-items:center;justify-content:center;font-weight:800;font-size:14px;flex-shrink:0;">2</div>
      <div>
        <div style="font-size:16px;font-weight:700;color:#0f172a;">Datos base del cliente</div>
        <div style="font-size:12px;color:#64748b;">Nombre, web y análisis. Cada campo se guarda en su carpeta correspondiente.</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Nombre + Web + Botón analizar ────────────────────────────
    col_n, col_w, col_btn = st.columns([2, 3, 1])
    with col_n:
        nombre = st.text_input(
            "Nombre del cliente *",
            value=st.session_state.wiz_nombre,
            placeholder="Ej: GBS Logistics",
        )
        _folder_badge("01_ADMIN_CLIENTE/datos_cliente.md")
    with col_w:
        web = st.text_input(
            "Sitio web *",
            value=st.session_state.wiz_web,
            placeholder="Ej: https://www.empresa.com",
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")
    with col_btn:
        st.markdown('<div style="margin-top:28px;">', unsafe_allow_html=True)
        analizar_clicked = st.button("🔍 Analizar", use_container_width=True, help="Genera prompt para analizar web + archivos con Claude")
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Panel del botón Analizar ──────────────────────────────────
    if analizar_clicked and web.strip():
        url_analisis = web if web.startswith("http") else f"https://{web}"
        archivos_nombres = [a["nombre"] for a in st.session_state.wiz_archivos]
        archivos_str = "\n".join(f"- {n}" for n in archivos_nombres) if archivos_nombres else "- (ninguno subido aún)"
        prompt_web = f"""Analiza la empresa del sitio web {url_analisis} usando también los archivos adjuntos si los tienes disponibles:
{archivos_str}

Responde en formato estructurado con estas 5 secciones EXACTAS (usa estos títulos):

## RESUMEN DEL SERVICIO
(2-3 líneas: qué hace la empresa y a quién ayuda)

## PROPUESTA DE VALOR PRINCIPAL
(qué resultado concreto obtienen sus clientes)

## PROBLEMA QUE RESUELVE
(dolor o frustración del cliente que esta empresa ataca)

## TIPO DE CLIENTE IDEAL
(tamaño de empresa, industria, cargo que toma la decisión, características clave)

## DIFERENCIACIÓN VS COMPETENCIA
(qué los hace únicos o mejores respecto a alternativas)"""

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown(f"""
        <div style="background:#fffbeb;border:1.5px solid #f59e0b;border-radius:10px;padding:14px 18px;margin-bottom:10px;">
          <b style="color:#92400e;">🔌 Próximamente automático con Claude API</b>
          <span style="font-size:12px;color:#78350f;margin-left:8px;">Por ahora: copia el prompt → pégalo en Claude → copia la respuesta en los campos de abajo.</span>
        </div>
        """, unsafe_allow_html=True)

        col_prompt, col_link = st.columns([4, 1])
        with col_prompt:
            st.text_area("Prompt para Claude:", value=prompt_web, height=200, key="prompt_analisis_gen")
        with col_link:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(f'<a href="{url_analisis}" target="_blank" style="font-size:13px;color:#3b82f6;display:block;margin-top:8px;">🌐 Abrir web ↗</a>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── 5 cards de análisis ───────────────────────────────────────
    st.markdown("""
    <div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:10px;">
      📊 Análisis del cliente
      <span style="font-size:11px;font-weight:400;color:#94a3b8;margin-left:8px;">Completa manualmente o pega la respuesta de Claude sección por sección</span>
    </div>
    """, unsafe_allow_html=True)

    card_col1, card_col2 = st.columns(2)

    with card_col1:
        st.markdown("""<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:10px 14px;margin-bottom:4px;">
          <span style="font-size:12px;font-weight:700;color:#15803d;">📝 Resumen del servicio</span></div>""", unsafe_allow_html=True)
        resumen_servicio = st.text_area(
            "resumen_servicio", label_visibility="collapsed",
            value=st.session_state.wiz_resumen_servicio,
            placeholder="Qué hace la empresa y a quién ayuda (2-3 líneas)",
            height=90, key="card_resumen"
        )
        _folder_badge("03_ANALISIS_CLIENTE/resumen_servicio.md")

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("""<div style="background:#fef3c7;border:1px solid #fde68a;border-radius:8px;padding:10px 14px;margin-bottom:4px;">
          <span style="font-size:12px;font-weight:700;color:#92400e;">⚡ Problema que resuelve</span></div>""", unsafe_allow_html=True)
        problema = st.text_area(
            "problema", label_visibility="collapsed",
            value=st.session_state.wiz_problema,
            placeholder="Dolor o frustración del cliente que esta empresa ataca",
            height=90, key="card_problema"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("""<div style="background:#fdf4ff;border:1px solid #e9d5ff;border-radius:8px;padding:10px 14px;margin-bottom:4px;">
          <span style="font-size:12px;font-weight:700;color:#7e22ce;">🏆 Diferenciación vs competencia</span></div>""", unsafe_allow_html=True)
        diferenciacion = st.text_area(
            "diferenciacion", label_visibility="collapsed",
            value=st.session_state.wiz_diferenciacion,
            placeholder="Qué los hace únicos o mejores respecto a alternativas",
            height=90, key="card_diferenciacion"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

    with card_col2:
        st.markdown("""<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:10px 14px;margin-bottom:4px;">
          <span style="font-size:12px;font-weight:700;color:#1e40af;">💎 Propuesta de valor principal</span></div>""", unsafe_allow_html=True)
        propuesta_valor = st.text_area(
            "propuesta_valor", label_visibility="collapsed",
            value=st.session_state.wiz_propuesta_valor,
            placeholder="Resultado concreto que obtienen sus clientes",
            height=90, key="card_propuesta"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("""<div style="background:#fff1f2;border:1px solid #fecdd3;border-radius:8px;padding:10px 14px;margin-bottom:4px;">
          <span style="font-size:12px;font-weight:700;color:#be123c;">🎯 Tipo de cliente ideal (ICP)</span></div>""", unsafe_allow_html=True)
        icp_tipo_cliente = st.text_area(
            "icp_tipo_cliente", label_visibility="collapsed",
            value=st.session_state.wiz_icp_tipo_cliente,
            placeholder="Tamaño de empresa, industria, cargo decisor, características clave",
            height=90, key="card_icp"
        )
        _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<hr style="border-color:#f1f5f9;">', unsafe_allow_html=True)

    # ── Campos importantes del cliente ───────────────────────────
    st.markdown("""
    <div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:4px;">
      🗂️ Información importante del cliente
      <span style="font-size:11px;font-weight:400;color:#94a3b8;margin-left:8px;">Pídela en el kick-off o completa lo que ya sepas</span>
    </div>
    """, unsafe_allow_html=True)

    ci_col1, ci_col2 = st.columns(2)

    with ci_col1:
        industrias_no = st.text_area(
            "🚫 Industrias a NO prospectar",
            value=st.session_state.wiz_industrias_no_prospectar,
            placeholder="Ej: retail, gobierno, startups sin financiamiento...",
            height=70, key="ci_industrias_no"
        )
        _folder_badge("04_ICP_ESTRATEGIA/exclusiones_icp.md")

        st.markdown("<br>", unsafe_allow_html=True)

        competidores = st.text_area(
            "⚔️ Competidores directos conocidos",
            value=st.session_state.wiz_competidores,
            placeholder="Ej: Empresa A, Empresa B... (los que menciona el cliente)",
            height=70, key="ci_competidores"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

        st.markdown("<br>", unsafe_allow_html=True)

        canales_actuales = st.text_input(
            "📡 Canales de venta actuales del cliente",
            value=st.session_state.wiz_canales_actuales,
            placeholder="Ej: referidos, LinkedIn, eventos, outbound propio...",
            key="ci_canales"
        )
        _folder_badge("01_ADMIN_CLIENTE/datos_cliente.md")

    with ci_col2:
        clientes_actuales = st.text_area(
            "🤝 Clientes actuales (ejemplos)",
            value=st.session_state.wiz_clientes_actuales,
            placeholder="Ej: Empresa X, Empresa Y... (ayuda a definir el ICP)",
            height=70, key="ci_clientes"
        )
        _folder_badge("01_ADMIN_CLIENTE/datos_cliente.md")

        st.markdown("<br>", unsafe_allow_html=True)

        ticket_promedio = st.text_input(
            "💰 Ticket o contrato promedio",
            value=st.session_state.wiz_ticket_promedio,
            placeholder="Ej: USD 2.000/mes · USD 15.000 anual · depende del proyecto",
            key="ci_ticket"
        )
        _folder_badge("01_ADMIN_CLIENTE/datos_cliente.md")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<hr style="border-color:#f1f5f9;">', unsafe_allow_html=True)

    # ── Objetivo comercial + Países ───────────────────────────────
    obj_col, pais_col = st.columns(2)

    with obj_col:
        objetivo = st.selectbox(
            "🎯 Objetivo comercial *",
            options=OBJETIVOS,
            index=OBJETIVOS.index(st.session_state.wiz_objetivo) if st.session_state.wiz_objetivo in OBJETIVOS else 0,
        )
        _folder_badge("01_ADMIN_CLIENTE/datos_cliente.md  ·  14_SUPABASE_METABASE/datos_para_supabase.json")

    with pais_col:
        paises = st.multiselect(
            "🌎 País o países objetivo *",
            options=PAISES_LA,
            default=st.session_state.wiz_paises,
            help="Puedes seleccionar varios"
        )
        _folder_badge("01_ADMIN_CLIENTE/datos_cliente.md")

    # descripcion queda como campo interno (suma de cards)
    descripcion = resumen_servicio

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown('<hr style="border-color:#f1f5f9;">', unsafe_allow_html=True)

    # ─────────────────────────────────────────
    # SECCIÓN 3: CONFIRMAR
    # ─────────────────────────────────────────
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:16px;margin-top:8px;">
      <div style="background:#3b82f6;color:white;width:28px;height:28px;border-radius:50%;
                  display:flex;align-items:center;justify-content:center;font-weight:800;font-size:14px;flex-shrink:0;">3</div>
      <div>
        <div style="font-size:16px;font-weight:700;color:#0f172a;">Crear cliente</div>
        <div style="font-size:12px;color:#64748b;">Se generan las carpetas y archivos base con la información ingresada.</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_resumen, col_btn_crear = st.columns([3, 1])

    with col_resumen:
        items_ok = []
        items_pend = []
        if nombre.strip():
            items_ok.append(f"✅ {nombre.strip()}")
        else:
            items_pend.append("⚠️ Nombre del cliente")
        if web.strip():
            items_ok.append(f"✅ {web.strip()}")
        else:
            items_pend.append("⚠️ Sitio web")
        if paises:
            items_ok.append(f"✅ {', '.join(paises)}")
        else:
            items_pend.append("⚠️ Países")
        analisis_cards = [resumen_servicio, propuesta_valor, problema, icp_tipo_cliente, diferenciacion]
        n_cards = sum(1 for c in analisis_cards if c.strip())
        if n_cards > 0:
            items_ok.append(f"✅ {n_cards}/5 campos de análisis")
        else:
            items_pend.append("⚪ Análisis del cliente (opcional)")
        if st.session_state.wiz_archivos:
            items_ok.append(f"✅ {len(st.session_state.wiz_archivos)} archivo(s)")
        if industrias_no.strip() or clientes_actuales.strip():
            items_ok.append("✅ Info importante del cliente")

        col_ok, col_pend = st.columns(2)
        with col_ok:
            for item in items_ok:
                st.markdown(f'<div style="font-size:13px;padding:2px 0;">{item}</div>', unsafe_allow_html=True)
        with col_pend:
            for item in items_pend:
                st.markdown(f'<div style="font-size:13px;padding:2px 0;color:#92400e;">{item}</div>', unsafe_allow_html=True)

    with col_btn_crear:
        st.markdown("<br>", unsafe_allow_html=True)
        crear = st.button("🚀 Crear cliente", type="primary", use_container_width=True)

    if crear:
        errores = []
        if not nombre.strip(): errores.append("El nombre es obligatorio")
        if not web.strip(): errores.append("El sitio web es obligatorio")
        if not paises: errores.append("Selecciona al menos un país")

        if errores:
            for e in errores:
                st.error(e)
        else:
            nombre_norm = normalizar(nombre.strip())
            dest = CLIENTES_DIR / nombre_norm

            if dest.exists():
                st.error(f"Ya existe un cliente **{nombre_norm}**. Cambia el nombre.")
            else:
                datos = {
                    # Datos básicos
                    "nombre_cliente": nombre.strip(),
                    "nombre_normalizado": nombre_norm,
                    "sitio_web": web.strip(),
                    "pais_objetivo": ", ".join(paises),
                    "objetivo_comercial": objetivo,
                    "objetivo_label": objetivo,
                    # Análisis (5 cards)
                    "resumen_servicio": resumen_servicio.strip(),
                    "propuesta_valor": propuesta_valor.strip(),
                    "problema_que_resuelve": problema.strip(),
                    "icp_tipo_cliente": icp_tipo_cliente.strip(),
                    "diferenciacion": diferenciacion.strip(),
                    "descripcion": resumen_servicio.strip(),
                    # Información importante del cliente
                    "industrias_no_prospectar": industrias_no.strip(),
                    "clientes_actuales": clientes_actuales.strip(),
                    "competidores": competidores.strip(),
                    "ticket_promedio": ticket_promedio.strip(),
                    "canales_actuales": canales_actuales.strip(),
                    # SDR / firma (se completa después)
                    "nombre_prospector": "",
                    "cargo_prospector": "",
                    "correo": "",
                    "telefono": "",
                    "linkedin": "",
                    "color_marca": "#1a56db",
                    "web_conprospeccion": "conprospeccion.com",
                    "notas_internas": "",
                    "moneda": "USD",
                }

                with st.spinner("Creando estructura de carpetas y archivos..."):
                    dest.mkdir(parents=True, exist_ok=True)
                    crear_estructura_carpetas(dest)
                    generar_todos_los_archivos(dest, datos)

                    fecha_hoy = datetime.now().strftime("%Y-%m-%d")

                    # Guardar cards de análisis en sus archivos individuales
                    if resumen_servicio.strip():
                        (dest / "03_ANALISIS_CLIENTE/resumen_servicio.md").write_text(
                            f"# Resumen del Servicio: {nombre.strip()}\n\n**Fecha:** {fecha_hoy}\n\n---\n\n{resumen_servicio.strip()}",
                            encoding="utf-8"
                        )
                    analisis_completo_parts = []
                    for titulo, contenido in [
                        ("Propuesta de Valor Principal", propuesta_valor),
                        ("Problema que Resuelve", problema),
                        ("Diferenciación vs Competencia", diferenciacion),
                    ]:
                        if contenido.strip():
                            analisis_completo_parts.append(f"## {titulo}\n\n{contenido.strip()}")
                    if analisis_completo_parts:
                        (dest / "03_ANALISIS_CLIENTE/analisis_web.md").write_text(
                            f"# Análisis Web: {nombre.strip()}\n\n**Fecha:** {fecha_hoy}\n\n---\n\n" + "\n\n".join(analisis_completo_parts),
                            encoding="utf-8"
                        )
                        actualizar_campo(dest, "estado_analisis", "analisis_listo")

                    # Guardar ICP borrador
                    if icp_tipo_cliente.strip():
                        icp_ruta = dest / "04_ICP_ESTRATEGIA/icp_borrador.md"
                        icp_existente = icp_ruta.read_text(encoding="utf-8") if icp_ruta.exists() else ""
                        icp_ruta.write_text(
                            icp_existente + f"\n\n## Tipo de Cliente Ideal (onboarding)\n\n{icp_tipo_cliente.strip()}",
                            encoding="utf-8"
                        )

                    # Guardar exclusiones ICP
                    if industrias_no.strip():
                        (dest / "04_ICP_ESTRATEGIA/exclusiones_icp.md").write_text(
                            f"# Exclusiones ICP: {nombre.strip()}\n\n**Fecha:** {fecha_hoy}\n\n## Industrias a NO Prospectar\n\n{industrias_no.strip()}",
                            encoding="utf-8"
                        )

                    # Guardar archivos subidos en sus carpetas correctas
                    for af in st.session_state.wiz_archivos:
                        res = guardar_archivo(dest, af["nombre"], af["bytes"])
                    if st.session_state.wiz_archivos:
                        actualizar_campo(dest, "estado_archivos", "archivos_subidos")

                # Limpiar wizard
                campos_limpiar = [
                    "wiz_nombre","wiz_web","wiz_analisis","wiz_descripcion","wiz_objetivo",
                    "wiz_resumen_servicio","wiz_propuesta_valor","wiz_problema",
                    "wiz_icp_tipo_cliente","wiz_diferenciacion",
                    "wiz_industrias_no_prospectar","wiz_clientes_actuales",
                    "wiz_competidores","wiz_ticket_promedio","wiz_canales_actuales",
                ]
                for k in campos_limpiar:
                    st.session_state[k] = ""
                st.session_state.wiz_paises = []
                st.session_state.wiz_archivos = []
                st.session_state.wiz_step = 1

                n_carpetas = sum(1 for _ in dest.rglob("*") if _.is_dir())
                n_archivos = sum(1 for _ in dest.rglob("*") if _.is_file())

                st.success(f"✅ **{nombre.strip()}** creado — {n_carpetas} carpetas, {n_archivos} archivos")
                st.balloons()

                import time; time.sleep(1)
                ir("cliente", nombre_norm)
                st.rerun()



# ─────────────────────────────────────────────
# WORKSPACE CLIENTE
# ─────────────────────────────────────────────
PASOS = [
    {"id": 1, "icono": "✅", "nombre": "Cliente creado",      "desc": "Datos básicos ingresados"},
    {"id": 2, "icono": "📤", "nombre": "Archivos e input",    "desc": "Sube documentos, logos y materiales del cliente"},
    {"id": 3, "icono": "🌐", "nombre": "Análisis web",        "desc": "Analiza el sitio web y extrae información clave"},
    {"id": 4, "icono": "🎯", "nombre": "ICP y estrategia",    "desc": "Define el perfil de cliente ideal y criterios de segmentación"},
    {"id": 5, "icono": "📨", "nombre": "Mensajería",          "desc": "Pitches, aperturas, mensajes por canal"},
    {"id": 6, "icono": "✍️",  "nombre": "Firma de email",     "desc": "Logo, nombre SDR, cargo → firma HTML lista"},
    {"id": 7, "icono": "📋", "nombre": "Playbook SDR",        "desc": "Brief para Codex y segmentos del playbook"},
    {"id": 8, "icono": "🚀", "nombre": "Listo para operar",   "desc": "Cliente en operación"},
]

def render_pasos(paso_actual: int):
    cols = st.columns(len(PASOS))
    for i, paso in enumerate(PASOS):
        with cols[i]:
            if paso["id"] < paso_actual:
                color_bg = "#d1fae5"
                color_text = "#065f46"
                color_border = "#10b981"
                estado_text = "✓"
            elif paso["id"] == paso_actual:
                color_bg = "#dbeafe"
                color_text = "#1e40af"
                color_border = "#3b82f6"
                estado_text = paso["icono"]
            else:
                color_bg = "#f8fafc"
                color_text = "#94a3b8"
                color_border = "#e2e8f0"
                estado_text = paso["icono"]

            st.markdown(f"""
            <div style="background:{color_bg};border:2px solid {color_border};border-radius:10px;
                        padding:10px 8px;text-align:center;margin:2px;">
              <div style="font-size:16px;">{estado_text}</div>
              <div style="font-size:10px;font-weight:700;color:{color_text};margin-top:3px;line-height:1.3;">{paso['nombre']}</div>
            </div>
            """, unsafe_allow_html=True)


def page_cliente(cid: str):
    path = CLIENTES_DIR / cid
    if not path.exists():
        st.error("Cliente no encontrado")
        if st.button("← Inicio"):
            ir("inicio"); st.rerun()
        return

    est = cargar_estado(path)
    nombre = est.get("nombre_cliente", cid)
    web = est.get("sitio_web", "")
    paises = est.get("pais_objetivo", "")
    objetivo = est.get("objetivo_label", "")
    paso_actual = est.get("etapa_actual", 1)

    # ── Header ──────────────────────────────────
    col_back, col_info, col_web = st.columns([1, 5, 3])
    with col_back:
        if st.button("← Inicio", key="back_inicio"):
            ir("inicio"); st.rerun()
    with col_info:
        st.markdown(f"""
        <div>
          <h2 style="margin:0;font-size:22px;font-weight:800;color:#0f172a;">{nombre}</h2>
          <div style="font-size:13px;color:#64748b;margin-top:2px;">
            📍 {paises or '—'} &nbsp;·&nbsp; 🎯 {objetivo or '—'}
          </div>
        </div>
        """, unsafe_allow_html=True)
    with col_web:
        if web:
            url = web if web.startswith("http") else f"https://{web}"
            st.markdown(f'<div style="text-align:right;padding-top:8px;"><a href="{url}" target="_blank" style="color:#3b82f6;font-size:13px;">🌐 {web} ↗</a></div>', unsafe_allow_html=True)

    # ── Barra de progreso compacta ───────────────
    prog = calcular_progreso(est)
    pct = prog["porcentaje"]
    comp = prog["completados"]
    tot  = prog["total"]
    color_prog = "#22c55e" if pct >= 80 else "#3b82f6" if pct >= 40 else "#f59e0b"
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:12px;padding:8px 0 4px;">
      <div style="flex:1;background:#e2e8f0;border-radius:99px;height:8px;">
        <div style="width:{pct}%;background:{color_prog};border-radius:99px;height:8px;transition:width 0.3s;"></div>
      </div>
      <div style="font-size:12px;font-weight:700;color:{color_prog};white-space:nowrap;">{pct}% completado</div>
      <div style="font-size:11px;color:#94a3b8;white-space:nowrap;">{comp}/{tot} módulos</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Tabs ────────────────────────────────────
    tabs = st.tabs([
        "🏢 Datos del cliente",
        "📤 Archivos",
        "🌐 Análisis web",
        "🎯 ICP",
        "📊 Bases y Mensajería",
        "📋 Playbook SDR",
        "✍️ Firma",
        "🔗 GHL Setup",
        "🗄️ Estado y datos",
        "💬 Chat",
        "📄 Archivos generados",
        "🗺️ Roadmap",
    ])

    with tabs[0]: tab_datos_cliente(path, est)
    with tabs[1]: tab_archivos(path, est)
    with tabs[2]: tab_analisis(path, est)
    with tabs[3]: tab_icp(path, est)
    with tabs[4]: tab_bases_apollo(path, est)
    with tabs[5]: tab_playbook(path, est)
    with tabs[6]: tab_firma(path, est)
    with tabs[7]: tab_ghl(path, est)
    with tabs[8]: tab_estado_datos(path, est)
    with tabs[9]: tab_chat(path, est)
    with tabs[10]: tab_archivos_generados(path, est)
    with tabs[11]: tab_roadmap(path, est)


# ─────────────────────────────────────────────
# TAB: DATOS DEL CLIENTE
# ─────────────────────────────────────────────
def tab_datos_cliente(path: Path, est: dict):
    import datetime as dt

    st.markdown("#### 🏢 Datos del cliente")
    _folder_badge("estado_cliente.json  ·  07_BASE_DATOS/comercial.json")

    nombre = est.get("nombre_cliente", "")

    # ── Sección 1: Información del cliente ────────────────────────
    st.markdown("##### Información del cliente")
    d1c1, d1c2, d1c3 = st.columns(3)
    with d1c1:
        v_nombre = st.text_input("Nombre del cliente", value=nombre, key="dc_nombre")
    with d1c2:
        v_web = st.text_input("Página web", value=est.get("sitio_web", ""), key="dc_web",
                              placeholder="gbslogistics.cl")
    with d1c3:
        v_paises = st.text_input("Países a prospectar", value=est.get("pais_objetivo", ""),
                                 key="dc_paises", placeholder="Chile, Perú, Colombia")

    d2c1, d2c2 = st.columns([2, 1])
    with d2c1:
        objetivo_opts = [""] + OBJETIVOS
        obj_actual = est.get("objetivo_comercial", "")
        try:
            obj_idx = objetivo_opts.index(obj_actual) if obj_actual in objetivo_opts else 0
        except:
            obj_idx = 0
        v_objetivo = st.selectbox("Objetivo comercial", options=objetivo_opts,
                                  index=obj_idx, key="dc_objetivo")
    with d2c2:
        v_industria = st.text_input("Industria / Rubro", value=est.get("industria", ""),
                                    key="dc_industria", placeholder="Logística, SaaS, Manufactura…")

    # ── Sección 2: Datos del prospector (SDR) ────────────────────
    st.markdown("---")
    st.markdown("##### Prospector / SDR asignado")
    p1, p2, p3 = st.columns(3)
    with p1:
        v_prosp_nombre = st.text_input("Nombre completo", value=est.get("nombre_prospector", "").strip(),
                                       key="dc_prosp_nombre", placeholder="Francisca Polanco")
    with p2:
        v_prosp_cargo = st.text_input("Cargo", value=est.get("cargo_prospector", "").strip(),
                                      key="dc_prosp_cargo", placeholder="Growth Manager")
    with p3:
        v_prosp_correo = st.text_input("Correo", value=est.get("correo", ""),
                                       key="dc_prosp_correo", placeholder="fran@conprospeccion.com")

    p4, p5, p6 = st.columns(3)
    with p4:
        v_prosp_tel = st.text_input("Teléfono", value=est.get("telefono", ""),
                                    key="dc_prosp_tel", placeholder="+56 9 1234 5678")
    with p5:
        v_prosp_linkedin = st.text_input("LinkedIn", value=est.get("linkedin", ""),
                                         key="dc_prosp_linkedin", placeholder="linkedin.com/in/fran")
    with p6:
        v_prosp_dir = st.text_input("Ciudad / Dirección (firma)", value=est.get("direccion_firma", ""),
                                    key="dc_prosp_dir", placeholder="Santiago, Chile")

    # ── Sección 3: Comercial básico ───────────────────────────────
    st.markdown("---")
    st.markdown("##### Contrato y comercial")

    # Leer comercial.json si existe
    com_file = path / "07_BASE_DATOS" / "comercial.json"
    com = {}
    if com_file.exists():
        try: com = json.loads(com_file.read_text(encoding="utf-8"))
        except: com = {}

    # ── Subir propuesta comercial ─────────────────────────────────
    prop_dir = path / "07_BASE_DATOS"
    prop_dir.mkdir(parents=True, exist_ok=True)

    # Buscar propuesta ya guardada
    _prop_existente = next((f for f in prop_dir.iterdir()
                            if f.stem.startswith("propuesta_comercial") and
                            f.suffix.lower() in {".pdf",".docx",".doc",".txt"}), None) \
                      if prop_dir.exists() else None

    with st.expander(
        f"📄 Propuesta comercial {'✅ ' + _prop_existente.name if _prop_existente else '— sube el documento aceptado'}",
        expanded=not _prop_existente
    ):
        prop_up = st.file_uploader(
            "Sube la propuesta comercial aceptada (PDF, DOCX, TXT)",
            type=["pdf", "docx", "doc", "txt"],
            key=f"dc_propuesta_{path.name}",
        )
        if prop_up:
            ext_prop = Path(prop_up.name).suffix.lower()
            dest_prop = prop_dir / f"propuesta_comercial_original{ext_prop}"
            dest_prop.write_bytes(prop_up.read())
            _prop_existente = dest_prop
            st.success(f"✅ Guardada: {dest_prop.name}")

        if _prop_existente and st.button("🤖 Analizar propuesta y auto-rellenar campos",
                                          key=f"dc_analizar_prop_{path.name}",
                                          use_container_width=True):
            with st.spinner("Leyendo propuesta con Claude..."):
                client = get_claude_client()
                _texto_prop = leer_documento(_prop_existente, max_chars=7000)
                _nombre_cl  = est.get("nombre_cliente", path.name)
                _datos_ext  = extraer_comercial_de_propuesta(client, _texto_prop, _nombre_cl)

            if _datos_ext:
                # Merge con com actual y guardar
                com_nuevo = dict(com)
                for k in ["moneda","monto_setup","monto_fijo","meses_contrato","semanas_setup",
                           "reuniones_garantizadas","costo_reunion"]:
                    if _datos_ext.get(k) is not None:
                        com_nuevo[k] = _datos_ext[k]
                # Guardar notas si vienen
                if _datos_ext.get("notas_comerciales"):
                    com_nuevo["notas_comerciales"] = _datos_ext["notas_comerciales"]
                com_file.write_text(json.dumps(com_nuevo, ensure_ascii=False, indent=2), encoding="utf-8")

                # Guardar resumen en MD del cliente
                md_com = prop_dir / "resumen_comercial.md"
                _fp = com_nuevo.get("fecha_inicio_prospeccion","—")
                md_com.write_text(
                    f"# Resumen comercial — {est.get('nombre_cliente', path.name)}\n\n"
                    f"| Campo | Valor |\n|---|---|\n"
                    f"| Moneda | {com_nuevo.get('moneda','—')} |\n"
                    f"| Pago de setup (único) | {com_nuevo.get('monto_setup',0):,} |\n"
                    f"| Fijo mensual | {com_nuevo.get('monto_fijo',0):,} |\n"
                    f"| Meses contrato | {com_nuevo.get('meses_contrato',5)} |\n"
                    f"| Semanas setup | {com_nuevo.get('semanas_setup',3)} |\n"
                    f"| Reuniones garantizadas | {com_nuevo.get('reuniones_garantizadas',0)} |\n"
                    f"| Costo por reunión (USD) | {com_nuevo.get('costo_reunion',0)} |\n"
                    f"| Inicio prospección activa | {_fp} |\n"
                    f"| Total contrato | {com_nuevo.get('monto_setup',0) + com_nuevo.get('monto_fijo',0)*com_nuevo.get('meses_contrato',5) + com_nuevo.get('reuniones_garantizadas',0)*com_nuevo.get('costo_reunion',0):,} |\n\n"
                    f"**Notas:** {com_nuevo.get('notas_comerciales','—')}\n",
                    encoding="utf-8",
                )

                # Actualizar resumen global
                try: actualizar_resumen_global_clientes()
                except: pass

                st.success("✅ Campos actualizados desde la propuesta")
                st.json(_datos_ext)
                com = com_nuevo  # Usar datos nuevos en el resto del tab
                st.rerun()
            else:
                st.warning("No se pudieron extraer datos. Revisa que la propuesta tenga montos y condiciones claras.")

        if _prop_existente:
            # Mostrar notas si existen
            _notas = com.get("notas_comerciales", "")
            if _notas:
                st.markdown(f'<div style="background:#f8fafc;border-radius:6px;padding:8px 12px;font-size:12px;color:#475569;margin-top:4px;">📝 <b>Notas:</b> {_notas}</div>', unsafe_allow_html=True)

    _planes = ["— Sin definir —", "Starter", "Pro"]
    _plan_guardado = com.get("plan_contratado", "— Sin definir —")
    _plan_idx = _planes.index(_plan_guardado) if _plan_guardado in _planes else 0
    v_plan = st.selectbox("📦 Plan contratado", _planes, index=_plan_idx, key="dc_plan")

    def _fmt(n, moneda):
        """Formatea un número con separador de miles según moneda."""
        if moneda == "CLP":
            return f"$ {int(n):,.0f}".replace(",", ".")
        else:
            return f"USD {n:,.2f}"

    cb1, cb2, cb3, cb4, cb4b = st.columns(5)
    with cb1:
        v_moneda = st.selectbox("Moneda", ["CLP", "USD"],
                                index=0 if com.get("moneda", "CLP") == "CLP" else 1,
                                key="dc_moneda")
    with cb2:
        paso = 10_000 if v_moneda == "CLP" else 100
        v_monto = st.number_input(f"Fijo mensual", min_value=0, step=paso,
                                  value=int(com.get("monto_fijo", 0)), key="dc_monto")
        if v_monto > 0:
            st.markdown(f'<div style="font-size:13px;font-weight:700;color:#0f172a;margin-top:-8px;">{_fmt(v_monto, v_moneda)}</div>', unsafe_allow_html=True)
    with cb3:
        v_setup_pago = st.number_input(f"Pago de setup", min_value=0, step=paso,
                                       value=int(com.get("monto_setup", 0)), key="dc_setup_pago",
                                       help="Cobro único al inicio del contrato por onboarding/setup")
        if v_setup_pago > 0:
            st.markdown(f'<div style="font-size:13px;font-weight:700;color:#0f172a;margin-top:-8px;">{_fmt(v_setup_pago, v_moneda)}</div>', unsafe_allow_html=True)
    with cb4:
        v_meses = st.number_input("Meses de contrato", min_value=1, max_value=24,
                                  value=int(com.get("meses_contrato", 5)), key="dc_meses")
    with cb4b:
        v_setup = st.number_input("Semanas de setup", min_value=1, max_value=8,
                                  value=int(com.get("semanas_setup", 3)), key="dc_setup")

    # ── Selector de qué fecha se conoce ───────────────────────────
    st.markdown('<div style="font-size:12px;color:#64748b;margin-top:4px;">📅 ¿Qué fecha conoces?</div>', unsafe_allow_html=True)
    modo_fecha_opts = ["Fecha de inicio del contrato", "Fecha de inicio de prospección activa"]
    modo_fecha = st.radio(
        "¿Qué fecha conoces?",
        modo_fecha_opts,
        horizontal=True,
        label_visibility="collapsed",
        key="dc_modo_fecha",
    )

    # Leer fechas guardadas como punto de partida
    _fc_str = com.get("fecha_inicio_contrato", dt.date.today().isoformat())
    _fp_str = com.get("fecha_inicio_prospeccion", "")
    try: _fc_def = dt.date.fromisoformat(_fc_str)
    except: _fc_def = dt.date.today()
    try: _fp_def = dt.date.fromisoformat(_fp_str) if _fp_str else (_fc_def + dt.timedelta(weeks=int(com.get("semanas_setup", 3))))
    except: _fp_def = _fc_def + dt.timedelta(weeks=3)

    cb5, cb6, cb7 = st.columns(3)
    with cb5:
        if modo_fecha == "Fecha de inicio del contrato":
            v_fecha = st.date_input("📋 Inicio contrato", value=_fc_def, key="dc_fecha")
            inicio_prosp = v_fecha + dt.timedelta(weeks=int(v_setup))
        else:
            v_fecha_prosp = st.date_input("🚀 Inicio prospección activa", value=_fp_def, key="dc_fecha_prosp")
            inicio_prosp = v_fecha_prosp
            v_fecha = v_fecha_prosp - dt.timedelta(weeks=int(v_setup))
    with cb6:
        v_reu_gar = st.number_input("Reuniones garantizadas (total)",
                                    min_value=0, value=int(com.get("reuniones_garantizadas", 0)),
                                    key="dc_reu_gar")
    with cb7:
        v_costo_reu = st.number_input("Costo por reunión (USD)",
                                      min_value=0, step=5,
                                      value=int(com.get("costo_reunion", 30)),
                                      key="dc_costo_reu")
        if v_costo_reu > 0:
            st.markdown(f'<div style="font-size:13px;font-weight:700;color:#0f172a;margin-top:-8px;">USD {v_costo_reu:,}</div>', unsafe_allow_html=True)

    # Calcular fin de contrato (siempre desde inicio prospección + meses)
    fin_ctto = inicio_prosp + dt.timedelta(days=int(v_meses * 30.4))

    # ── Línea de tiempo destacada (siempre visible) ────────────────
    col_fa, col_fb, col_fc = st.columns(3)
    with col_fa:
        st.markdown(
            f"""<div style="background:#f8fafc;border:1.5px solid #cbd5e1;border-radius:8px;
                padding:8px 14px;margin-top:8px;font-size:12px;color:#475569;text-align:center;">
              📋 <b>Inicio contrato</b><br>
              <span style="font-size:15px;font-weight:700;color:#334155;">{v_fecha.strftime('%d/%m/%Y')}</span>
            </div>""",
            unsafe_allow_html=True,
        )
    with col_fb:
        st.markdown(
            f"""<div style="background:#f0fdf4;border:1.5px solid #86efac;border-radius:8px;
                padding:8px 14px;margin-top:8px;font-size:12px;color:#166534;text-align:center;">
              🚀 <b>Inicio prospección activa</b><br>
              <span style="font-size:15px;font-weight:700;">{inicio_prosp.strftime('%d/%m/%Y')}</span>
              <span style="font-size:10px;display:block;opacity:.7;">+{int(v_setup)} sem. setup</span>
            </div>""",
            unsafe_allow_html=True,
        )
    with col_fc:
        st.markdown(
            f"""<div style="background:#fef3c7;border:1.5px solid #fcd34d;border-radius:8px;
                padding:8px 14px;margin-top:8px;font-size:12px;color:#92400e;text-align:center;">
              🏁 <b>Fin contrato</b><br>
              <span style="font-size:15px;font-weight:700;">{fin_ctto.strftime('%d/%m/%Y')}</span>
              <span style="font-size:10px;display:block;opacity:.7;">+{int(v_meses)} meses</span>
            </div>""",
            unsafe_allow_html=True,
        )

    # Resumen financiero (solo si tiene algún monto)
    if v_monto > 0 or v_reu_gar > 0 or v_setup_pago > 0:
        ingreso_fijo = v_monto * v_meses
        ingreso_var  = v_reu_gar * v_costo_reu
        total_ctto   = v_setup_pago + ingreso_fijo + ingreso_var
        partes = []
        if v_setup_pago > 0:
            partes.append(f"<span>🔧 <b>Setup (único):</b> {v_moneda} {v_setup_pago:,.0f}</span>")
        if v_monto > 0:
            partes.append(f"<span>📆 <b>Fijo mensual:</b> {v_moneda} {v_monto:,.0f} × {v_meses} m.</span>")
        if v_reu_gar > 0:
            partes.append(f"<span>🤝 <b>Reu. gar.:</b> {v_reu_gar} × {v_costo_reu} USD</span>")
        partes.append(f"<span style='font-weight:800;font-size:13px;'>💰 <b>TOTAL: {v_moneda} {total_ctto:,.0f}</b></span>")
        st.markdown(f"""
        <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
                    padding:10px 16px;margin-top:6px;font-size:12px;color:#0369a1;display:flex;gap:20px;flex-wrap:wrap;align-items:center;">
          {''.join(partes)}
        </div>""", unsafe_allow_html=True)

    # ── Guardar (manual + auto-save al cambiar) ───────────────────
    st.markdown("<br>", unsafe_allow_html=True)

    def _guardar_datos():
        campos_est = {
            "nombre_cliente": st.session_state.get("dc_nombre", nombre).strip(),
            "sitio_web": st.session_state.get("dc_web", "").strip(),
            "pais_objetivo": st.session_state.get("dc_paises", "").strip(),
            "objetivo_comercial": st.session_state.get("dc_objetivo", ""),
            "industria": st.session_state.get("dc_industria", "").strip(),
            "nombre_prospector": st.session_state.get("dc_prosp_nombre", "").strip(),
            "cargo_prospector": st.session_state.get("dc_prosp_cargo", "").strip(),
            "correo": st.session_state.get("dc_prosp_correo", "").strip(),
            "telefono": st.session_state.get("dc_prosp_tel", "").strip(),
            "linkedin": st.session_state.get("dc_prosp_linkedin", "").strip(),
            "direccion_firma": st.session_state.get("dc_prosp_dir", "").strip(),
        }
        for campo, valor in campos_est.items():
            if valor:
                actualizar_campo(path, campo, valor)
        (path / "07_BASE_DATOS").mkdir(parents=True, exist_ok=True)
        com_nuevo = dict(com)
        com_nuevo.update({
            "plan_contratado": st.session_state.get("dc_plan", "— Sin definir —"),
            "moneda": st.session_state.get("dc_moneda", "CLP"),
            "monto_fijo": st.session_state.get("dc_monto", 0),
            "monto_setup": st.session_state.get("dc_setup_pago", 0),
            "meses_contrato": st.session_state.get("dc_meses", 5),
            "semanas_setup": st.session_state.get("dc_setup", 3),
            # v_fecha e inicio_prosp se calculan correctamente sea cual sea el modo
            "fecha_inicio_contrato": v_fecha.isoformat(),
            "fecha_inicio_prospeccion": inicio_prosp.isoformat(),
            "reuniones_garantizadas": st.session_state.get("dc_reu_gar", 0),
            "costo_reunion": st.session_state.get("dc_costo_reu", 30),
        })
        com_file.write_text(json.dumps(com_nuevo, ensure_ascii=False, indent=2), encoding="utf-8")
        _sync_supabase_cliente(path)

    gc1, gc2 = st.columns([2, 3])
    with gc1:
        if st.button("💾 Guardar y continuar →", type="primary", use_container_width=True, key="dc_guardar"):
            _guardar_datos()
            st.toast("✅ Datos guardados", icon="💾")
            ir_a_tab(1)   # Avanza a la pestaña "Archivos"
    with gc2:
        st.markdown('<div style="font-size:12px;color:#94a3b8;padding-top:12px;">💡 Los campos también se guardan automáticamente al hacer clic fuera de ellos</div>', unsafe_allow_html=True)

    # Auto-save al cambiar cualquier campo — comparamos como string para manejar int/float/date
    _autosave_keys = [
        "dc_nombre","dc_web","dc_paises","dc_objetivo","dc_industria",
        "dc_prosp_nombre","dc_prosp_cargo","dc_prosp_correo",
        "dc_prosp_tel","dc_prosp_linkedin","dc_prosp_dir",
        "dc_plan","dc_moneda","dc_monto","dc_setup_pago",
        "dc_meses","dc_setup","dc_fecha","dc_fecha_prosp","dc_reu_gar","dc_costo_reu",
    ]
    _did_autosave = False
    for k in _autosave_keys:
        _val_actual = str(st.session_state.get(k, ""))
        _prev_key   = f"_prev_{k}"
        if _prev_key not in st.session_state:
            st.session_state[_prev_key] = _val_actual
        elif _val_actual != st.session_state[_prev_key]:
            st.session_state[_prev_key] = _val_actual
            if not _did_autosave:          # guardar una sola vez aunque cambien varios
                _guardar_datos()
                _did_autosave = True
    if _did_autosave:
        st.toast("💾 Guardado automáticamente", icon="✅")


# ─────────────────────────────────────────────
# TAB: ARCHIVOS
# ─────────────────────────────────────────────
def tab_archivos(path: Path, est: dict):
    nombre = est.get("nombre_cliente", "")

    st.markdown("#### 📤 Archivos del cliente")
    _folder_badge("00_INPUT_CLIENTE/")
    st.markdown("Sube documentos, presentaciones, bases de datos y logos. Se ordenan automáticamente en subcarpetas.")

    col1, col2 = st.columns([3, 2])
    with col1:
        archivos = st.file_uploader(
            "📎 Selecciona o arrastra archivos aquí",
            accept_multiple_files=True,
            type=["pdf", "docx", "txt", "md", "csv", "xlsx", "png", "jpg", "jpeg", "webp", "html", "doc", "pptx"],
        )
        if archivos:
            if st.button("⬆️ Subir archivos", type="primary", use_container_width=True, key="btn_subir_archivos"):
                resultados = []
                for f in archivos:
                    res = guardar_archivo(path, f.name, f.getvalue())
                    resultados.append(res)
                categorias_icons = {"logo": "🎨", "documento": "📄", "imagen": "🖼️", "base": "📊", "minuta": "📝", "html": "🌐", "otro": "📦"}
                st.markdown("---")
                c1, c2 = st.columns(2)
                for i, r in enumerate(resultados):
                    ico = categorias_icons.get(r["categoria"], "📦")
                    with (c1 if i % 2 == 0 else c2):
                        st.markdown(f"""
                        <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:8px 12px;margin:4px 0;">
                          {ico} <b>{r['archivo']}</b><br>
                          <span style="font-size:11px;color:#16a34a;font-family:monospace;">→ {r['subcarpeta']}</span>
                        </div>
                        """, unsafe_allow_html=True)
                actualizar_campo(path, "estado_archivos", "archivos_subidos")
                logos = listar_logos(path)
                if logos:
                    st.toast(f"🎨 {len(logos)} logo(s) detectado(s)", icon="🎨")
                st.toast(f"✅ {len(resultados)} archivo(s) guardados", icon="📤")
                ir_a_tab(2)   # Avanza a Análisis web
    with col2:
        st.markdown("""
        <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:14px 16px;font-size:12px;color:#475569;">
          <b>Clasificación automática:</b><br><br>
          🎨 <code>logo*.*</code> → logos/<br>
          📄 PDF, DOCX → documentos/<br>
          📊 CSV, XLSX → bases/<br>
          🖼️ Imágenes → imagenes/<br>
          📝 Minutas → minutas/<br>
          📦 Otros → otros/
        </div>
        """, unsafe_allow_html=True)

    # Mostrar archivos existentes
    st.markdown("---")
    st.markdown("#### Archivos en el sistema")

    archivos_por_carpeta = listar_archivos_por_carpeta(path)
    carpetas_input = {k: v for k, v in archivos_por_carpeta.items() if "00_INPUT" in k and v}

    if not carpetas_input:
        st.markdown("""
        <div class="placeholder-box">
          <div style="font-size:32px;margin-bottom:8px;">📁</div>
          <div style="font-weight:600;margin-bottom:4px;">Sin archivos subidos aún</div>
          <div style="font-size:13px;">Sube archivos del cliente arriba para empezar</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        total = sum(len(v) for v in carpetas_input.values())
        st.caption(f"{total} archivos subidos")
        for carpeta, archivos_list in carpetas_input.items():
            carpeta_corta = carpeta.replace("00_INPUT_CLIENTE/", "")
            with st.expander(f"📂 {carpeta_corta} ({len(archivos_list)})", expanded=True):
                for a in archivos_list:
                    icons = {".pdf": "📕", ".docx": "📘", ".xlsx": "📗", ".png": "🖼️", ".jpg": "🖼️", ".jpeg": "🖼️", ".csv": "📊", ".md": "📝"}
                    ico = icons.get(a["extension"], "📄")
                    st.markdown(f'<span class="file-tag">{ico} {a["nombre"]}</span>', unsafe_allow_html=True)


# ─────────────────────────────────────────────
# TAB: ANÁLISIS WEB
# ─────────────────────────────────────────────
def _leer_archivo(path: Path, ruta_rel: str, fallback: str = "") -> str:
    p = path / ruta_rel
    if p.exists():
        try:
            return p.read_text(encoding="utf-8")
        except Exception:
            return fallback
    return fallback


def _card_analisis(titulo: str, icono: str, color_bg: str, color_borde: str, color_text: str,
                   valor: str, placeholder: str, key: str) -> str:
    st.markdown(f"""
    <div style="background:{color_bg};border:1.5px solid {color_borde};border-radius:10px;
                padding:12px 16px;margin-bottom:6px;">
      <span style="font-size:12px;font-weight:700;color:{color_text};">{icono} {titulo}</span>
    </div>
    """, unsafe_allow_html=True)
    return st.text_area(
        titulo, label_visibility="collapsed",
        value=valor, placeholder=placeholder,
        height=110, key=key
    )


def tab_analisis(path: Path, est: dict):
    nombre = est.get("nombre_cliente", "")
    web = est.get("sitio_web", "")
    url = web if web.startswith("http") else f"https://{web}"

    st.markdown("#### 🌐 Análisis del cliente")
    _folder_badge("03_ANALISIS_CLIENTE/  ·  04_ICP_ESTRATEGIA/icp_borrador.md")

    # ── Estado del análisis ───────────────────────────────────────
    campos_analisis = ["resumen_servicio","propuesta_valor","problema_que_resuelve","icp_tipo_cliente","diferenciacion"]
    n_completos = sum(1 for c in campos_analisis if est.get(c, "").strip())

    if n_completos == 5:
        st.success(f"✅ Análisis completo — todos los campos completados")
    elif n_completos > 0:
        st.warning(f"⚡ Análisis parcial — {n_completos}/5 campos completados")
    else:
        st.info("⏳ Análisis pendiente — completa los campos abajo o usa el botón Analizar")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Botón Analizar con Claude ─────────────────────────────────
    client = get_claude_client()
    ba1, ba2 = st.columns([2, 3])
    with ba1:
        analizar = st.button("🤖 Analizar con Claude", type="primary",
                             use_container_width=True, key="btn_analizar_web",
                             disabled=(not client))
        if not client:
            st.caption("Configura ANTHROPIC_API_KEY en .env")
        if web:
            st.markdown(f'<a href="{url}" target="_blank" style="font-size:12px;color:#3b82f6;">🌐 Abrir {web} ↗</a>', unsafe_allow_html=True)
    with ba2:
        st.markdown("""
        <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
                    padding:10px 14px;font-size:12px;color:#0369a1;">
          Claude analiza la web del cliente y los documentos subidos para completar los 5 campos automáticamente.
          Puedes editar el resultado antes de guardar.
        </div>""", unsafe_allow_html=True)

    if analizar and client:
        with st.spinner(f"Analizando {web}..."):
            # Intentar leer HTML de la web
            web_content = ""
            try:
                import requests as req
                r = req.get(url, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
                if r.status_code == 200:
                    raw_html = r.text[:15000]
                    # Limpiar HTML básico
                    import re as re2
                    web_content = re2.sub(r'<[^>]+>', ' ', raw_html)
                    web_content = re2.sub(r'\s+', ' ', web_content).strip()[:8000]
            except Exception:
                web_content = ""

            # Leer docs subidos (PDF, DOCX, TXT, MD, etc.)
            docs_content = leer_todos_los_documentos(path, max_por_doc=3000)

            system_analisis = f"""Eres un experto en análisis de empresas B2B para prospección outbound.
Analiza la empresa {nombre} ({web}) y devuelve EXACTAMENTE este JSON (sin markdown):
{{
  "resumen_servicio": "2-3 líneas: qué hace la empresa y a quién ayuda",
  "propuesta_valor": "resultado concreto que obtienen sus clientes",
  "problema_que_resuelve": "dolor o frustración del cliente que esta empresa ataca",
  "icp_tipo_cliente": "tamaño de empresa, industria, cargo decisor y características clave",
  "diferenciacion": "qué los hace únicos o mejores respecto a alternativas"
}}
Sé específico y accionable. Usa la información del sitio web y documentos provistos."""

            user_msg = f"Empresa: {nombre}\nWeb: {url}\n"
            if web_content:
                user_msg += f"\nContenido extraído del sitio web:\n{web_content}"
            if docs_content:
                user_msg += f"\nDocumentos del cliente:{docs_content}"
            if not web_content and not docs_content:
                user_msg += "\n(No se pudo acceder al sitio web. Usa tu conocimiento general sobre esta empresa y sus servicios típicos del rubro.)"

            raw = llamar_claude(client, [{"role": "user", "content": user_msg}], system_analisis, max_tokens=1500)
            try:
                # Extracción robusta de JSON
                txt = raw.strip()
                # Intentar extraer entre llaves
                start = txt.find('{')
                end = txt.rfind('}')
                if start >= 0 and end > start:
                    txt = txt[start:end+1]
                resultado = json.loads(txt)
                campos_ok = 0
                for campo, val in resultado.items():
                    if val and str(val).strip():
                        actualizar_campo(path, campo, str(val).strip())
                        campos_ok += 1
                actualizar_campo(path, "estado_analisis", "analisis_listo")
                st.success(f"✅ Análisis completado — {campos_ok} campos completados. Puedes editarlos abajo.")
                st.rerun()
            except Exception as e:
                st.error(f"Error parseando respuesta de Claude: {e}")
                st.code(raw[:800])

    st.markdown("<br>", unsafe_allow_html=True)

    # ── 5 cards editables ─────────────────────────────────────────
    st.markdown('<div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:12px;">📊 Resultados del análisis</div>', unsafe_allow_html=True)

    col_izq, col_der = st.columns(2)

    with col_izq:
        v_resumen = _card_analisis(
            "Resumen del servicio", "📝", "#f0fdf4", "#bbf7d0", "#15803d",
            est.get("resumen_servicio", ""),
            "Qué hace la empresa y a quién ayuda (2-3 líneas)",
            "an_resumen"
        )
        _folder_badge("03_ANALISIS_CLIENTE/resumen_servicio.md")

        st.markdown("<br>", unsafe_allow_html=True)

        v_problema = _card_analisis(
            "Problema que resuelve", "⚡", "#fef3c7", "#fde68a", "#92400e",
            est.get("problema_que_resuelve", ""),
            "Dolor o frustración del cliente que esta empresa ataca",
            "an_problema"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

        st.markdown("<br>", unsafe_allow_html=True)

        v_diferenciacion = _card_analisis(
            "Diferenciación vs competencia", "🏆", "#fdf4ff", "#e9d5ff", "#7e22ce",
            est.get("diferenciacion", ""),
            "Qué los hace únicos o mejores respecto a alternativas",
            "an_diferenciacion"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

    with col_der:
        v_propuesta = _card_analisis(
            "Propuesta de valor principal", "💎", "#eff6ff", "#bfdbfe", "#1e40af",
            est.get("propuesta_valor", ""),
            "Resultado concreto que obtienen sus clientes",
            "an_propuesta"
        )
        _folder_badge("03_ANALISIS_CLIENTE/analisis_web.md")

        st.markdown("<br>", unsafe_allow_html=True)

        v_icp = _card_analisis(
            "Tipo de cliente ideal (ICP)", "🎯", "#fff1f2", "#fecdd3", "#be123c",
            est.get("icp_tipo_cliente", ""),
            "Tamaño de empresa, industria, cargo decisor, características clave",
            "an_icp"
        )
        _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md")

    st.markdown("<br>", unsafe_allow_html=True)

    col_save, col_dl = st.columns([2, 1])
    with col_save:
        if st.button("💾 Guardar análisis", type="primary", use_container_width=True):
            fecha = datetime.now().strftime("%Y-%m-%d")

            # Guardar en estado_cliente.json
            actualizar_campo(path, "resumen_servicio", v_resumen.strip())
            actualizar_campo(path, "propuesta_valor", v_propuesta.strip())
            actualizar_campo(path, "problema_que_resuelve", v_problema.strip())
            actualizar_campo(path, "icp_tipo_cliente", v_icp.strip())
            actualizar_campo(path, "diferenciacion", v_diferenciacion.strip())

            # Guardar archivos individuales
            if v_resumen.strip():
                (path / "03_ANALISIS_CLIENTE/resumen_servicio.md").write_text(
                    f"# Resumen del Servicio: {nombre}\n\n**Fecha:** {fecha}\n\n---\n\n{v_resumen.strip()}",
                    encoding="utf-8"
                )
            partes = []
            for titulo, val in [
                ("Propuesta de Valor Principal", v_propuesta),
                ("Problema que Resuelve", v_problema),
                ("Diferenciación vs Competencia", v_diferenciacion),
            ]:
                if val.strip():
                    partes.append(f"## {titulo}\n\n{val.strip()}")
            if partes:
                (path / "03_ANALISIS_CLIENTE/analisis_web.md").write_text(
                    f"# Análisis Web: {nombre}\n\n**Fecha:** {fecha}\n\n---\n\n" + "\n\n".join(partes),
                    encoding="utf-8"
                )
            if v_icp.strip():
                icp_ruta = path / "04_ICP_ESTRATEGIA/icp_borrador.md"
                icp_existente = icp_ruta.read_text(encoding="utf-8") if icp_ruta.exists() else f"# ICP Borrador: {nombre}\n\n"
                if "## Tipo de Cliente Ideal" not in icp_existente:
                    icp_ruta.write_text(icp_existente + f"\n\n## Tipo de Cliente Ideal\n\n{v_icp.strip()}", encoding="utf-8")

            n_nuevos = sum(1 for v in [v_resumen, v_propuesta, v_problema, v_icp, v_diferenciacion] if v.strip())
            actualizar_campo(path, "estado_analisis", "analisis_listo" if n_nuevos > 0 else "pendiente")
            st.toast(f"✅ Análisis guardado — {n_nuevos}/5 campos", icon="💾")
            ir_a_tab(3)   # Avanza a ICP

    with col_dl:
        resumen_export = "\n\n".join([
            f"## Resumen del Servicio\n{v_resumen}",
            f"## Propuesta de Valor\n{v_propuesta}",
            f"## Problema que Resuelve\n{v_problema}",
            f"## Tipo de Cliente Ideal\n{v_icp}",
            f"## Diferenciación\n{v_diferenciacion}",
        ])
        st.download_button(
            "⬇️ Exportar análisis",
            data=resumen_export.encode("utf-8"),
            file_name=f"analisis_{nombre.lower().replace(' ','_')}.md",
            mime="text/plain",
            use_container_width=True,
        )

    # ── Documentos fuente disponibles ────────────────────────────
    st.markdown("---")
    st.markdown('<div style="font-size:13px;font-weight:700;color:#374151;margin-bottom:8px;">📎 Archivos fuente del cliente</div>', unsafe_allow_html=True)
    archivos_cliente = listar_archivos_por_carpeta(path)
    docs = archivos_cliente.get("00_INPUT_CLIENTE/documentos", [])
    if docs:
        cols = st.columns(4)
        for i, doc in enumerate(docs):
            with cols[i % 4]:
                st.markdown(f'<span class="file-tag">📄 {doc["nombre"]}</span>', unsafe_allow_html=True)
    else:
        st.caption("Sin documentos subidos. Ve a **📤 Archivos** para subir materiales del cliente.")


# ─────────────────────────────────────────────
# TAB: ICP
# ─────────────────────────────────────────────
ACCIONES_RAPIDAS_ICP = [
    ("📋 Definir ICP completo", "Con toda la información que tienes del cliente, define el ICP completo incluyendo: macro cargos (niveles jerárquicos), cargos específicos, macro industrias, micro industrias, tamaño de empresa (empleados), criterios de prioridad (qué hace que un lead sea A vs B), criterios de descarte, keywords para Apollo y países foco. Para cargos e industrias incluye la versión separada por coma lista para Apollo."),
    ("👔 Cargos y macro cargos para Apollo", "Dame los macro cargos (niveles: C-Level, VP, Director, Gerente, Jefe) y luego los cargos específicos separados por coma para Apollo. Incluye variaciones en español e inglés. Ordena por prioridad."),
    ("🏭 Industrias por coma para Apollo", "Dame las macro industrias y micro industrias separadas por coma, listas para pegar directamente en Apollo. Incluye las variaciones de nombres exactas que usa Apollo.io."),
    ("🔑 Keywords para Apollo y búsqueda", "Dame las keywords más relevantes para encontrar empresas que calzan con el ICP en Apollo (tecnologías que usan, términos de industria, certificaciones, descripciones de empresa). Sepáralas por coma."),
    ("🏢 Busca empresas reales que calcen", "Basándote en el ICP y en la información que tienes del cliente, dame 15 empresas reales y específicas (con nombre real) que serían clientes ideales en Latinoamérica. Para cada una: nombre, país, industria, tamaño aproximado y por qué calza con el ICP. Si tienes información de empresas mencionadas por el cliente, priorízalas."),
    ("🌐 Analiza la web del cliente y encuentra empresas objetivo", "Revisando la web del cliente y su propuesta de valor, ¿qué tipo de empresas específicas debería prospectar? Busca empresas reales por nombre que calcen con el servicio que ofrece. Considera sus clientes actuales como referencia del perfil ideal."),
    ("⛔ Excluir clientes actuales y descartes", "Según los clientes actuales que ya tiene el cliente, agrégalos a la lista de exclusión para no prospectarlos. Además dame los criterios de descarte completos: qué industrias, tamaños, cargos o características hacen que una empresa NO sea buen prospecto."),
    ("📂 Extrae empresas foco de los archivos", "Si en los archivos que subió el cliente hay listas de empresas objetivo, clientes potenciales o empresas mencionadas, extráelas y organízalas. Si no hay archivos con esa info, dime qué le pediría al cliente para tener esa lista."),
    ("🎯 Criterios de prioridad (Tier A/B/C)", "Define los criterios para clasificar leads en Tier A (ideal, entrar directo), Tier B (calza pero con fricción) y Tier C (prospecto frío). ¿Qué hace que una empresa sea Tier A para este cliente?"),
    ("📏 Ajusta para PyMEs (10-200 empleados)", "Ajusta el ICP para enfocarse en PyMEs de 10 a 200 empleados. ¿Cómo cambian los cargos decisores, el proceso de compra, el ticket y la mensajería?"),
    ("🏛️ Ajusta para empresas grandes (500+)", "Ajusta el ICP para empresas grandes de más de 500 empleados. ¿Cómo cambian los cargos, el ciclo de venta y qué criterios adicionales son relevantes?"),
    ("❓ ¿Qué info me falta del cliente?", "Según lo que sabes, ¿qué información adicional necesitarías para definir el ICP con más precisión? Dame una lista de preguntas específicas para hacerle al cliente en el próximo call."),
    ("✅ Resumen ejecutivo del ICP", "Genera un resumen ejecutivo del ICP definido, en máximo 6 líneas, listo para compartir con el equipo SDR. Incluye: a quién prospectamos, en qué industrias, en qué países y con qué objetivo."),
    ("🔀 Consolida en pocas combinaciones para Snov.io", "Teniendo en cuenta todos los cargos e industrias definidos en el ICP, consolídalos en el MÍNIMO número de combinaciones posibles para cargar en Snov.io o Apollo. El objetivo es cubrir el mercado con máximo 4-5 búsquedas distintas, no 10+. Dame el resultado como tabla: | Combinación | Macro Cargo | Macro Industria | Países | Prioridad |. Que cada fila sea una campaña distinta."),
]


def tab_icp(path: Path, est: dict):
    nombre = est.get("nombre_cliente", "")
    estado_icp = est.get("estado_icp", "pendiente")
    chat_key = f"icp_chat_{path.name}"
    input_key = f"icp_input_{path.name}"

    if chat_key not in st.session_state:
        st.session_state[chat_key] = []
    if input_key not in st.session_state:
        st.session_state[input_key] = ""

    st.markdown("#### 🎯 ICP y Estrategia")
    _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md  →  icp_master.md")

    # ── Estado + botón aprobar ────────────────────────────────────
    col_estado, col_accion = st.columns([3, 1])
    with col_estado:
        if estado_icp == "icp_aprobado":
            st.success("✅ ICP aprobado — `icp_master.md` es la fuente oficial")
        elif estado_icp in ("icp_generado", "icp_pendiente_revision"):
            st.warning("⚡ ICP en revisión — revisa y aprueba cuando esté listo")
        else:
            st.info("⏳ ICP pendiente — usa el chat de abajo para definirlo con Claude")
    with col_accion:
        if estado_icp in ("icp_generado", "icp_pendiente_revision", "icp_aprobado"):
            if st.button("✅ Aprobar ICP", type="primary", use_container_width=True):
                borrador = path / "04_ICP_ESTRATEGIA/icp_borrador.md"
                master = path / "04_ICP_ESTRATEGIA/icp_master.md"
                hist = path / "99_HISTORICO/versiones_anteriores"
                hist.mkdir(parents=True, exist_ok=True)
                if master.exists():
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    shutil.copy2(master, hist / f"icp_master_{ts}.md")
                if borrador.exists():
                    shutil.copy2(borrador, master)
                actualizar_campo(path, "estado_icp", "icp_aprobado")
                st.success("✅ ICP aprobado y guardado en icp_master.md")
                _folder_badge("04_ICP_ESTRATEGIA/icp_master.md")
                st.rerun()

    st.markdown("---")

    # ── Verificar API key ────────────────────────────────────────
    client = get_claude_client()
    if not client:
        st.warning("⚠️ Configura ANTHROPIC_API_KEY en .env para usar Claude.")

    # ── GENERAR ICP AUTOMÁTICO ────────────────────────────────────
    st.markdown("##### ✨ Generar ICP automático")
    st.markdown("""
    <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
                padding:10px 14px;font-size:12px;color:#0369a1;margin-bottom:10px;">
      Claude cruza el <b>análisis web</b> + los <b>documentos del cliente</b> (brief, propuesta, presentación)
      para generar el ICP completo. Revisa y ajusta con el chat de abajo.
    </div>""", unsafe_allow_html=True)

    # Mostrar qué fuentes tiene disponibles
    tiene_analisis = any(est.get(c, "").strip() for c in
                         ["resumen_servicio","propuesta_valor","problema_que_resuelve","icp_tipo_cliente","diferenciacion"])
    docs_disponibles = []
    for carpeta_rel in ["00_INPUT_CLIENTE/documentos", "00_INPUT_CLIENTE/otros", "00_INPUT_CLIENTE/bases"]:
        p_c = path / carpeta_rel
        if p_c.exists():
            docs_disponibles += [f for f in p_c.iterdir() if f.is_file()]

    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        st.markdown(
            f'<div style="background:{"#f0fdf4" if tiene_analisis else "#f8fafc"};border:1px solid '
            f'{"#22c55e" if tiene_analisis else "#e2e8f0"};border-radius:8px;padding:8px 12px;font-size:12px;">'
            f'{"✅" if tiene_analisis else "⬜"} <b>Análisis web</b><br>'
            f'<span style="color:#64748b;">{"5 campos completados" if tiene_analisis else "Pendiente — tab Análisis web"}</span></div>',
            unsafe_allow_html=True)
    with fc2:
        st.markdown(
            f'<div style="background:{"#f0fdf4" if docs_disponibles else "#f8fafc"};border:1px solid '
            f'{"#22c55e" if docs_disponibles else "#e2e8f0"};border-radius:8px;padding:8px 12px;font-size:12px;">'
            f'{"✅" if docs_disponibles else "⬜"} <b>Docs del cliente</b><br>'
            f'<span style="color:#64748b;">{len(docs_disponibles)} archivo(s) disponibles</span></div>',
            unsafe_allow_html=True)
    with fc3:
        tiene_icp_guardado = any(est.get(c, "").strip() for c in ["icp_cargos","icp_industrias","icp_macro_cargos"])
        st.markdown(
            f'<div style="background:{"#f0fdf4" if tiene_icp_guardado else "#f8fafc"};border:1px solid '
            f'{"#22c55e" if tiene_icp_guardado else "#e2e8f0"};border-radius:8px;padding:8px 12px;font-size:12px;">'
            f'{"✅" if tiene_icp_guardado else "⬜"} <b>ICP estructurado</b><br>'
            f'<span style="color:#64748b;">{"Campos completados" if tiene_icp_guardado else "Por generar"}</span></div>',
            unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    gen_col1, gen_col2 = st.columns([2, 3])
    with gen_col1:
        generar_icp = st.button("✨ Generar ICP completo", type="primary",
                                use_container_width=True, key="btn_gen_icp_auto",
                                disabled=(not client))
    with gen_col2:
        if not tiene_analisis and not docs_disponibles:
            st.warning("Sube documentos del cliente o completa el análisis web primero.")

    if generar_icp and client:
        with st.spinner("Cruzando análisis web + documentos del cliente para armar el ICP..."):
            # Leer análisis web
            analisis_ctx = ""
            if tiene_analisis:
                analisis_ctx = f"""
ANÁLISIS WEB DEL CLIENTE:
- Resumen del servicio: {est.get('resumen_servicio','')}
- Propuesta de valor: {est.get('propuesta_valor','')}
- Problema que resuelve: {est.get('problema_que_resuelve','')}
- Tipo de cliente (preliminar): {est.get('icp_tipo_cliente','')}
- Diferenciación: {est.get('diferenciacion','')}"""

            # Leer documentos del cliente (PDF, DOCX, TXT, MD, etc.)
            docs_ctx = leer_todos_los_documentos(path, max_por_doc=4000)

            system_gen_icp = f"""Eres un experto senior en prospección B2B outbound para Latinoamérica.
Tu tarea es construir un ICP (Ideal Customer Profile) completo y accionable para el cliente {nombre}.

REGLA FUNDAMENTAL:
El ICP son los COMPRADORES/CLIENTES de {nombre}, NUNCA sus competidores.
Si {nombre} opera en logística, sus clientes son importadores, exportadores, retailers, manufactureros — los que CONTRATAN logística.
Las empresas del mismo rubro que {nombre} son competencia, NO van en el ICP.
Antes de incluir una industria, pregúntate: "¿esta empresa le pagaría a {nombre} por su servicio?" Si no, no va.

FUENTES DISPONIBLES:
{analisis_ctx}
{"DOCUMENTOS DEL CLIENTE:" + docs_ctx if docs_ctx else ""}

Devuelve EXACTAMENTE este JSON (sin markdown):
{{
  "icp_macro_cargos": "niveles jerárquicos objetivo separados por coma (ej: Director, Gerente, VP)",
  "icp_cargos": "cargos específicos separados por coma, listos para Apollo.io",
  "icp_industrias": "macro industrias separadas por coma, listas para Apollo.io",
  "icp_micro_industrias": "nichos específicos dentro de las industrias, separados por coma",
  "icp_tamano_empresa": "rango de empleados y/o facturación anual aproximada",
  "icp_paises_foco": "países prioritarios separados por coma",
  "icp_keywords": "palabras clave tecnológicas o sectoriales (WMS, ERP, ISO, etc.), separadas por coma",
  "icp_criterios_prioridad": "características que hacen a una empresa Tier A (prioridad máxima)",
  "icp_criterios_descarte": "empresas o perfiles que nunca se deben prospectar",
  "icp_empresas_foco": "3-5 empresas reales de la región que calzan perfectamente con el ICP"
}}

Sé específico. Usa terminología real de Apollo.io. Considera el contexto B2B latinoamericano."""

            raw = llamar_claude(client,
                [{"role": "user", "content": f"Genera el ICP completo para {nombre} ({est.get('sitio_web','')})."}],
                system_gen_icp, max_tokens=2000)
            try:
                txt = raw.strip()
                start = txt.find('{'); end = txt.rfind('}')
                if start >= 0 and end > start: txt = txt[start:end+1]
                icp_result = json.loads(txt)
                for campo, valor in icp_result.items():
                    actualizar_campo(path, campo, valor)
                actualizar_campo(path, "estado_icp", "icp_pendiente_revision")
                # Guardar en icp_borrador.md
                borrador_path = path / "04_ICP_ESTRATEGIA/icp_borrador.md"
                ts = datetime.now().strftime("%Y-%m-%d %H:%M")
                borrador_txt = f"# ICP Borrador: {nombre}\n\nGenerado: {ts}\n\n"
                for campo, valor in icp_result.items():
                    borrador_txt += f"## {campo}\n{valor}\n\n"
                borrador_path.write_text(borrador_txt, encoding="utf-8")
                st.success("✅ ICP generado. Revisa los campos abajo — puedes editarlos y refinar con el chat.")
                st.rerun()
            except Exception as e:
                st.error(f"Error parseando respuesta: {e}")
                st.code(raw[:600])

    st.markdown("---")

    # ── CHAT ─────────────────────────────────────────────────────
    st.markdown("""
    <div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:10px;">
      💬 Chat ICP con Claude
      <span style="font-size:12px;font-weight:400;color:#64748b;margin-left:8px;">El contexto del cliente se inyecta automáticamente</span>
    </div>
    """, unsafe_allow_html=True)

    # Helper para enviar mensaje
    def _enviar_icp(msg_texto: str):
        if not msg_texto.strip() or not client:
            return
        system_prompt = build_icp_system_prompt(est, path)
        hist = st.session_state[chat_key]
        messages_api = [{"role": m["role"], "content": m["content"]} for m in hist]
        messages_api.append({"role": "user", "content": msg_texto.strip()})
        with st.spinner("Claude está pensando..."):
            respuesta = llamar_claude(client, messages_api, system_prompt)
        st.session_state[chat_key].append({"role": "user",      "content": msg_texto.strip()})
        st.session_state[chat_key].append({"role": "assistant", "content": respuesta})
        st.session_state[input_key] = ""
        # Guardar en borrador
        borrador_path = path / "04_ICP_ESTRATEGIA/icp_borrador.md"
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        entrada = f"\n\n---\n### Chat {ts}\n**Pregunta:** {msg_texto.strip()}\n\n**Respuesta:**\n{respuesta}"
        base = borrador_path.read_text(encoding="utf-8") if borrador_path.exists() else f"# ICP Borrador: {nombre}\n"
        borrador_path.write_text(base + entrada, encoding="utf-8")
        actualizar_campo(path, "estado_icp", "icp_pendiente_revision")
        st.rerun()

    # Acciones rápidas — ahora auto-envían
    st.markdown('<div style="font-size:12px;color:#64748b;margin-bottom:6px;">Acciones rápidas (hacen clic y se envían automáticamente):</div>', unsafe_allow_html=True)
    chip_cols = st.columns(3)
    for i, (label, prompt_accion) in enumerate(ACCIONES_RAPIDAS_ICP):
        with chip_cols[i % 3]:
            if st.button(label, key=f"chip_icp_{i}", use_container_width=True, disabled=(not client)):
                _enviar_icp(prompt_accion)

    st.markdown("<br>", unsafe_allow_html=True)

    # Historial del chat
    chat_history = st.session_state[chat_key]
    if chat_history:
        for msg in chat_history:
            if msg["role"] == "user":
                st.markdown(
                    f'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:10px;'
                    f'padding:10px 14px;margin:6px 0;margin-left:20%;">'
                    f'<span style="font-size:11px;color:#1e40af;font-weight:700;">TÚ</span><br>'
                    f'<span style="font-size:13px;color:#1e293b;">{msg["content"]}</span></div>',
                    unsafe_allow_html=True)
            else:
                st.markdown(
                    f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;'
                    f'padding:12px 16px;margin:6px 0;margin-right:10%;">'
                    f'<span style="font-size:11px;color:#475569;font-weight:700;">CLAUDE</span><br>'
                    f'<div style="font-size:13px;color:#1e293b;line-height:1.7;white-space:pre-wrap;">{msg["content"]}</div></div>',
                    unsafe_allow_html=True)
        if st.button("🗑 Limpiar conversación", key="icp_limpiar"):
            st.session_state[chat_key] = []
            st.rerun()
    else:
        st.markdown("""
        <div style="background:#f8fafc;border:2px dashed #e2e8f0;border-radius:12px;
                    padding:24px;text-align:center;color:#94a3b8;font-size:13px;margin-bottom:12px;">
          💬 Usa una acción rápida o escribe / habla tu pregunta abajo
        </div>
        """, unsafe_allow_html=True)

    # ── Input: texto + voz del servidor ─────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)

    voz_key = f"icp_voz_{path.name}"
    col_voz, col_dur = st.columns([2, 1])
    with col_dur:
        duracion_voz = st.selectbox("Duración", [10, 20, 30, 45], index=1,
                                    key=f"icp_dur_{path.name}",
                                    format_func=lambda x: f"{x} seg")
    with col_voz:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🎤 Grabar y enviar", key=f"icp_grabar_{path.name}",
                     use_container_width=True, disabled=(not client)):
            with st.spinner(f"🔴 Grabando {duracion_voz} seg — habla ahora..."):
                texto_voz = grabar_y_transcribir(duracion=duracion_voz)
            if texto_voz.startswith("ERROR"):
                st.error(f"No se pudo transcribir: {texto_voz}")
            else:
                st.info(f"🎤 Reconocido: *{texto_voz}*")
                _enviar_icp(texto_voz)

    col_input, col_send = st.columns([5, 1])
    with col_input:
        textarea_key = f"icp_textarea_{path.name}"
        user_msg = st.text_area(
            "Mensaje", label_visibility="collapsed",
            placeholder="O escribe tu pregunta aquí...",
            height=80, key=textarea_key
        )
    with col_send:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Enviar →", type="primary", use_container_width=True, disabled=(not client)):
            _enviar_icp(user_msg)

    st.markdown("---")

    # ── ICP ESTRUCTURADO ─────────────────────────────────────────
    st.markdown("""
    <div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:4px;">
      📊 ICP Estructurado
      <span style="font-size:12px;font-weight:400;color:#64748b;margin-left:8px;">Completa o edita tras la conversación con Claude — se guarda en icp_borrador.md y estado_cliente.json</span>
    </div>
    """, unsafe_allow_html=True)

    def _icp_field(titulo, icono, bg, borde, tc, valor_est, placeholder, key, height=80, tipo="area"):
        st.markdown(f"""<div style="background:{bg};border:1.5px solid {borde};border-radius:8px;
                        padding:8px 14px;margin-bottom:4px;font-size:12px;font-weight:700;color:{tc};">
                        {icono} {titulo}</div>""", unsafe_allow_html=True)
        if tipo == "area":
            return st.text_area(titulo, label_visibility="collapsed", value=valor_est,
                placeholder=placeholder, height=height, key=key)
        else:
            return st.text_input(titulo, label_visibility="collapsed", value=valor_est,
                placeholder=placeholder, key=key)

    # Fila 1: Cargos
    ic1, ic2 = st.columns(2)
    with ic1:
        icp_macro_cargos = _icp_field(
            "Macro cargos (nivel jerárquico)", "🏢",
            "#dbeafe", "#93c5fd", "#1e3a8a",
            est.get("icp_macro_cargos", ""),
            "C-Level, VP, Director, Gerente, Jefe de área...",
            "icp_s_macro_cargos", height=70
        )
        _folder_badge("04_ICP_ESTRATEGIA/macro_cargos.md")

    with ic2:
        icp_cargos = _icp_field(
            "Cargos específicos", "👔",
            "#eff6ff", "#bfdbfe", "#1e40af",
            est.get("icp_cargos", ""),
            "CEO, Director de Operaciones, Gerente Logística, Head of Supply Chain...",
            "icp_s_cargos", height=70
        )
        _folder_badge("04_ICP_ESTRATEGIA/cargos_apollo_por_coma.md")
        if icp_cargos.strip():
            st.code(", ".join([c.strip() for c in icp_cargos.replace("\n",",").split(",") if c.strip()]), language=None)

    st.markdown("<br>", unsafe_allow_html=True)

    # Fila 2: Industrias
    ic3, ic4 = st.columns(2)
    with ic3:
        icp_industrias = _icp_field(
            "Macro industrias", "🏭",
            "#f0fdf4", "#bbf7d0", "#15803d",
            est.get("icp_industrias", ""),
            "Logística, Supply Chain, Manufactura, Retail con distribución...",
            "icp_s_industrias", height=70
        )
        _folder_badge("04_ICP_ESTRATEGIA/industrias_apollo_por_coma.md")
        if icp_industrias.strip():
            st.code(", ".join([i.strip() for i in icp_industrias.replace("\n",",").split(",") if i.strip()]), language=None)

    with ic4:
        icp_micro = _icp_field(
            "Micro industrias / nicho", "🔍",
            "#fdf4ff", "#e9d5ff", "#7e22ce",
            est.get("icp_micro_industrias", ""),
            "3PL, Cold chain, Last mile delivery, Freight forwarding...",
            "icp_s_micro", height=70
        )
        _folder_badge("04_ICP_ESTRATEGIA/micro_industrias.md")

    st.markdown("<br>", unsafe_allow_html=True)

    # Fila 3: Tamaño + Países + Keywords
    ic5, ic6, ic7 = st.columns(3)
    with ic5:
        icp_tamano = _icp_field(
            "Tamaño de empresa", "📏",
            "#fef3c7", "#fde68a", "#92400e",
            est.get("icp_tamano_empresa", ""),
            "50-500 empleados · USD 1M-20M facturación",
            "icp_s_tamano", tipo="input"
        )
        _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md")

    with ic6:
        icp_paises_foco = _icp_field(
            "Países prioritarios", "🌎",
            "#f0f9ff", "#bae6fd", "#0369a1",
            est.get("icp_paises_foco", est.get("pais_objetivo", "")),
            "Chile, México, Colombia, Perú...",
            "icp_s_paises", tipo="input"
        )
        _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md")

    with ic7:
        icp_keywords = _icp_field(
            "Keywords de búsqueda", "🔑",
            "#ecfdf5", "#6ee7b7", "#065f46",
            est.get("icp_keywords", ""),
            "WMS, TMS, ERP logístico, ISO 9001...",
            "icp_s_keywords", tipo="input"
        )
        _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md")

    st.markdown("<br>", unsafe_allow_html=True)

    # Fila 4: Prioridad + Descarte
    ic8, ic9 = st.columns(2)
    with ic8:
        icp_prioridad = _icp_field(
            "Criterios de prioridad (Tier A)", "🎯",
            "#fefce8", "#fef08a", "#713f12",
            est.get("icp_criterios_prioridad", ""),
            "Empresa con 100+ empleados + industria logística + cargo Director o superior + en Chile o México...",
            "icp_s_prioridad", height=90
        )
        _folder_badge("04_ICP_ESTRATEGIA/criterios_prioridad.md")

    with ic9:
        icp_descarte = _icp_field(
            "Criterios de descarte", "🚫",
            "#fff1f2", "#fecdd3", "#be123c",
            est.get("icp_criterios_descarte", ""),
            "Gobierno, startups sin funding, retail B2C, clientes actuales, empresas <10 empleados...",
            "icp_s_descarte", height=90
        )
        _folder_badge("04_ICP_ESTRATEGIA/criterios_descarte.md")

    st.markdown("<br>", unsafe_allow_html=True)

    # Fila 5: Empresas foco
    icp_empresas_foco = _icp_field(
        "Empresas foco (target list)", "🏹",
        "#f8fafc", "#e2e8f0", "#334155",
        est.get("icp_empresas_foco", ""),
        "Empresa X (Chile, logística, 200 emp.) · Empresa Y (México, manufactura) · ... — pégalas del chat o súbelas como archivo",
        "icp_s_empresas_foco", height=80
    )
    _folder_badge("07_APOLLO_Y_BUSQUEDAS/empresas_foco.md  ·  estado_cliente.json")

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("💾 Guardar ICP estructurado", type="primary", use_container_width=False):
        campos_guardar = {
            "icp_macro_cargos": icp_macro_cargos.strip(),
            "icp_cargos": icp_cargos.strip(),
            "icp_industrias": icp_industrias.strip(),
            "icp_micro_industrias": icp_micro.strip(),
            "icp_tamano_empresa": icp_tamano.strip(),
            "icp_paises_foco": icp_paises_foco.strip(),
            "icp_keywords": icp_keywords.strip(),
            "icp_criterios_prioridad": icp_prioridad.strip(),
            "icp_criterios_descarte": icp_descarte.strip(),
            "icp_empresas_foco": icp_empresas_foco.strip(),
        }
        for campo, valor in campos_guardar.items():
            actualizar_campo(path, campo, valor)

        fecha = datetime.now().strftime("%Y-%m-%d")

        def _a_coma(texto):
            return ", ".join([x.strip() for x in texto.replace("\n", ",").split(",") if x.strip()])

        if icp_macro_cargos.strip():
            (path / "04_ICP_ESTRATEGIA/macro_cargos.md").write_text(
                f"# Macro Cargos: {nombre}\n\n**Fecha:** {fecha}\n\n```\n{_a_coma(icp_macro_cargos)}\n```",
                encoding="utf-8"
            )
        if icp_cargos.strip():
            (path / "04_ICP_ESTRATEGIA/cargos_apollo_por_coma.md").write_text(
                f"# Cargos Apollo: {nombre}\n\n**Fecha:** {fecha}\n\n```\n{_a_coma(icp_cargos)}\n```",
                encoding="utf-8"
            )
        if icp_industrias.strip():
            (path / "04_ICP_ESTRATEGIA/industrias_apollo_por_coma.md").write_text(
                f"# Industrias Apollo: {nombre}\n\n**Fecha:** {fecha}\n\n```\n{_a_coma(icp_industrias)}\n```",
                encoding="utf-8"
            )
        if icp_micro.strip():
            (path / "04_ICP_ESTRATEGIA/micro_industrias.md").write_text(
                f"# Micro Industrias: {nombre}\n\n**Fecha:** {fecha}\n\n{icp_micro.strip()}",
                encoding="utf-8"
            )
        if icp_prioridad.strip():
            (path / "04_ICP_ESTRATEGIA/criterios_prioridad.md").write_text(
                f"# Criterios de Prioridad: {nombre}\n\n**Fecha:** {fecha}\n\n{icp_prioridad.strip()}",
                encoding="utf-8"
            )
        if icp_descarte.strip():
            (path / "04_ICP_ESTRATEGIA/criterios_descarte.md").write_text(
                f"# Criterios de Descarte: {nombre}\n\n**Fecha:** {fecha}\n\n{icp_descarte.strip()}",
                encoding="utf-8"
            )
        if icp_empresas_foco.strip():
            carpeta_apollo = path / "07_APOLLO_Y_BUSQUEDAS"
            carpeta_apollo.mkdir(parents=True, exist_ok=True)
            (carpeta_apollo / "empresas_foco.md").write_text(
                f"# Empresas Foco: {nombre}\n\n**Fecha:** {fecha}\n\n{icp_empresas_foco.strip()}",
                encoding="utf-8"
            )

        actualizar_campo(path, "estado_icp", "icp_pendiente_revision")
        st.toast("✅ ICP guardado", icon="🎯")
        ir_a_tab(4)   # Avanza a Bases y Mensajería

    st.markdown("---")

    # ── Borrador completo ─────────────────────────────────────────
    with st.expander("📄 Ver / editar icp_borrador.md completo", expanded=False):
        borrador = path / "04_ICP_ESTRATEGIA/icp_borrador.md"
        if borrador.exists():
            contenido_bor = borrador.read_text(encoding="utf-8")
            nuevo_bor = st.text_area("icp_borrador.md", value=contenido_bor, height=350, key="icp_bor_full")
            col_gb, col_db = st.columns(2)
            with col_gb:
                if st.button("💾 Guardar borrador"):
                    borrador.write_text(nuevo_bor, encoding="utf-8")
                    actualizar_campo(path, "estado_icp", "icp_pendiente_revision")
                    st.success("Guardado")
                    _folder_badge("04_ICP_ESTRATEGIA/icp_borrador.md")
            with col_db:
                st.download_button("⬇️ Descargar", data=contenido_bor.encode(), file_name="icp_borrador.md")

    with st.expander("📋 Ver icp_master.md (oficial aprobado)", expanded=False):
        master = path / "04_ICP_ESTRATEGIA/icp_master.md"
        if master.exists():
            contenido_master = master.read_text(encoding="utf-8")
            st.text_area("icp_master.md", value=contenido_master, height=300, disabled=True, key="icp_master_view")
            st.download_button("⬇️ Descargar master", data=contenido_master.encode(), file_name="icp_master.md")


# ─────────────────────────────────────────────
# APOLLO.IO — Valores reales de filtros
# ─────────────────────────────────────────────
APOLLO_SENIORITY = {
    "C-Suite / Owner":    "c_suite",
    "Founder":            "founder",
    "Partner":            "partner",
    "VP":                 "vp",
    "Head of":            "head",
    "Director":           "director",
    "Manager / Gerente":  "manager",
    "Senior":             "senior",
    "Entry Level":        "entry",
}

APOLLO_DEPARTAMENTOS = {
    "Ventas / Sales":           "sales",
    "Marketing":                "marketing",
    "Operaciones":              "operations",
    "Tecnologia / IT":          "information_technology",
    "Finanzas":                 "finance",
    "Recursos Humanos":         "human_resources",
    "Ingenieria":               "engineering",
    "Producto":                 "product_management",
    "Atencion al cliente":      "customer_success",
    "Desarrollo de negocios":   "business_development",
    "Legal":                    "legal",
    "Contabilidad":             "accounting",
    "Diseno":                   "design",
    "Compras / Procurement":    "procurement",
    "Comunicacion / PR":        "media_and_communication",
    "Administrativo":           "administrative",
}

APOLLO_TAMANOS = {
    "1-10 empleados":         "1,10",
    "11-50 empleados":        "11,50",
    "51-100 empleados":       "51,100",
    "101-200 empleados":      "101,200",
    "201-500 empleados":      "201,500",
    "501-1,000 empleados":    "501,1000",
    "1,001-2,000 empleados":  "1001,2000",
    "2,001-5,000 empleados":  "2001,5000",
    "5,001-10,000 empleados": "5001,10000",
    "10,001+ empleados":      "10001,",
}

APOLLO_INDUSTRIAS = sorted([
    "Accounting", "Aerospace & Defense", "Agriculture", "Airlines/Aviation",
    "Architecture & Planning", "Automotive", "Banking", "Biotechnology",
    "Broadcast Media", "Capital Markets", "Chemicals", "Civil Engineering",
    "Computer Hardware", "Computer Networking", "Computer & Network Security",
    "Computer Software", "Construction", "Consumer Goods",
    "Design", "E-Learning", "Education Management",
    "Electrical/Electronic Manufacturing", "Entertainment", "Environmental Services",
    "Events Services", "Facilities Services", "Farming", "Financial Services",
    "Food & Beverages", "Food Production", "Government Administration",
    "Health, Wellness and Fitness", "Higher Education", "Hospital & Health Care",
    "Hospitality", "Human Resources", "Import and Export", "Industrial Automation",
    "Information Technology and Services", "Insurance", "Internet",
    "Investment Banking", "Law Practice", "Legal Services",
    "Logistics and Supply Chain", "Luxury Goods & Jewelry", "Machinery",
    "Management Consulting", "Manufacturing", "Maritime",
    "Marketing and Advertising", "Media Production", "Medical Devices",
    "Military", "Mining & Metals", "Non-Profit Organization Management",
    "Oil & Energy", "Online Media", "Outsourcing/Offshoring",
    "Package/Freight Delivery", "Pharmaceuticals", "Plastics",
    "Primary/Secondary Education", "Professional Training & Coaching",
    "Public Policy", "Publishing", "Real Estate", "Renewables & Environment",
    "Research", "Restaurants", "Retail", "Semiconductors",
    "Staffing and Recruiting", "Supermarkets", "Telecommunications",
    "Transportation/Trucking/Railroad", "Utilities",
    "Venture Capital & Private Equity", "Warehousing", "Wireless",
])


def _icp_a_apollo_filtros(est: dict) -> dict:
    """Mapea campos del ICP guardado a filtros de Apollo (retorna display labels)."""
    import re as _re
    txt_cargos = ((est.get("icp_macro_cargos") or "") + " " + (est.get("icp_cargos") or "")).lower()
    txt_tamano = (est.get("icp_tamano_empresa") or "").lower()
    txt_ind    = ((est.get("icp_industrias") or "") + " " + (est.get("icp_micro_industrias") or "")).lower()
    txt_desc   = (est.get("icp_criterios_descarte") or "").lower()

    sen_kw = {
        "c-suite": "C-Suite / Owner", "c suite": "C-Suite / Owner",
        "ceo": "C-Suite / Owner", "coo": "C-Suite / Owner",
        "cto": "C-Suite / Owner", "cfo": "C-Suite / Owner",
        "owner": "C-Suite / Owner", "propietar": "C-Suite / Owner",
        "fundador": "Founder", "founder": "Founder",
        "partner": "Partner", "socio": "Partner",
        "vicepresidente": "VP", "vice president": "VP", " vp ": "VP",
        "director": "Director",
        "head of": "Head of",
        "gerente": "Manager / Gerente", "manager": "Manager / Gerente", "jefe": "Manager / Gerente",
        "senior": "Senior",
    }
    seniority = []
    for kw, lbl in sen_kw.items():
        if kw in txt_cargos and lbl not in seniority:
            seniority.append(lbl)
    if not seniority:
        seniority = ["C-Suite / Owner", "Director", "VP"]

    ind_kw = {
        "logis": "Logistics and Supply Chain", "supply chain": "Logistics and Supply Chain",
        "transport": "Transportation/Trucking/Railroad",
        "wareho": "Warehousing", "almacen": "Warehousing",
        "manufactur": "Manufacturing",
        "retail": "Retail",
        "tecnolog": "Information Technology and Services",
        "software": "Computer Software",
        "salud": "Hospital & Health Care", "health": "Hospital & Health Care",
        "farmac": "Pharmaceuticals",
        "construc": "Construction",
        "inmobil": "Real Estate",
        "financ": "Financial Services",
        "banca": "Banking",
        "seguro": "Insurance",
        "educaci": "Education Management",
        "marketing": "Marketing and Advertising",
        "consultor": "Management Consulting",
        "aliment": "Food & Beverages", "food": "Food & Beverages",
        "energia": "Oil & Energy",
        "mining": "Mining & Metals", "miner": "Mining & Metals",
        "automotr": "Automotive",
        "import": "Import and Export", "export": "Import and Export",
        "agro": "Agriculture",
    }
    industrias = []
    for kw, ind in ind_kw.items():
        if kw in txt_ind and ind not in industrias:
            industrias.append(ind)

    excl_kw = {
        "gobierno": "Government Administration", "government": "Government Administration",
        "militar": "Military",
        "non-profit": "Non-Profit Organization Management", "ong": "Non-Profit Organization Management",
    }
    ind_excluidas = []
    for kw, ind in excl_kw.items():
        if kw in txt_desc and ind not in ind_excluidas:
            ind_excluidas.append(ind)

    nums = [int(_re.sub(r"[,\.]", "", n)) for n in _re.findall(r"\d[\d,\.]*", txt_tamano)]
    tamano = []
    rango_map = [
        (1, 10, "1-10 empleados"), (11, 50, "11-50 empleados"),
        (51, 100, "51-100 empleados"), (101, 200, "101-200 empleados"),
        (201, 500, "201-500 empleados"), (501, 1000, "501-1,000 empleados"),
        (1001, 2000, "1,001-2,000 empleados"), (2001, 5000, "2,001-5,000 empleados"),
        (5001, 10000, "5,001-10,000 empleados"), (10001, 9999999, "10,001+ empleados"),
    ]
    if len(nums) >= 2:
        lo, hi = min(nums), max(nums)
        for r_lo, r_hi, lbl in rango_map:
            if r_lo <= hi and r_hi >= lo:
                tamano.append(lbl)
    elif len(nums) == 1:
        v = nums[0]
        for r_lo, r_hi, lbl in rango_map:
            if r_lo <= v <= r_hi:
                tamano.append(lbl)

    return {"seniority": seniority, "industrias": industrias,
            "ind_excluidas": ind_excluidas, "tamano": tamano}


# ─────────────────────────────────────────────
# APOLLO.IO API
# ─────────────────────────────────────────────
def _apollo_buscar(cargos: list, paises: list, keywords: list = None,
                   seniority: list = None, departments: list = None,
                   tamano: list = None,
                   pagina: int = 1, por_pagina: int = 100) -> dict:
    """Llama a Apollo People Search API."""
    import requests as _req
    api_key = os.environ.get("APOLLO_API_KEY", "").strip()
    if not api_key:
        return {"error": "APOLLO_API_KEY no configurada en .env"}
    payload: dict = {"page": pagina, "per_page": min(por_pagina, 100)}
    if cargos:      payload["person_titles"]                    = [c.strip() for c in cargos[:10]  if c.strip()]
    if paises:      payload["person_locations"]                 = [p.strip() for p in paises[:10]  if p.strip()]
    if keywords:    payload["q_organization_keyword_tags"]      = [k.strip() for k in keywords     if k.strip()]
    if seniority:   payload["person_seniorities"]               = seniority
    if departments: payload["person_departments"]               = departments
    if tamano:      payload["organization_num_employees_ranges"] = tamano
    try:
        r = _req.post("https://api.apollo.io/api/v1/mixed_people/api_search",
                      json=payload,
                      headers={"Content-Type": "application/json", "x-api-key": api_key},
                      timeout=30)
        return r.json() if r.status_code == 200 else {"error": f"Apollo {r.status_code}: {r.text[:400]}"}
    except Exception as e:
        return {"error": str(e)}


def _apollo_a_df(people: list) -> "pd.DataFrame":
    import pandas as _pd
    rows = []
    for p in people:
        org = p.get("organization") or {}

        # Nombre: first/last o split del campo "name"
        first = p.get("first_name") or ""
        last  = p.get("last_name")  or ""
        if not first and not last:
            parts = (p.get("name") or "").split(" ", 1)
            first = parts[0]
            last  = parts[1] if len(parts) > 1 else ""

        # Teléfonos: separar móvil de fijo
        phones   = p.get("phone_numbers") or []
        mobile   = next((ph.get("sanitized_number", "") for ph in phones
                         if ph.get("type") in ("mobile", "cell") and ph.get("sanitized_number")), "")
        landline = next((ph.get("sanitized_number", "") for ph in phones
                         if ph.get("type") not in ("mobile", "cell") and ph.get("sanitized_number")), "")
        if not mobile and not landline and phones:
            mobile = phones[0].get("sanitized_number", "")

        # Industria: persona > org
        industry = (p.get("industry") or org.get("industry")
                    or org.get("primary_industry") or "")

        # Empleados
        employees = (org.get("estimated_num_employees")
                     or org.get("num_employees")
                     or p.get("estimated_num_employees") or "")

        # Ubicación
        country = p.get("country") or org.get("country") or ""
        city    = p.get("city")    or org.get("city")    or ""
        state   = p.get("state")   or org.get("state")   or ""

        rows.append({
            "First Name":       first,
            "Last Name":        last,
            "Title":            p.get("title", ""),
            "Company":          p.get("organization_name", "") or org.get("name", ""),
            "Industry":         industry,
            "# Employees":      employees,
            "Country":          country,
            "State":            state,
            "City":             city,
            "Email":            p.get("email", ""),
            "Email Status":     p.get("email_status", ""),
            "Mobile Phone":     mobile,
            "Direct Phone":     landline,
            "Person LinkedIn":  p.get("linkedin_url", ""),
            "Company LinkedIn": org.get("linkedin_url", ""),
            "Website":          org.get("website_url", ""),
            "Apollo ID":        p.get("id", ""),
        })
    return _pd.DataFrame(rows)


# ─────────────────────────────────────────────
# SNOV.IO API
# ─────────────────────────────────────────────
def _snov_token() -> str | None:
    import requests as _req
    cid = os.environ.get("SNOV_CLIENT_ID", "").strip()
    sec = os.environ.get("SNOV_CLIENT_SECRET", "").strip()
    if not cid or not sec:
        return None
    try:
        r = _req.post("https://api.snov.io/v1/oauth/access_token",
                      data={"grant_type": "client_credentials", "client_id": cid, "client_secret": sec},
                      timeout=15)
        return r.json().get("access_token") if r.status_code == 200 else None
    except Exception:
        return None


def _snov_buscar(token: str, cargos: list, industrias: list, paises: list,
                 pagina: int = 1, por_pagina: int = 100) -> dict:
    import requests as _req
    payload: dict = {"access_token": token, "per_page": min(por_pagina, 100), "page": pagina}
    if cargos:     payload["positions"]  = cargos[:5]
    if industrias: payload["industries"] = industrias[:5]
    if paises:     payload["countries"]  = paises[:5]
    try:
        r = _req.post("https://api.snov.io/v2/api/search-contacts",
                      json=payload,
                      headers={"Content-Type": "application/json"},
                      timeout=30)
        return r.json() if r.status_code == 200 else {"error": f"Snov {r.status_code}: {r.text[:400]}"}
    except Exception as e:
        return {"error": str(e)}


def _snov_a_df(data: dict) -> "pd.DataFrame":
    import pandas as _pd
    contacts = data.get("data", data.get("contacts", data.get("prospects", [])))
    if not isinstance(contacts, list):
        return _pd.DataFrame()
    rows = []
    for c in contacts:
        rows.append({
            "First Name":      c.get("firstName", c.get("first_name", "")),
            "Last Name":       c.get("lastName",  c.get("last_name", "")),
            "Title":           c.get("position",  c.get("title", "")),
            "Company":         c.get("company",   c.get("companyName", "")),
            "Industry":        c.get("industry", ""),
            "# Employees":     c.get("companySize", c.get("employees", "")),
            "Country":         c.get("country", ""),
            "City":            c.get("city", ""),
            "Email":           c.get("email", ""),
            "Email Status":    c.get("emailStatus", ""),
            "Phone":           c.get("phone", ""),
            "Person LinkedIn": c.get("linkedInUrl", c.get("linkedin", "")),
            "Website":         c.get("companyWebsite", c.get("website", "")),
            "Snov ID":         c.get("id", ""),
        })
    return _pd.DataFrame(rows)


def _acumular_base(ruta_csv: Path, df_nuevo: "pd.DataFrame", id_col: str) -> tuple:
    """Agrega df_nuevo a ruta_csv deduplicando por id_col. Retorna (df_total, n_nuevos)."""
    import pandas as _pd
    if ruta_csv.exists():
        df_old = _pd.read_csv(str(ruta_csv), encoding="utf-8-sig", dtype=str)
        ids_old = set(df_old[id_col].dropna().values) if id_col in df_old.columns else set()
        df_nuevos = df_nuevo[~df_nuevo[id_col].astype(str).isin(ids_old)]
        df_total  = _pd.concat([df_old, df_nuevos], ignore_index=True)
        n_nuevos  = len(df_nuevos)
    else:
        df_total = df_nuevo.copy()
        n_nuevos = len(df_nuevo)
    df_total.to_csv(str(ruta_csv), index=False, encoding="utf-8-sig")
    return df_total, n_nuevos


# ─────────────────────────────────────────────
# TAB: BASES APOLLO
# ─────────────────────────────────────────────
def _parsear_lista_icp(texto: str) -> list[str]:
    """Convierte texto de ICP (separado por comas o saltos) en lista de keywords en minúscula."""
    return [x.strip().lower() for x in texto.replace("\n", ",").split(",") if x.strip()]


def _match_keywords(texto: str, keywords: list[str]) -> str | None:
    """Retorna el primer keyword que matchea en el texto (insensible a mayúsculas)."""
    texto_l = texto.lower()
    for kw in keywords:
        # Chequeo de palabras individuales del keyword
        palabras = [p for p in kw.split() if len(p) > 2]
        if palabras and all(p in texto_l for p in palabras):
            return kw.title()
        if kw and kw in texto_l:
            return kw.title()
    return None


def _parse_empleados(texto: str) -> int | None:
    """Extrae número de empleados de strings como '51-200', '500+', '1001-5000'."""
    import re
    if not texto or str(texto).strip() in ("", "nan", "None"):
        return None
    s = str(texto).replace(",", "").replace(".", "")
    nums = re.findall(r"\d+", s)
    if nums:
        return int(nums[0])
    return None


def _clasificar_base(df, icp: dict) -> "pd.DataFrame":
    import pandas as pd

    cargos_kw      = _parsear_lista_icp(icp.get("icp_cargos", ""))
    macro_cargos_kw = _parsear_lista_icp(icp.get("icp_macro_cargos", ""))
    industrias_kw  = _parsear_lista_icp(icp.get("icp_industrias", ""))
    micro_ind_kw   = _parsear_lista_icp(icp.get("icp_micro_industrias", ""))
    paises_kw      = _parsear_lista_icp(icp.get("icp_paises_foco", icp.get("pais_objetivo", "")))
    descarte_kw    = _parsear_lista_icp(icp.get("icp_criterios_descarte", ""))
    keywords_kw    = _parsear_lista_icp(icp.get("icp_keywords", ""))
    empresas_excl  = _parsear_lista_icp(icp.get("clientes_actuales", ""))

    # Detectar columnas Apollo (nombres varían según exportación)
    cols = {c.lower().strip(): c for c in df.columns}
    def _col(*nombres):
        for n in nombres:
            if n in cols: return cols[n]
        return None

    col_titulo    = _col("title", "job title", "cargo", "título")
    col_empresa   = _col("company", "company name", "empresa", "company name for emails")
    col_industria = _col("industry", "industria")
    col_empleados = _col("# employees", "employees", "employee count", "empleados", "num employees")
    col_pais      = _col("country", "país", "pais", "company country")
    col_email     = _col("email", "correo", "work email")
    col_keywords  = _col("keywords", "seo description", "technologies")

    resultados = []
    for _, row in df.iterrows():
        titulo    = str(row[col_titulo]    if col_titulo    else "").strip()
        empresa   = str(row[col_empresa]   if col_empresa   else "").strip()
        industria = str(row[col_industria] if col_industria else "").strip()
        pais      = str(row[col_pais]      if col_pais      else "").strip()
        kw_raw    = str(row[col_keywords]  if col_keywords  else "").strip()
        emp_num   = _parse_empleados(row[col_empleados] if col_empleados else None)

        # Macro cargo
        macro_cargo = (_match_keywords(titulo, macro_cargos_kw)
                       or _match_keywords(titulo, cargos_kw)
                       or "Sin clasificar")

        # Cargo específico (coincidencia más fina)
        cargo_esp = _match_keywords(titulo, cargos_kw) or "—"

        # Macro industria
        macro_ind = (_match_keywords(industria, industrias_kw)
                     or _match_keywords(industria + " " + kw_raw, industrias_kw)
                     or "Sin clasificar")

        # Micro industria
        micro_ind = (_match_keywords(industria + " " + kw_raw, micro_ind_kw) or "—")

        # País OK
        pais_ok = bool(paises_kw) and bool(_match_keywords(pais, paises_kw))
        if not paises_kw:
            pais_ok = True

        # Keyword match
        kw_match = bool(keywords_kw and _match_keywords(titulo + " " + industria + " " + kw_raw, keywords_kw))

        # Empresa excluida (cliente actual)
        excluida = bool(empresas_excl and _match_keywords(empresa, empresas_excl))

        # Descarte por criteria
        descarte_match = bool(descarte_kw and (
            _match_keywords(titulo, descarte_kw) or
            _match_keywords(industria, descarte_kw) or
            _match_keywords(empresa, descarte_kw)
        ))

        # Score y Tier
        score = 0
        if macro_cargo != "Sin clasificar": score += 3
        if cargo_esp   != "—":             score += 1
        if macro_ind   != "Sin clasificar": score += 3
        if pais_ok:                         score += 2
        if kw_match:                        score += 1

        if excluida or descarte_match:
            tier = "❌ Excluir"
        elif score >= 7:
            tier = "🟢 Tier A"
        elif score >= 4:
            tier = "🟡 Tier B"
        elif score >= 2:
            tier = "🟠 Tier C"
        else:
            tier = "⚫ Sin match"

        resultados.append({
            "CP_Tier":           tier,
            "CP_Macro_Cargo":    macro_cargo,
            "CP_Cargo_Esp":      cargo_esp,
            "CP_Macro_Industria": macro_ind,
            "CP_Micro_Industria": micro_ind,
            "CP_Pais_OK":        "✓" if pais_ok else "✗",
            "CP_KW_Match":       "✓" if kw_match else "",
            "CP_Excluir":        "⚠️ Sí" if excluida else "",
        })

    result_df = pd.DataFrame(resultados)
    return pd.concat([result_df, df.reset_index(drop=True)], axis=1)


def tab_bases_apollo(path: Path, est: dict):
    import pandas as pd
    import io

    nombre = est.get("nombre_cliente", "")
    fecha_str = datetime.now().strftime("%Y%m%d")

    st.markdown("#### 📊 Bases y Mensajería")
    _folder_badge("08_BASES_Y_CALIFICACION/")

    # ── ICP status ────────────────────────────────────────────────
    icp_ok = any(est.get(k, "").strip() for k in ["icp_cargos", "icp_industrias", "icp_macro_cargos"])
    if icp_ok:
        st.success("✅ ICP cargado del paso 4 — usando cargos, industrias y prioridades ya definidas")
    else:
        st.warning("⚠️ Aún no hay ICP definido. Ve al tab **🎯 ICP** y completa al menos cargos e industrias.")

    with st.expander("📋 Ver criterios activos del ICP", expanded=False):
        ci1, ci2, ci3 = st.columns(3)
        with ci1:
            st.markdown(f"**👔 Macro cargos:**\n{est.get('icp_macro_cargos','—') or '—'}")
            st.markdown(f"**👔 Cargos:**\n{est.get('icp_cargos','—') or '—'}")
        with ci2:
            st.markdown(f"**🏭 Industrias:**\n{est.get('icp_industrias','—') or '—'}")
            st.markdown(f"**🔍 Micro industrias:**\n{est.get('icp_micro_industrias','—') or '—'}")
        with ci3:
            st.markdown(f"**🌎 Países:** {est.get('icp_paises_foco', est.get('pais_objetivo','—')) or '—'}")
            st.markdown(f"**🚫 Descarte:** {est.get('icp_criterios_descarte','—') or '—'}")
            st.markdown(f"**⛔ Excluir:** {est.get('clientes_actuales','—') or '—'}")

    st.markdown("---")

    # ═════════════════════════════════════════════════════════════
    # BÚSQUEDA DIRECTA EN APIs
    # ═════════════════════════════════════════════════════════════
    st.markdown("""
    <div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:2px;">🔍 Buscar prospectos directamente</div>
    <div style="font-size:12px;color:#64748b;margin-bottom:10px;">
      Usa el ICP del paso 4 para buscar contactos en Apollo.io y Snov.io. Los resultados se acumulan
      en <code>BASES_APOLLO/</code> y <code>BASES_SNOV/</code> para reusar en futuras campañas.
    </div>
    """, unsafe_allow_html=True)

    # Parse ICP fields to lists
    def _plist(txt: str) -> list:
        return [x.strip() for x in txt.replace("\n", ",").split(",") if x.strip()]

    _cargos_icp   = _plist((est.get("icp_cargos","") or "") + "," + (est.get("icp_macro_cargos","") or ""))
    _ind_icp      = _plist((est.get("icp_industrias","") or "") + "," + (est.get("icp_micro_industrias","") or ""))
    _paises_icp   = _plist(est.get("icp_paises_foco", est.get("pais_objetivo","")) or "")
    _kw_icp       = _plist(est.get("icp_keywords","") or "")

    # Estado de credenciales
    _apollo_ok = bool(os.environ.get("APOLLO_API_KEY","").strip())
    _snov_ok   = bool(os.environ.get("SNOV_CLIENT_ID","").strip() and os.environ.get("SNOV_CLIENT_SECRET","").strip())

    # ── GLOBAL paths ────────────────────────────────────────────────
    _base_dir      = path.parent.parent   # ConprospeccionOS/
    _global_apollo = _base_dir / "BASES_APOLLO" / nombre
    _global_snov   = _base_dir / "BASES_SNOV"   / nombre
    _global_apollo.mkdir(parents=True, exist_ok=True)
    _global_snov.mkdir(parents=True, exist_ok=True)
    _ruta_apollo_all = _global_apollo / "apollo_all.csv"
    _ruta_snov_all   = _global_snov   / "snov_all.csv"


    # -- Pool de bases
    import pandas as _pd_pool

    def _pool_stats(ruta):
        if not ruta.exists():
            return 0, 0
        try:
            _df = _pd_pool.read_csv(str(ruta), dtype=str, encoding='utf-8-sig')
            _nem = int(_df['Email'].astype(str).str.contains('@').sum()) if 'Email' in _df.columns else 0
            return len(_df), _nem
        except Exception:
            return 0, 0

    def _detectar_id_col(df):
        for col in ['Apollo ID', 'Snov ID', 'id', 'ID', 'Email', 'email',
                    'Person Linkedin Url', 'LinkedIn URL', 'linkedin_url']:
            if col in df.columns:
                return col
        return df.columns[0] if len(df.columns) > 0 else 'Email'

    _n_ap, _nem_ap = _pool_stats(_ruta_apollo_all)
    _n_sn, _nem_sn = _pool_stats(_ruta_snov_all)

    st.markdown('<div style="font-size:13px;font-weight:700;color:#0f172a;margin-bottom:8px;">Base acumulada de contactos</div>', unsafe_allow_html=True)
    _pc1, _pc2, _pc3, _pc4 = st.columns(4)
    with _pc1:
        st.markdown(f'<div style="background:#eff6ff;color:#1e40af;border-radius:8px;padding:10px;text-align:center;"><div style="font-size:22px;font-weight:800;">{_n_ap:,}</div><div style="font-size:10px;">Apollo (total)</div></div>', unsafe_allow_html=True)
    with _pc2:
        st.markdown(f'<div style="background:#eff6ff;color:#1e40af;border-radius:8px;padding:10px;text-align:center;"><div style="font-size:22px;font-weight:800;">{_nem_ap:,}</div><div style="font-size:10px;">Apollo (con email)</div></div>', unsafe_allow_html=True)
    with _pc3:
        st.markdown(f'<div style="background:#f0fdf4;color:#15803d;border-radius:8px;padding:10px;text-align:center;"><div style="font-size:22px;font-weight:800;">{_n_sn:,}</div><div style="font-size:10px;">Snov (total)</div></div>', unsafe_allow_html=True)
    with _pc4:
        st.markdown(f'<div style="background:#f0fdf4;color:#15803d;border-radius:8px;padding:10px;text-align:center;"><div style="font-size:22px;font-weight:800;">{_nem_sn:,}</div><div style="font-size:10px;">Snov (con email)</div></div>', unsafe_allow_html=True)

    st.markdown('<br>', unsafe_allow_html=True)

    # -- Upload al pool
    _up_col1, _up_col2 = st.columns(2)
    with _up_col1:
        _uf_ap = st.file_uploader('Agregar base Apollo al pool (CSV o Excel)', type=['csv','xlsx'],
                                   key=f'pool_up_apollo_{path.name}')
        if _uf_ap:
            try:
                _dfu = (_pd_pool.read_excel(_uf_ap, dtype=str)
                        if _uf_ap.name.endswith('.xlsx')
                        else _pd_pool.read_csv(_uf_ap, dtype=str, encoding='utf-8-sig', errors='replace'))
                st.caption(f'Archivo leido: {len(_dfu):,} filas | Columnas: {list(_dfu.columns[:6])}')
                _id_col_ap = _detectar_id_col(_dfu)
                _dft, _nnew = _acumular_base(_ruta_apollo_all, _dfu, _id_col_ap)
                st.success(f'{_nnew:,} contactos nuevos agregados. Pool Apollo: {len(_dft):,} total.')
                st.rerun()
            except Exception as _ex:
                st.error(f'Error al leer archivo: {_ex}')
    with _up_col2:
        _uf_sn = st.file_uploader('Agregar base Snov al pool (CSV o Excel)', type=['csv','xlsx'],
                                   key=f'pool_up_snov_{path.name}')
        if _uf_sn:
            try:
                _dfu = (_pd_pool.read_excel(_uf_sn, dtype=str)
                        if _uf_sn.name.endswith('.xlsx')
                        else _pd_pool.read_csv(_uf_sn, dtype=str, encoding='utf-8-sig', errors='replace'))
                st.caption(f'Archivo leido: {len(_dfu):,} filas | Columnas: {list(_dfu.columns[:6])}')
                _id_col_sn = _detectar_id_col(_dfu)
                _dft, _nnew = _acumular_base(_ruta_snov_all, _dfu, _id_col_sn)
                st.success(f'{_nnew:,} contactos nuevos agregados. Pool Snov: {len(_dft):,} total.')
                st.rerun()
            except Exception as _ex:
                st.error(f'Error al leer archivo: {_ex}')

    # Preview del pool actual
    if _ruta_apollo_all.exists() or _ruta_snov_all.exists():
        with st.expander('Vista previa del pool (primeras 10 filas)', expanded=False):
            for _rp, _lbl in [(_ruta_apollo_all, 'Apollo'), (_ruta_snov_all, 'Snov')]:
                if _rp.exists():
                    _dfp = _pd_pool.read_csv(str(_rp), dtype=str, encoding='utf-8-sig', nrows=10)
                    st.caption(f'{_lbl} - columnas: {list(_dfp.columns)}')
                    st.dataframe(_dfp, use_container_width=True, height=150)

    st.markdown('---')

    # -- Seccion 2: Buscar / filtrar en base
    st.markdown('<div style="font-size:13px;font-weight:700;color:#0f172a;margin-bottom:8px;">Filtrar base segun criterios</div>', unsafe_allow_html=True)

    _pk2 = path.name
    _sb_col, _sb_info = st.columns([2, 5])
    with _sb_col:
        if st.button('Cargar filtros desde ICP', key=f'icp_fill_search_{_pk2}', use_container_width=True):
            _f2 = _icp_a_apollo_filtros(est)
            st.session_state[f'sb_cargos_{_pk2}']     = chr(10).join(_cargos_icp[:15])
            st.session_state[f'sb_ind_{_pk2}']        = _f2['industrias']
            st.session_state[f'sb_ind_excl_{_pk2}']   = _f2['ind_excluidas']
            st.session_state[f'sb_paises_{_pk2}']     = ', '.join(_paises_icp)
            st.session_state[f'sb_min_emp_{_pk2}']    = 0
            st.rerun()
    with _sb_info:
        st.caption('Pre-rellena filtros desde el ICP del Paso 4. Los cargos buscan PALABRAS (no frases exactas) - mas flexible.')

    _sf1, _sf2 = st.columns(2)
    with _sf1:
        _sb_cargos = st.text_area(
            'Palabras clave de cargo (uno por linea — busca cada palabra por separado en Title)',
            key=f'sb_cargos_{_pk2}', height=120,
            placeholder='Gerente\nDirector\nSupply Chain\nLogistica\nAbastecimiento\nOperaciones',
            help='Cada linea puede ser una palabra o frase. Busca si CUALQUIER palabra significativa aparece en el cargo.'
        )
    with _sf2:
        _sb_ind = st.multiselect(
            'Industrias a INCLUIR (solo contactos de estas industrias)',
            options=APOLLO_INDUSTRIAS,
            key=f'sb_ind_{_pk2}',
            placeholder='Sin filtro = muestra todas...',
        )

    _sf2b_1, _sf2b_2 = st.columns(2)
    with _sf2b_1:
        _sb_ind_excl = st.multiselect(
            'Industrias a EXCLUIR (elimina estos contactos)',
            options=APOLLO_INDUSTRIAS,
            key=f'sb_ind_excl_{_pk2}',
            placeholder='Ej: Import and Export, Government Administration...',
        )
    with _sf2b_2:
        _sb_paises = st.text_input(
            'Paises (separados por coma, filtra columna Country)',
            key=f'sb_paises_{_pk2}',
            placeholder='Chile, Mexico, Argentina',
        )

    _sf3, _sf4 = st.columns([2, 3])
    with _sf3:
        _sb_min_emp = st.number_input(
            'Empleados minimos (0 = sin filtro)', min_value=0, step=10, value=0,
            key=f'sb_min_emp_{_pk2}',
        )
    with _sf4:
        _sb_fuente = st.selectbox(
            'Fuente', options=['Apollo + Snov', 'Solo Apollo', 'Solo Snov'],
            key=f'sb_fuente_{_pk2}',
        )

    if st.button('Filtrar base', type='primary', key=f'btn_buscar_base_{_pk2}'):
        _kw_raw = [x.strip() for x in (_sb_cargos or '').split(chr(10)) if x.strip()]
        _kw_paises = [x.strip() for x in (_sb_paises or '').split(',') if x.strip()]

        # Palabras significativas (no stop words) para cargo
        _stop = {'de','del','la','el','los','las','y','e','o','en','un','una',
                 'of','the','and','or','in','a','for','to','at','by','con','por'}
        _words = set()
        for kw in _kw_raw:
            for w in kw.lower().split():
                if w not in _stop and len(w) >= 3:
                    _words.add(w)

        _dfs_to_search = []
        if _sb_fuente in ('Apollo + Snov', 'Solo Apollo') and _ruta_apollo_all.exists():
            _dfs_to_search.append(_pd_pool.read_csv(str(_ruta_apollo_all), dtype=str, encoding='utf-8-sig'))
        if _sb_fuente in ('Apollo + Snov', 'Solo Snov') and _ruta_snov_all.exists():
            _dfs_to_search.append(_pd_pool.read_csv(str(_ruta_snov_all), dtype=str, encoding='utf-8-sig'))

        if not _dfs_to_search:
            st.warning('No hay bases en el pool. Sube archivos en la seccion de arriba.')
        else:
            _df_search = (_pd_pool.concat(_dfs_to_search, ignore_index=True)
                          if len(_dfs_to_search) > 1 else _dfs_to_search[0])
            _n_inicio = len(_df_search)
            _mask = _pd_pool.Series([True] * _n_inicio)

            # Filtro cargo: palabra individual en Title
            if _words:
                _title_col = next((c for c in _df_search.columns
                                   if c.lower() in ('title','cargo','position','titulo','job title')), None)
                if _title_col:
                    _mask &= _df_search[_title_col].fillna('').str.lower().apply(
                        lambda v: any(w in v for w in _words)
                    )
                else:
                    st.warning(f'No se encontro columna de cargo. Columnas disponibles: {list(_df_search.columns)}')

            _n_tras_cargo = int(_mask.sum())

            # Filtro industrias INCLUIR
            if _sb_ind:
                _ind_lower = [i.lower() for i in _sb_ind]
                _ind_col = next((c for c in _df_search.columns
                                 if c.lower() in ('industry','industria')), None)
                if _ind_col:
                    _mask &= _df_search[_ind_col].fillna('').str.lower().apply(
                        lambda v: any(ind in v for ind in _ind_lower) if v.strip() else True
                    )

            # Filtro industrias EXCLUIR
            if _sb_ind_excl:
                _excl_lower = [i.lower() for i in _sb_ind_excl]
                _ind_col2 = next((c for c in _df_search.columns
                                  if c.lower() in ('industry','industria')), None)
                if _ind_col2:
                    _mask &= ~_df_search[_ind_col2].fillna('').str.lower().apply(
                        lambda v: any(ex in v for ex in _excl_lower)
                    )

            # Filtro pais
            if _kw_paises:
                _p_lower = [p.lower() for p in _kw_paises]
                _country_col = next((c for c in _df_search.columns
                                     if c.lower() in ('country','pais','location')), None)
                if _country_col:
                    _mask &= _df_search[_country_col].fillna('').str.lower().apply(
                        lambda v: any(p in v for p in _p_lower)
                    )

            # Filtro empleados
            if _sb_min_emp and _sb_min_emp > 0:
                _emp_col = next((c for c in _df_search.columns
                                 if 'employ' in c.lower() or 'empleado' in c.lower()), None)
                if _emp_col:
                    def _emp_ok(v):
                        try:
                            return int(str(v).replace(',','').replace('.','').split('-')[0]) >= _sb_min_emp
                        except Exception:
                            return True
                    _mask &= _df_search[_emp_col].apply(_emp_ok)

            _df_result = _df_search[_mask].reset_index(drop=True)
            st.info(f'Total base: {_n_inicio:,} | Tras filtro cargo: {_n_tras_cargo:,} | Resultado final: {len(_df_result):,}')
            st.session_state[f'sb_result_{_pk2}'] = _df_result

    _df_result = st.session_state.get(f'sb_result_{_pk2}')
    if _df_result is not None:
        if _df_result.empty:
            st.warning('Ningun contacto matcheo los filtros. Proba con menos filtros o palabras mas cortas.')
        else:
            _n_em_r = int(_df_result['Email'].astype(str).str.contains('@').sum()) if 'Email' in _df_result.columns else 0
            st.success(f'{len(_df_result):,} contactos encontrados  --  {_n_em_r:,} con email')
            st.dataframe(_df_result.head(500), use_container_width=True, height=320)
            _rb1, _rb2, _rb3 = st.columns(3)
            with _rb1:
                st.download_button(
                    'Descargar lista para campana',
                    data=_df_result.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig'),
                    file_name=f'campana_{nombre.replace(" ","_")}_{fecha_str}.csv',
                    mime='text/csv', key=f'dl_result_{_pk2}', use_container_width=True,
                )
            with _rb2:
                if st.button('Clasificar con ICP', key=f'btn_clasif_result_{_pk2}',
                              use_container_width=True, type='primary'):
                    st.session_state[f'base_preloaded_{path.name}'] = _df_result
                    st.success('Cargado. Baja a la seccion Calificar.')
            with _rb3:
                st.download_button(
                    'Descargar TODO el pool (sin filtros)',
                    data=_df_search.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig') if '_df_search' in dir() else b'',
                    file_name=f'pool_completo_{nombre.replace(" ","_")}.csv',
                    mime='text/csv', key=f'dl_pool_{_pk2}', use_container_width=True,
                )

    # ── Upload MANUAL ─────────────────────────────────────────────
    # Preload desde búsqueda API si está en session_state
    _df_preload = st.session_state.pop(f"base_preloaded_{path.name}", None)

    st.markdown('<div style="font-size:14px;font-weight:700;color:#0f172a;margin-bottom:4px;">📂 O sube tu base manualmente</div>', unsafe_allow_html=True)
    st.caption("Exporta desde tu herramienta de prospección y sube el archivo aquí. Acepta .xlsx y .csv.")

    archivo = st.file_uploader("Base de contactos", type=["xlsx", "xls", "csv"],
                               label_visibility="collapsed", key="apollo_base_upload")

    if not archivo:
        carpeta_bases = path / "08_BASES_Y_CALIFICACION"
        if carpeta_bases.exists():
            bases = [f for f in sorted(carpeta_bases.iterdir(), reverse=True) if f.suffix in (".xlsx", ".csv")]
            if bases:
                st.markdown("**Bases clasificadas anteriores:**")
                for b in bases[:5]:
                    c1, c2 = st.columns([5, 1])
                    with c1: st.markdown(f'<span class="file-tag">📊 {b.name}</span>', unsafe_allow_html=True)
                    with c2: st.download_button("⬇️", data=b.read_bytes(), file_name=b.name,
                                                key=f"dl_base_{b.name}", use_container_width=True)
        return

    # ── Leer ──────────────────────────────────────────────────────
    try:
        df_raw = pd.read_csv(archivo, encoding="utf-8-sig") if archivo.name.endswith(".csv") \
                 else pd.read_excel(archivo, engine="openpyxl")
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}"); return

    df_raw = df_raw[[c for c in df_raw.columns if not c.startswith("CP_")]]

    col_i1, col_i2 = st.columns([3, 1])
    with col_i1:
        st.success(f"✅ {len(df_raw):,} contactos · {len(df_raw.columns)} columnas")
        st.caption(f"Columnas: {', '.join(df_raw.columns[:8])}{'...' if len(df_raw.columns) > 8 else ''}")
    with col_i2:
        if st.button("👁 Ver muestra", use_container_width=True):
            st.dataframe(df_raw.head(5), use_container_width=True)

    st.markdown("---")

    # ── Calificar ─────────────────────────────────────────────────
    if st.button("⚡ Calificar base con el ICP", type="primary", use_container_width=False):
        with st.spinner(f"Clasificando {len(df_raw):,} contactos..."):
            df_cal = _clasificar_base(df_raw, est)
        st.session_state[f"base_cal_{path.name}"] = df_cal
        st.session_state[f"base_nombre_{path.name}"] = archivo.name

    df_cal = st.session_state.get(f"base_cal_{path.name}")
    if df_cal is None:
        return

    # ── Stats ─────────────────────────────────────────────────────
    tier_counts = df_cal["CP_Tier"].value_counts()
    total = len(df_cal)
    colores_tier = {
        "🟢 Tier A": "#d1fae5", "🟡 Tier B": "#fef3c7",
        "🟠 Tier C": "#fed7aa", "❌ Excluir": "#fee2e2", "⚫ Sin match": "#f1f5f9",
    }
    TIER_DEFS = {
        "🟢 Tier A": ("Prioridad máxima", "Cargo + industria coinciden exactamente con el ICP. Contactar primero."),
        "🟡 Tier B": ("Buena coincidencia", "Cargo o industria coincide parcialmente. Vale la pena contactar."),
        "🟠 Tier C": ("Coincidencia débil", "Poco match con el ICP pero no se descarta. Contactar al final."),
        "❌ Excluir": ("Descartado", "No cumple criterios mínimos o está explícitamente en lista de descarte."),
        "⚫ Sin match": ("Sin datos", "No fue posible clasificar por falta de cargo o industria en la fila."),
    }
    st.markdown("<br>**Resultado de la calificación:**", unsafe_allow_html=True)
    stat_cols = st.columns(5)
    for i, (tier, color) in enumerate(colores_tier.items()):
        n = tier_counts.get(tier, 0)
        subtitulo, tooltip = TIER_DEFS.get(tier, ("",""))
        with stat_cols[i]:
            st.markdown(f"""<div style="background:{color};border-radius:10px;padding:12px;text-align:center;"
                            title="{tooltip}">
              <div style="font-size:22px;font-weight:800;">{n:,}</div>
              <div style="font-size:11px;font-weight:700;color:#374151;">{tier}</div>
              <div style="font-size:10px;color:#475569;margin-top:2px;">{subtitulo}</div>
              <div style="font-size:10px;color:#94a3b8;">{int(n/total*100) if total else 0}%</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    cols_cp = [c for c in df_cal.columns if c.startswith("CP_")]
    cols_key = []
    for nc in ["First Name","Last Name","Title","Company","Industry","# Employees","Country","Email","Phone"]:
        for c in df_cal.columns:
            if c.lower().strip() == nc.lower() and c not in cols_key: cols_key.append(c)
    cols_vista = cols_cp + [c for c in cols_key if c not in cols_cp]
    tier_filtro = st.selectbox("Filtrar:", ["Todos"] + list(colores_tier.keys()), key="filtro_tier_v2")
    df_vista = df_cal if tier_filtro == "Todos" else df_cal[df_cal["CP_Tier"] == tier_filtro]
    st.dataframe(df_vista[cols_vista].head(300), use_container_width=True, height=280)
    if len(df_vista) > 300:
        st.caption(f"Mostrando 300 de {len(df_vista):,}")

    # ── Helpers internos ──────────────────────────────────────────
    def _detectar_cols_base(df):
        cols_l = {c.lower().strip(): c for c in df.columns}
        def _c(*ns):
            for n in ns:
                if n in cols_l: return cols_l[n]
            return None
        return {
            "email":     _c("email","work email","correo","email address"),
            "phone":     _c("phone","phone number","teléfono","telefono","mobile phone","direct phone","work direct phone"),
            "first":     _c("first name","nombre","first_name"),
            "last":      _c("last name","apellido","last_name"),
            "title":     _c("title","job title","cargo","título"),
            "company":   _c("company","company name","empresa","company name for emails"),
            "industry":  _c("industry","industria"),
            "emp":       _c("# employees","employees","employee count","empleados","num employees","company employee count"),
            "linkedin_p":_c("person linkedin url","linkedin url","linkedin","person linkedin"),
            "linkedin_c":_c("company linkedin url","company linkedin","linkedin company"),
            "website":   _c("website","company website","web","company domain","domain"),
            "city":      _c("city","ciudad"),
            "country":   _c("country","país","pais","company country"),
        }

    def _stats_base(df, dc):
        def _n(col):
            if not col or col not in df.columns: return 0
            return int(df[col].apply(lambda x: str(x).strip() not in ("","nan","None","NaN")).sum())
        return len(df), _n(dc["email"]), _n(dc["phone"])

    def _a_excel_col(df_exp) -> bytes:
        from openpyxl.styles import PatternFill, Font
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df_exp.to_excel(writer, index=False, sheet_name="Base")
            ws = writer.sheets["Base"]
            tier_fills = {"🟢 Tier A":"D1FAE5","🟡 Tier B":"FEF3C7",
                          "🟠 Tier C":"FED7AA","❌ Excluir":"FEE2E2","⚫ Sin match":"F1F5F9"}
            tier_col = next((c for c in ["CP_Tier","Tier"] if c in df_exp.columns), None)
            if tier_col:
                ti = df_exp.columns.get_loc(tier_col) + 1
                for r in range(2, len(df_exp) + 2):
                    tv = str(ws.cell(r, ti).value or "")
                    for tk, fc in tier_fills.items():
                        if tk in tv:
                            fill = PatternFill("solid", fgColor=fc)
                            for cc in range(1, len(df_exp.columns)+1):
                                ws.cell(r, cc).fill = fill
                            break
            for cc in range(1, len(df_exp.columns)+1):
                ws.cell(1, cc).font = Font(bold=True)
        return buf.getvalue()

    def _preparar_snov(df_camp, dc) -> "pd.DataFrame":
        cols_orden, rename = [], {}
        for src, dst in [
            (dc["first"],     "Nombre"),
            (dc["last"],      "Apellido"),
            (dc["title"],     "Cargo"),
            (dc["company"],   "Empresa"),
            (dc["industry"],  "Industria"),
            (dc["emp"],       "Tamaño_Empresa"),
            (dc["email"],     "Email"),
            (dc["phone"],     "Teléfono"),
            (dc["linkedin_p"],"LinkedIn_Personal"),
            (dc["linkedin_c"],"LinkedIn_Empresa"),
            (dc["website"],   "Sitio_Web"),
            (dc["city"],      "Ciudad"),
            (dc["country"],   "País"),
        ]:
            if src and src in df_camp.columns:
                cols_orden.append(src)
                rename[src] = dst
        df_out = df_camp[cols_orden].rename(columns=rename).copy()
        df_out["Macro_Industria"] = df_camp["CP_Macro_Industria"].values
        df_out["Micro_Industria"] = df_camp["CP_Micro_Industria"].values
        df_out["Tier"]            = df_camp["CP_Tier"].values
        df_out["Macro_Cargo"]     = df_camp["CP_Macro_Cargo"].values
        return df_out.reset_index(drop=True)

    def _parsear_mensajes(texto: str) -> dict:
        import re as _re
        partes = _re.split(r'##\s*MENSAJE\s*[123][^\n]*\n', texto, flags=_re.IGNORECASE)
        msgs = {"msg1": "", "msg2": "", "msg3": ""}
        if len(partes) >= 4:
            msgs["msg1"] = partes[1].strip()
            msgs["msg2"] = partes[2].strip()
            msgs["msg3"] = partes[3].strip()
        elif len(partes) == 3:
            msgs["msg1"] = partes[1].strip()
            msgs["msg2"] = partes[2].strip()
        else:
            msgs["msg1"] = texto.strip()
        return msgs

    def _build_msgs_system(mc, mi, cargos_enc, n_contactos, estilo):
        resumen   = est.get("resumen_servicio","") or "(no completado)"
        propuesta = est.get("propuesta_valor","") or "(no completado)"
        problema  = est.get("problema_que_resuelve","") or "(no completado)"
        diferenc  = est.get("diferenciacion","") or "(no completado)"
        ticket    = est.get("ticket_promedio","") or "(no especificado)"
        web_c     = est.get("sitio_web","")
        paises_c  = est.get("pais_objetivo","")
        objetivo_c= est.get("objetivo_comercial","")
        return f"""Eres un experto en prospección B2B y copywriting de outreach en frío para la agencia Conprospección.

CLIENTE: {nombre}
Web: {web_c}  |  Países: {paises_c}  |  Objetivo: {objetivo_c}  |  Ticket promedio: {ticket}

CONTEXTO:
- Servicio: {resumen}
- Propuesta de valor: {propuesta}
- Problema que resuelve: {problema}
- Diferenciación: {diferenc}

CAMPAÑA:
- Macro Cargo objetivo: {mc}
- Macro Industria: {mi}
- Cargos específicos en la base: {cargos_enc}
- Total contactos: {n_contactos:,}
- Estilo de mensajes solicitado: {estilo}

INSTRUCCIONES:
- Genera secuencia de 3 emails en frío para esta campaña
- Máximo 100-120 palabras por mensaje
- Incluye Subject: [asunto] al inicio de cada mensaje
- Tono "{estilo}": adapta el lenguaje exactamente a ese estilo
- CTA claro y específico, nunca genérico
- Sin frases vacías ("espero que estés bien", "me permito contactarte")
- Formato EXACTO (sin variaciones):

## MENSAJE 1 — APERTURA
Subject: [asunto]

[cuerpo]

## MENSAJE 2 — FOLLOW-UP
Subject: [asunto]

[cuerpo]

## MENSAJE 3 — CIERRE
Subject: [asunto]

[cuerpo]"""

    # ── Aviso ICP vacío ───────────────────────────────────────────
    icp_vacio = not any(est.get(k,"").strip() for k in ["icp_cargos","icp_macro_cargos","icp_industrias"])
    if icp_vacio:
        st.markdown("""
        <div style="background:#fef3c7;border:1.5px solid #f59e0b;border-radius:10px;padding:12px 16px;margin-bottom:8px;font-size:13px;">
          ⚠️ <b>ICP vacío — por eso todo quedó en Tier C</b><br>
          El algoritmo necesita cargos e industrias del <b>paso 4 (ICP)</b> para asignar Tier A y B.
          Sin esos datos, todos los contactos quedan en Tier C por defecto.<br>
          <span style="color:#78350f;">→ Ve al tab 🎯 ICP, completa los campos estructurados y vuelve a calificar.</span>
        </div>
        """, unsafe_allow_html=True)

    # ── Datos comunes ─────────────────────────────────────────────
    dc = _detectar_cols_base(df_cal)
    # GHL y Snov incluyen Tier A+B+C (no solo A+B) para que siempre haya contactos
    TIERS_ACTIVOS = ["🟢 Tier A", "🟡 Tier B", "🟠 Tier C"]
    df_ghl  = df_cal[df_cal["CP_Tier"].isin(TIERS_ACTIVOS)].copy()
    df_snov_base = df_cal[df_cal["CP_Tier"].isin(TIERS_ACTIVOS)].copy()
    nombre_base = archivo.name.rsplit(".", 1)[0]
    client = get_claude_client()

    def _a_csv(df_exp) -> bytes:
        return df_exp.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

    # ═════════════════════════════════════════════════════════════
    # SECCIÓN 1: SDR / GHL
    # ═════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("""
    <div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:4px;">👥 Archivos para SDRs — CRM</div>
    <div style="font-size:12px;color:#64748b;margin-bottom:10px;">
      Una base CSV por prioridad (Tier). Los SDRs trabajan Prioridad 1 primero, luego 2, luego 3.
    </div>
    """, unsafe_allow_html=True)
    _folder_badge("08_BASES_Y_CALIFICACION/GHL/")

    PRIORIDADES_GHL = [
        ("🟢 Tier A", "Prioridad 1", "#d1fae5", "#065f46"),
        ("🟡 Tier B", "Prioridad 2", "#fef3c7", "#92400e"),
        ("🟠 Tier C", "Prioridad 3", "#fed7aa", "#9a3412"),
    ]

    carpeta_ghl = path / "08_BASES_Y_CALIFICACION" / "GHL"
    carpeta_ghl.mkdir(parents=True, exist_ok=True)

    # Header tabla
    hdr = st.columns([1, 3, 1, 1, 1, 1])
    for h, t in zip(hdr, ["PRIORIDAD", "TIER", "TOTAL", "✉️ EMAIL", "📞 TEL", "CSV"]):
        h.markdown(f'<div style="font-size:10px;color:#94a3b8;font-weight:700;text-align:center;">{t}</div>', unsafe_allow_html=True)

    hay_ghl = False
    for tier_label, prio_label, bg, tc in PRIORIDADES_GHL:
        df_tier = df_cal[df_cal["CP_Tier"] == tier_label].copy()
        if df_tier.empty:
            continue
        hay_ghl = True
        n_t, n_em, n_ph = _stats_base(df_tier, dc)
        c0, c1, c2, c3, c4, c5 = st.columns([1, 3, 1, 1, 1, 1])
        with c0:
            st.markdown(f'<div style="background:{bg};color:{tc};font-size:12px;font-weight:800;text-align:center;border-radius:6px;padding:7px 4px;">{prio_label}</div>', unsafe_allow_html=True)
        with c1:
            st.markdown(f'<div style="background:{bg};color:{tc};font-size:12px;font-weight:600;border-radius:6px;padding:6px 10px;">{tier_label}</div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div style="text-align:center;font-weight:700;font-size:16px;padding-top:4px;">{n_t:,}</div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div style="text-align:center;font-size:14px;color:#10b981;padding-top:6px;">{n_em:,}</div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div style="text-align:center;font-size:14px;color:#3b82f6;padding-top:6px;">{n_ph:,}</div>', unsafe_allow_html=True)
        with c5:
            prio_num = prio_label.split()[-1]
            fname_ghl = f"SDR_P{prio_num}_{tier_label.split()[-1]}_{nombre_base[:15]}_{fecha_str}.csv"
            csv_data = _a_csv(df_tier)
            st.download_button("⬇️ CSV", data=csv_data, file_name=fname_ghl,
                               key=f"dl_ghl_{tier_label}", use_container_width=True, mime="text/csv")
            ruta_ghl_f = carpeta_ghl / fname_ghl
            if not ruta_ghl_f.exists():
                ruta_ghl_f.write_bytes(csv_data)

    if not hay_ghl:
        st.info("No hay contactos Tier A/B/C. Sube una base y califica primero.")

    st.markdown("<br>", unsafe_allow_html=True)
    if not df_ghl.empty:
        n_t_ghl, n_em_ghl, n_ph_ghl = _stats_base(df_ghl, dc)
        cTd1, cTd2 = st.columns([4, 1])
        with cTd1:
            st.markdown(f'<div style="font-size:12px;color:#64748b;">📦 Todos (A+B+C): <b>{n_t_ghl:,}</b> · ✉️ <b>{n_em_ghl:,}</b> · 📞 <b>{n_ph_ghl:,}</b></div>', unsafe_allow_html=True)
        with cTd2:
            st.download_button("⬇️ Todos CSV", data=_a_csv(df_ghl),
                file_name=f"SDR_Todos_{nombre_base[:15]}_{fecha_str}.csv",
                key="dl_sdr_todos", use_container_width=True, mime="text/csv")

    # ═════════════════════════════════════════════════════════════
    # SECCIÓN 2: Snov.io + Mensajes por campaña
    # ═════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("""
    <div style="font-size:15px;font-weight:700;color:#0f172a;margin-bottom:4px;">📧 Campañas para Snov.io + Mensajes</div>
    <div style="font-size:12px;color:#64748b;margin-bottom:10px;">
      Máximo 3 campañas por Macro Cargo + Macro Industria. Cada una incluye el Excel para Snov y la
      secuencia de 3 mensajes generada con Claude.
    </div>
    """, unsafe_allow_html=True)
    _folder_badge("08_BASES_Y_CALIFICACION/Snov/  ·  05_MENSAJERIA_COMERCIAL/")

    if df_snov_base.empty:
        st.info("No hay contactos calificados para campañas Snov.io.")
        st.markdown("---")
    else:
        combos = (df_snov_base.groupby(["CP_Macro_Cargo","CP_Macro_Industria"])
                  .size().reset_index(name="n")
                  .sort_values("n", ascending=False)
                  .head(6))

        ESTILOS_MSG = [
            "Profesional directo",
            "Desde el dolor / problema",
            "Consultivo",
            "Challenger / Provocador",
            "Persuasivo con prueba social",
            "Basado en datos",
            "Asesor de confianza",
            "Casual y cercano",
            "Storytelling",
            "AIDA (Atención → Interés → Deseo → Acción)",
        ]

        max_camp = min(5, len(combos))
        if max_camp > 1:
            n_campanas = st.slider("¿Cuántas campañas quieres generar?", 1, max_camp, min(max_camp, 3), key="n_campanas_snov")
        else:
            n_campanas = 1
            st.markdown('<div style="font-size:12px;color:#64748b;padding-top:8px;">📋 1 campaña disponible</div>', unsafe_allow_html=True)

        if not client:
            st.markdown("""
            <div style="background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;padding:10px 14px;margin:6px 0;font-size:12px;">
              ⚠️ <b>API key de Claude no configurada</b> — generación de mensajes no disponible.
              Configura <code>ANTHROPIC_API_KEY</code> en el archivo <code>.env</code> y reinicia la app.
            </div>
            """, unsafe_allow_html=True)

        for idx, (_, row_c) in enumerate(combos.head(n_campanas).iterrows()):
            mc = row_c["CP_Macro_Cargo"]
            mi = row_c["CP_Macro_Industria"]
            df_camp = df_snov_base[(df_snov_base["CP_Macro_Cargo"] == mc) & (df_snov_base["CP_Macro_Industria"] == mi)].copy()
            n_t, n_em, n_ph = _stats_base(df_camp, dc)

            # Cargos específicos en este segmento
            col_title = dc.get("title")
            cargos_enc = "—"
            if col_title and col_title in df_camp.columns:
                top_c = df_camp[col_title].value_counts().head(5)
                cargos_enc = ", ".join(top_c.index.tolist())

            chat_key  = f"snov_chat_{path.name}_{idx}"
            input_key = f"snov_input_{path.name}_{idx}"
            msgs_key  = f"snov_msgs_{path.name}_{idx}"
            mk1 = f"snov_m1_{path.name}_{idx}"
            mk2 = f"snov_m2_{path.name}_{idx}"
            mk3 = f"snov_m3_{path.name}_{idx}"

            for k, v in [(chat_key, []), (input_key, ""), (msgs_key, {"msg1":"","msg2":"","msg3":""})]:
                if k not in st.session_state:
                    st.session_state[k] = v
            for mk in [mk1, mk2, mk3]:
                if mk not in st.session_state:
                    st.session_state[mk] = ""

            # Aplicar actualizaciones pendientes del chat ANTES de renderizar los text_area
            for mk, pk in [(mk1, f"_pend_{mk1}"), (mk2, f"_pend_{mk2}"), (mk3, f"_pend_{mk3}")]:
                if pk in st.session_state:
                    st.session_state[mk] = st.session_state.pop(pk)

            with st.expander(f"📧 Campaña {idx+1}: {mc} · {mi}  ({n_t:,} contactos)", expanded=True):

                # ── Resumen de la campaña ─────────────────────────
                rCol1, rCol2 = st.columns([3, 1])
                with rCol1:
                    st.markdown(f"""
                    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;padding:14px 18px;margin-bottom:8px;">
                      <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:8px;">
                        <span style="background:#eff6ff;color:#1e40af;font-size:12px;font-weight:700;padding:3px 10px;border-radius:6px;">👔 {mc}</span>
                        <span style="background:#f0fdf4;color:#15803d;font-size:12px;font-weight:700;padding:3px 10px;border-radius:6px;">🏭 {mi}</span>
                      </div>
                      <div style="font-size:12px;color:#475569;margin-bottom:4px;"><b>Cargos en esta base:</b> {cargos_enc}</div>
                      <div style="font-size:12px;color:#475569;">
                        👥 <b>{n_t:,}</b> contactos &nbsp;·&nbsp; ✉️ <b>{n_em:,}</b> con email &nbsp;·&nbsp; 📞 <b>{n_ph:,}</b> con teléfono
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
                with rCol2:
                    sdr_nombre = st.text_input("SDR", key=f"snov_sdr_{idx}",
                                               placeholder="Nombre SDR", label_visibility="collapsed")
                    df_snov = _preparar_snov(df_camp, dc)
                    if sdr_nombre.strip():
                        df_snov.insert(0, "SDR_Asignado", sdr_nombre.strip())
                    mc_safe = mc.replace(' ','_').replace('/','_')
                    fname_snov = f"Snov_C{idx+1}_{mc_safe}_{fecha_str}.xlsx"
                    excel_snov = _a_excel_col(df_snov)
                    st.download_button("⬇️ Excel Snov.io",
                        data=excel_snov, file_name=fname_snov,
                        key=f"dl_snov_{idx}", use_container_width=True,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    # Guardar en carpeta del cliente
                    carpeta_snov = path / "08_BASES_Y_CALIFICACION" / "Snov"
                    carpeta_snov.mkdir(parents=True, exist_ok=True)
                    ruta_snov = carpeta_snov / fname_snov
                    if not ruta_snov.exists():
                        ruta_snov.write_bytes(excel_snov)

                # ── Estilo por campaña + botón generar ────────────
                estilo_key = f"snov_estilo_{path.name}_{idx}"
                col_est2, col_gen = st.columns([3, 2])
                with col_est2:
                    estilo = st.selectbox("🎨 Estilo de mensajes", ESTILOS_MSG,
                                          key=estilo_key, label_visibility="visible")
                with col_gen:
                    st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                    generar_btn = st.button(f"🤖 Generar mensajes", key=f"gen_msgs_{idx}",
                                            disabled=not client, use_container_width=True)

                if generar_btn:
                    sys_p = _build_msgs_system(mc, mi, cargos_enc, n_t, estilo)
                    with st.spinner("Claude generando mensajes..."):
                        resp = llamar_claude(client,
                            [{"role":"user","content":f"Genera los 3 mensajes. Campaña: {mc} en {mi}. Estilo: {estilo}."}],
                            sys_p)
                    parsed = _parsear_mensajes(resp)
                    st.session_state[msgs_key] = parsed
                    # Usar _pend_ para evitar conflicto con widgets ya instanciados
                    st.session_state[f"_pend_{mk1}"] = parsed["msg1"]
                    st.session_state[f"_pend_{mk2}"] = parsed["msg2"]
                    st.session_state[f"_pend_{mk3}"] = parsed["msg3"]
                    st.rerun()

                m1c, m2c, m3c = st.columns(3)
                with m1c:
                    st.markdown('<div style="background:#dbeafe;border-radius:6px;padding:5px 10px;font-size:11px;font-weight:700;color:#1e40af;margin-bottom:3px;">📧 MENSAJE 1 — APERTURA</div>', unsafe_allow_html=True)
                    msg1 = st.text_area("M1", key=mk1, height=190, label_visibility="collapsed",
                                        placeholder="Haz clic en 'Generar mensajes' o escribe directamente aquí...")
                with m2c:
                    st.markdown('<div style="background:#d1fae5;border-radius:6px;padding:5px 10px;font-size:11px;font-weight:700;color:#065f46;margin-bottom:3px;">📧 MENSAJE 2 — FOLLOW-UP</div>', unsafe_allow_html=True)
                    msg2 = st.text_area("M2", key=mk2, height=190, label_visibility="collapsed",
                                        placeholder="Se genera junto al mensaje 1...")
                with m3c:
                    st.markdown('<div style="background:#fef3c7;border-radius:6px;padding:5px 10px;font-size:11px;font-weight:700;color:#92400e;margin-bottom:3px;">📧 MENSAJE 3 — CIERRE</div>', unsafe_allow_html=True)
                    msg3 = st.text_area("M3", key=mk3, height=190, label_visibility="collapsed",
                                        placeholder="Se genera junto al mensaje 1...")

                cs1, cs2 = st.columns([2, 1])
                with cs1:
                    if st.button("💾 Guardar mensajes", key=f"save_msgs_{idx}", use_container_width=True):
                        fecha = datetime.now().strftime("%Y-%m-%d")
                        carpeta_msg = path / "05_MENSAJERIA_COMERCIAL"
                        carpeta_msg.mkdir(parents=True, exist_ok=True)
                        mc_safe = mc.replace(' ','_').replace('/','_')
                        contenido_md = (
                            f"# Mensajes Campaña {idx+1}: {mc} · {mi}\n\n"
                            f"**Fecha:** {fecha}  |  **Estilo:** {estilo}\n\n---\n\n"
                            f"## MENSAJE 1 — APERTURA\n\n{msg1}\n\n"
                            f"## MENSAJE 2 — FOLLOW-UP\n\n{msg2}\n\n"
                            f"## MENSAJE 3 — CIERRE\n\n{msg3}\n"
                        )
                        fname_md = f"mensajes_c{idx+1}_{mc_safe}.md"
                        (carpeta_msg / fname_md).write_text(contenido_md, encoding="utf-8")
                        st.success(f"✅ Guardado en 05_MENSAJERIA_COMERCIAL/{fname_md}")
                        _folder_badge(f"05_MENSAJERIA_COMERCIAL/{fname_md}")
                with cs2:
                    if msg1 or msg2 or msg3:
                        dl_txt = f"CAMPAÑA {idx+1}: {mc} · {mi}\nEstilo: {estilo}\n\n---\nMENSAJE 1\n{msg1}\n\n---\nMENSAJE 2\n{msg2}\n\n---\nMENSAJE 3\n{msg3}"
                        st.download_button("⬇️ TXT mensajes", data=dl_txt.encode(),
                            file_name=f"mensajes_c{idx+1}.txt", key=f"dl_msgs_{idx}",
                            use_container_width=True)

                # ── Chat refinamiento ─────────────────────────────
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("""
                <div style="background:#f0f9ff;border:1.5px solid #7dd3fc;border-radius:10px;padding:10px 14px;margin-bottom:8px;">
                  <div style="font-size:13px;font-weight:700;color:#0369a1;margin-bottom:2px;">💬 Refinar mensajes con chat</div>
                  <div style="font-size:11px;color:#475569;">
                    Escribe qué quieres cambiar y Claude actualiza los 3 recuadros automáticamente.<br>
                    Ej: <em>"Cambia el asunto del 1"</em> · <em>"El mensaje 2 muy largo, acórtalo"</em> · <em>"El 3 más directo y agresivo"</em> · <em>"Rehaz el 1 desde el dolor"</em>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                chat_h = st.session_state[chat_key]
                if chat_h:
                    for msg_c in chat_h[-8:]:
                        if msg_c["role"] == "user":
                            st.markdown(f'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:7px 12px;margin:3px 0;margin-left:15%;font-size:12px;"><b style="color:#1e40af;">TÚ</b><br>{msg_c["content"]}</div>', unsafe_allow_html=True)
                        else:
                            st.markdown(f'<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:7px 12px;margin:3px 0;margin-right:5%;font-size:12px;"><b style="color:#15803d;">✅ Actualizado</b><br><div style="white-space:pre-wrap;color:#374151;">{msg_c["content"]}</div></div>', unsafe_allow_html=True)

                chat_ic, chat_is = st.columns([5, 1])
                with chat_ic:
                    chat_input = st.text_input(f"chat_c{idx}", key=input_key,
                        label_visibility="collapsed",
                        placeholder='¿Qué quieres cambiar? Ej: "Cambia el asunto del 1" · "El 2 más corto" · "Rehaz el 3 desde el dolor"')
                with chat_is:
                    enviar_c = st.button("Enviar →", key=f"send_chat_{idx}", disabled=not client, use_container_width=True)
                    if chat_h:
                        if st.button("🗑 Limpiar", key=f"clear_chat_{idx}", use_container_width=True):
                            st.session_state[chat_key] = []
                            st.rerun()

                if enviar_c and chat_input.strip() and client:
                    sys_chat = (
                        _build_msgs_system(mc, mi, cargos_enc, n_t, estilo)
                        + "\n\nIMPORTANTE: Cuando el usuario pida cambios, SIEMPRE devuelve los 3 mensajes completos "
                        "con el formato exacto:\n## MENSAJE 1 — APERTURA\n[texto]\n\n## MENSAJE 2 — FOLLOW-UP\n[texto]\n\n## MENSAJE 3 — CIERRE\n[texto]\n"
                        "Aunque solo pida cambiar uno, devuelve los 3 (los que no cambian, cópialos igual). "
                        "Así los recuadros se actualizan automáticamente."
                    )
                    ctx_msgs = (
                        f"Mensajes actuales:\n\n"
                        f"## MENSAJE 1 — APERTURA\n{msg1}\n\n"
                        f"## MENSAJE 2 — FOLLOW-UP\n{msg2}\n\n"
                        f"## MENSAJE 3 — CIERRE\n{msg3}"
                    )
                    api_msgs = (
                        [{"role":"user","content": ctx_msgs},
                         {"role":"assistant","content":"Entendido, tengo los 3 mensajes actuales. ¿Qué quieres cambiar?"}]
                        + [{"role": m["role"], "content": m["content"]} for m in chat_h]
                        + [{"role":"user","content": chat_input.strip()}]
                    )
                    with st.spinner("Claude refinando..."):
                        resp_c = llamar_claude(client, api_msgs, sys_chat)
                    # Guardar pendientes ANTES del rerun — se aplican al inicio del próximo ciclo
                    parsed_c = _parsear_mensajes(resp_c)
                    if parsed_c.get("msg1"):
                        st.session_state[f"_pend_{mk1}"] = parsed_c["msg1"]
                    if parsed_c.get("msg2"):
                        st.session_state[f"_pend_{mk2}"] = parsed_c["msg2"]
                    if parsed_c.get("msg3"):
                        st.session_state[f"_pend_{mk3}"] = parsed_c["msg3"]
                    resumen_chat = f"Cambios aplicados según: \"{chat_input.strip()}\""
                    st.session_state[chat_key].append({"role":"user","content":chat_input.strip()})
                    st.session_state[chat_key].append({"role":"assistant","content":resumen_chat})
                    st.session_state[input_key] = ""
                    st.rerun()

    # ═════════════════════════════════════════════════════════════
    # Base completa calificada
    st.markdown("---")
    st.markdown('<div style="font-size:14px;font-weight:700;color:#0f172a;margin-bottom:8px;">📥 Base completa calificada</div>', unsafe_allow_html=True)
    col_dl1, col_dl2 = st.columns(2)
    with col_dl1:
        nombre_base_local = st.session_state.get(f"base_nombre_{path.name}", "base").rsplit(".", 1)[0]
        st.download_button("⬇️ Todos los contactos (con columnas CP_)",
            data=_a_excel_col(df_cal),
            file_name=f"{nombre_base_local}_calificada_{fecha_str}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary")
    with col_dl2:
        df_excl = df_cal[df_cal["CP_Tier"] == "❌ Excluir"]
        if not df_excl.empty:
            st.download_button(f"❌ Solo excluidos ({len(df_excl):,}) — para verificar",
                data=_a_excel_col(df_excl),
                file_name=f"{nombre_base_local}_EXCLUIDOS_{fecha_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    carpeta_bases_f = path / "08_BASES_Y_CALIFICACION"
    carpeta_bases_f.mkdir(parents=True, exist_ok=True)
    ruta_g = carpeta_bases_f / f"{nombre_base_local}_calificada_{fecha_str}.xlsx"
    if not ruta_g.exists():
        ruta_g.write_bytes(_a_excel_col(df_cal))
    _folder_badge(f"08_BASES_Y_CALIFICACION/{ruta_g.name}")


# ─────────────────────────────────────────────
# TAB: MENSAJERÍA
# ─────────────────────────────────────────────
def tab_mensajeria(path: Path, est: dict):
    st.markdown("#### 📨 Mensajería Comercial")
    _folder_badge("05_MENSAJERIA_COMERCIAL/")
    st.info("⏳ Completa el análisis web y el ICP antes de trabajar la mensajería.")

    carpeta = path / "05_MENSAJERIA_COMERCIAL"
    archivos = sorted(carpeta.iterdir()) if carpeta.exists() else []
    archivos_md = [f for f in archivos if f.is_file()]

    if archivos_md:
        archivo_sel = st.selectbox("Ver archivo", [f.name for f in archivos_md])
        f_sel = carpeta / archivo_sel
        if f_sel.exists():
            contenido = f_sel.read_text(encoding="utf-8")
            nuevo = st.text_area(archivo_sel, value=contenido, height=350, key=f"msg_{archivo_sel}")
            col1, col2 = st.columns(2)
            with col1:
                if st.button("💾 Guardar cambios"):
                    f_sel.write_text(nuevo, encoding="utf-8")
                    st.success("Guardado")
                    _folder_badge(f"05_MENSAJERIA_COMERCIAL/{archivo_sel}")
            with col2:
                st.download_button("⬇️ Descargar", data=contenido.encode(), file_name=archivo_sel)


# ─────────────────────────────────────────────
# TAB: PLAYBOOK SDR
# ─────────────────────────────────────────────
def _playbook_datos(est: dict) -> dict:
    """Extrae datos del cliente para el playbook."""
    return {
        "nombre":   est.get("nombre_cliente", ""),
        "web":      est.get("sitio_web", ""),
        "paises":   est.get("pais_objetivo", ""),
        "industria":est.get("icp_industrias","") or est.get("icp_macro_industrias",""),
        "servicio": est.get("resumen_servicio","") or est.get("descripcion_empresa",""),
        "propuesta":est.get("propuesta_valor",""),
        "diferenc": est.get("diferenciacion",""),
        "problema": est.get("problema_que_resuelve",""),
        "ticket":   est.get("ticket_promedio",""),
        "cargos":   est.get("icp_cargos",""),
        "mc_cargos":est.get("icp_macro_cargos",""),
        "mc_ind":   est.get("icp_industrias",""),
        "proceso":  est.get("proceso_comercial",""),
        "objetivo": est.get("objetivo_comercial",""),
        "color":    est.get("color_marca","#1a56db"),
        "color2":   est.get("color_secundario","#64748b"),
    }


def _build_playbook_prompt(est: dict, logo_b64: str = "", idioma: str = "Español", notas: str = "") -> str:
    """Prompt PARTE 1: genera <!DOCTYPE> … hasta sección 10, SIN cerrar </body></html>."""
    d = _playbook_datos(est)
    nombre = d["nombre"];  color = d["color"];  color2 = d["color2"]
    logo_tag = (f'<img src="{logo_b64}" style="max-height:60px;object-fit:contain;" alt="{nombre}" />'
                if logo_b64 else f'<b style="font-size:20px;color:{color};">{nombre}</b>')
    notas_extra = f"\nNotas: {notas}" if notas.strip() else ""

    return f"""Eres un consultor comercial B2B senior. Generas la PARTE 1 de un playbook en HTML.

REGLAS CRÍTICAS:
- Devuelve SOLO HTML puro. Sin markdown, sin bloques ```, sin texto fuera del HTML.
- Empieza exactamente con: <!DOCTYPE html>
- NO cierres </body> ni </html> — esto es PARTE 1, la PARTE 2 continuará.
- CSS completo en <style> en el <head> (Inter font, fondo #f8fafc, color primario {color}).
- Secciones compactas: H2 + bullets cortos. Sin párrafos largos.
- Usa "(Inf.)" para inferencias.

CLIENTE: {nombre} | {d['web']} | {d['paises']}
SERVICIO: {d['servicio'][:300]}
PROPUESTA: {d['propuesta'][:200]}
ICP CARGOS: {d['mc_cargos']} / {d['cargos'][:150]}
ICP INDUSTRIAS: {d['mc_ind'][:150]}
DOLOR: {d['problema'][:200]}
DIFERENCIADORES: {d['diferenc'][:200]}
OBJETIVO: {d['objetivo'][:150]}
COLOR: {color} / {color2} | IDIOMA: {idioma}{notas_extra}

ESTRUCTURA PARTE 1 (secciones 1-10):
1. PORTADA — logo: {logo_tag} | "Playbook Comercial · {nombre}" | "powered by Conprospección" | fondo oscuro {color}
2. ÍNDICE — lista con anclas a #s3…#s12
3. RESUMEN EJECUTIVO — qué hace, a quién vende, problema que resuelve, cómo usar este playbook
4. ICP COMPLETO — tipo empresa, tamaño, rubro, geografía, señales fit/no-fit, tabla resumen
5. BUYER PERSONAS — por cada cargo: dolor, motivación, objeción típica, cómo abordarlo
6. PROPUESTA DE VALOR — mensaje principal, 3 beneficios tangibles, diferenciadores vs alternativas
7. PROBLEMAS Y PAINS — 5-8 dolores específicos del mercado, impacto, cómo conecta la solución
8. MENSAJERÍA — mensaje paraguas, hooks de apertura por canal, frases de impacto, qué NO decir
9. OBJECIONES — tabla: objeción | significado real | respuesta corta | respuesta estratégica
10. PROCESO COMERCIAL — etapas, objetivo de cada una, señales de avance, errores típicos
"""


def _build_playbook_prompt_2(est: dict, idioma: str = "Español") -> str:
    """Prompt PARTE 2: secciones 11-16 + cierre </body></html>."""
    d = _playbook_datos(est)
    nombre = d["nombre"];  color = d["color"]

    return f"""Eres un consultor comercial B2B senior. Generas la PARTE 2 de un playbook en HTML.

REGLAS CRÍTICAS:
- Devuelve SOLO HTML puro. Sin markdown, sin ```, sin texto fuera del HTML.
- Empieza directamente con el contenido HTML (sin <!DOCTYPE>, sin <html>, sin <head>).
- Cierra al final con </main></body></html>.
- Usa el mismo estilo visual que la PARTE 1: color {color}, fondo #f8fafc, fuente Inter.
- Secciones compactas: H2 + bullets. Sin párrafos largos.

CLIENTE: {nombre} | COLOR: {color} | IDIOMA: {idioma}
CARGOS ICP: {d['mc_cargos']} / {d['cargos'][:150]}
SERVICIO: {d['servicio'][:200]}

ESTRUCTURA PARTE 2 (secciones 11-16):
11. SCRIPTS COMERCIALES — email apertura, email follow-up, WhatsApp apertura, WhatsApp follow-up, LinkedIn mensaje. Cada uno: asunto/gancho + cuerpo completo listo para usar.
12. DISCOVERY — 10 preguntas clave por etapa (contexto, dolor, urgencia, decisión). Señales de oportunidad real vs pérdida de tiempo.
13. CALIFICACIÓN — criterios lead calificado vs no. Checklist de 8 puntos. Cuándo avanzar, cuándo descartar.
14. CIERRE Y SIGUIENTE PASO — cómo pedir reunión, confirmar interés, dejar fecha, resumir valor.
15. CHECKLIST OPERATIVO — antes de contactar / durante / después / antes de handoff. Formato checklist.
16. KPIs SUGERIDOS — tabla: métrica | meta referencial | cómo medirla. Al menos 8 KPIs de prospección.
"""


def _netlify_deploy(html: str, nombre_cliente: str, token: str, site_id: str = None) -> dict:
    """Sube el HTML a Netlify y devuelve {'url': ..., 'site_id': ...}."""
    import requests, zipfile, io, re as _re
    slug = _re.sub(r'[^a-z0-9]+', '-', nombre_cliente.lower()).strip('-')
    slug = f"playbook-{slug}-cp"[:63]  # Netlify max 63 chars

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('index.html', html)
    zip_bytes = buf.getvalue()

    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/zip'}

    if site_id:
        r = requests.post(f"https://api.netlify.com/api/v1/sites/{site_id}/deploys",
                          headers=headers, data=zip_bytes, timeout=60)
    else:
        r = requests.post("https://api.netlify.com/api/v1/sites",
                          headers=headers, params={'name': slug}, data=zip_bytes, timeout=60)

    if r.status_code not in (200, 201):
        raise RuntimeError(f"Netlify {r.status_code}: {r.text[:300]}")

    data = r.json()
    return {
        'url': data.get('ssl_url') or data.get('url', ''),
        'site_id': data.get('site_id') or data.get('id', ''),
    }


def _limpiar_html_claude(raw: str) -> str:
    """Limpia la respuesta de Claude para extraer solo el HTML."""
    raw = raw.strip()
    if "```" in raw:
        partes = raw.split("```")
        for p in partes:
            if p.strip().startswith("html"):
                return p.strip()[4:].strip()
            if p.strip().startswith("<!DOCTYPE") or p.strip().startswith("<html"):
                return p.strip()
    if raw.startswith("<!DOCTYPE") or raw.startswith("<html"):
        return raw
    # Buscar inicio de HTML en cualquier posición
    idx = raw.find("<!DOCTYPE")
    if idx == -1:
        idx = raw.find("<html")
    if idx >= 0:
        return raw[idx:]
    return raw


def tab_playbook(path: Path, est: dict):
    import json as _json
    nombre = est.get("nombre_cliente", "")
    st.markdown("#### 📋 Playbook SDR")
    _folder_badge("06_PLAYBOOK_SDR/")

    client = get_claude_client()
    playbook_file = path / "06_PLAYBOOK_SDR" / "playbook_sdr.html"
    datos_file    = path / "06_PLAYBOOK_SDR" / "playbook_meta.json"
    (path / "06_PLAYBOOK_SDR").mkdir(parents=True, exist_ok=True)

    # ── Estado: sin generar / borrador / aprobado ─────────────────────────────
    # pb_borrador_html = HTML generado por Claude, pendiente de aprobación
    # playbook_file    = aprobado y guardado en disco

    pb_borrador_html = st.session_state.get("pb_borrador_html")

    playbook_aprobado = playbook_file.exists() and playbook_file.stat().st_size > 100

    def _guardar_meta(html: str):
        """Guarda metadatos del playbook en JSON para tracking/Supabase futuro."""
        import datetime
        meta = {
            "cliente": nombre,
            "fecha": datetime.date.today().isoformat(),
            "archivo": str(playbook_file),
            "tamanio_bytes": len(html.encode("utf-8")),
            "estado": "aprobado",
        }
        datos_file.write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        # También guardar en carpeta 07_BASE_DATOS para futura sync a Supabase
        db_dir = path / "07_BASE_DATOS"
        db_dir.mkdir(parents=True, exist_ok=True)
        (db_dir / "playbook_log.json").write_text(_json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")

    # ─────────────────────────────────────────────────────────────────────────
    # VISTA: sin playbook todavía
    # ─────────────────────────────────────────────────────────────────────────
    if not pb_borrador_html and not playbook_aprobado:
        st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("""
        <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:10px;padding:14px 18px;margin-bottom:16px;">
          <div style="font-size:13px;font-weight:700;color:#0369a1;margin-bottom:4px;">📋 Playbook Comercial Premium — 20 secciones</div>
          <div style="font-size:12px;color:#0c4a6e;line-height:1.6;">
            Genera un playbook completo y listo para entregar: ICP, buyer personas, propuesta de valor, mensajería,
            scripts de prospección, manejo de objeciones, proceso comercial, KPIs y más.<br>
            <b>Tiempo estimado de generación: 45–90 segundos.</b>
          </div>
        </div>
        """, unsafe_allow_html=True)

        cc1, cc2 = st.columns([2, 3])
        with cc1:
            idioma = st.selectbox("Idioma del playbook", ["Español", "Inglés", "Portugués"], key="pb_idioma")
        with cc2:
            notas = st.text_input("Instrucción adicional (opcional)",
                placeholder="Ej: foco en empresas medianas, agrega sección de WhatsApp, tono muy directo…",
                key="pb_notas")

        st.markdown("<br>", unsafe_allow_html=True)

        if not client:
            st.warning("Configura ANTHROPIC_API_KEY en .env para habilitar la generación.")
        elif st.button("🤖 Generar Playbook Completo", type="primary", use_container_width=False, key="pb_generar"):
            logos = listar_logos(path)
            logo_b64 = imagen_b64(logos[0]["path"]) if logos else ""

            pb_html_draft = ""
            error_pb = ""

            # — PARTE 1: portada + secciones 1-10 —
            prompt1 = _build_playbook_prompt(est, logo_b64, idioma, notas)
            with st.spinner("📝 Parte 1/2 — generando portada, ICP, propuesta, mensajería... (30–50 seg)"):
                raw1 = llamar_claude(client,
                    [{"role": "user", "content": "Genera la PARTE 1 del playbook en HTML. Recuerda: NO cierres </body></html>."}],
                    prompt1, max_tokens=8000)

            if raw1.startswith("❌"):
                error_pb = raw1
            else:
                parte1 = _limpiar_html_claude(raw1)
                # Quitar cierre accidental en parte 1
                for tag in ["</body>", "</html>"]:
                    parte1 = parte1.replace(tag, "")

                # — PARTE 2: scripts, discovery, checklists, KPIs —
                import time as _time_pb
                _time_pb.sleep(10)  # pausa para no superar el límite de tokens/minuto
                prompt2 = _build_playbook_prompt_2(est, idioma)
                with st.spinner("📝 Parte 2/2 — generando scripts, checklists, KPIs... (30–50 seg)"):
                    raw2 = llamar_claude(client,
                        [{"role": "user", "content": "Genera la PARTE 2 del playbook en HTML. Empieza con contenido, cierra con </main></body></html>."}],
                        prompt2, max_tokens=8000)

                if raw2.startswith("❌"):
                    error_pb = raw2
                else:
                    parte2 = raw2.strip()
                    # Si parte2 tiene DOCTYPE/html, extraer solo el body content
                    if "<!DOCTYPE" in parte2 or "<html" in parte2:
                        import re as _re2
                        m = _re2.search(r'<body[^>]*>(.*)', parte2, _re2.DOTALL | _re2.IGNORECASE)
                        if m:
                            parte2 = m.group(1)
                    pb_html_draft = parte1 + "\n" + parte2
                    # Asegurar cierre correcto
                    if not pb_html_draft.rstrip().endswith("</html>"):
                        pb_html_draft += "\n</body></html>"

            if error_pb:
                st.error(f"Error de API: {error_pb}")
            elif len(pb_html_draft) < 1000:
                st.error("El playbook generado está vacío o demasiado corto. Intenta de nuevo.")
                st.code(pb_html_draft[:300] or raw1[:300])
            else:
                st.session_state["pb_borrador_html"] = pb_html_draft
                st.rerun()

    # ─────────────────────────────────────────────────────────────────────────
    # VISTA: borrador generado — revisar y aprobar
    # ─────────────────────────────────────────────────────────────────────────
    elif pb_borrador_html and not playbook_aprobado:
        st.markdown("""
        <div style="background:#fef3c7;border:1.5px solid #f59e0b;border-radius:10px;padding:12px 18px;margin-bottom:12px;">
          <b style="color:#92400e;">👁️ Revisa el playbook antes de aprobarlo</b><br>
          <span style="font-size:12px;color:#78350f;">Cuando estés conforme haz clic en <b>✅ Aprobar y guardar</b>.
          Si quieres cambios, escríbelos en el chat de abajo — el playbook se actualiza completo.</span>
        </div>
        """, unsafe_allow_html=True)

        ba1, ba2, ba3 = st.columns(3)
        with ba1:
            if st.button("✅ Aprobar y guardar", type="primary", use_container_width=True, key="pb_aprobar"):
                playbook_file.write_text(pb_borrador_html, encoding="utf-8")
                _guardar_meta(pb_borrador_html)
                st.session_state.pop("pb_borrador_html", None)
                st.session_state.pop("pb_chat", None)
                st.toast("✅ Playbook aprobado y guardado", icon="📋")
                ir_a_tab(6)   # Avanza a Firma
        with ba2:
            if st.button("🔄 Regenerar desde cero", use_container_width=True, key="pb_regen2"):
                st.session_state.pop("pb_borrador_html", None)
                st.session_state.pop("pb_chat", None)
                st.rerun()
        with ba3:
            st.download_button("⬇️ Descargar borrador",
                data=pb_borrador_html.encode("utf-8"),
                file_name=f"BORRADOR_Playbook_{nombre.replace(' ','_')}.html",
                mime="text/html", use_container_width=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.components.v1.html(pb_borrador_html, height=620, scrolling=True)

        # Chat para pedir cambios — actualiza el HTML completo
        st.markdown("---")
        st.markdown("""
        <div style="font-size:13px;font-weight:700;color:#0f172a;margin-bottom:4px;">💬 Pide cambios al playbook</div>
        <div style="font-size:12px;color:#64748b;margin-bottom:8px;">
          El playbook completo se actualizará con tu solicitud.<br>
          Ej: "Agrega un script de WhatsApp" · "Las objeciones son muy largas, acórtalas" · "El ICP no incluye empresas de 50-200 empleados"
        </div>
        """, unsafe_allow_html=True)

        if "pb_chat" not in st.session_state:
            st.session_state["pb_chat"] = []

        for m in st.session_state["pb_chat"][-6:]:
            if m["role"] == "user":
                st.markdown(f'<div style="background:#eff6ff;border-radius:8px;padding:8px 12px;margin:3px 0;margin-left:20%;font-size:12px;"><b style="color:#1e40af;">TÚ</b><br>{m["content"]}</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:8px 12px;margin:3px 0;margin-right:5%;font-size:12px;"><b style="color:#15803d;">✅ Playbook actualizado</b></div>', unsafe_allow_html=True)

        ci, cs = st.columns([5, 1])
        with ci:
            chat_pb = st.text_input("", key="pb_chat_input", label_visibility="collapsed",
                placeholder='Escribe qué quieres cambiar o agregar…')
        with cs:
            if st.button("Enviar →", key="pb_chat_send", disabled=not client):
                if chat_pb.strip():
                    sys_chat = """Eres el autor de este playbook comercial premium. El usuario pide modificaciones.
Devuelve ÚNICAMENTE el HTML completo actualizado (<!DOCTYPE html>...</html>). Sin texto antes ni después.
Mantén el mismo estilo visual, estructura y calidad del original. Aplica solo el cambio solicitado."""
                    with st.spinner("Actualizando playbook..."):
                        raw2 = llamar_claude(client,
                            [{"role": "user", "content": f"HTML actual del playbook:\n\n{pb_borrador_html[:6000]}\n\n[...resto del documento...]\n\nCambio solicitado: {chat_pb.strip()}"}],
                            sys_chat, max_tokens=8000)
                    nuevo_html = _limpiar_html_claude(raw2)
                    if len(nuevo_html) > 500:
                        st.session_state["pb_borrador_html"] = nuevo_html
                        st.session_state["pb_chat"].append({"role": "user", "content": chat_pb.strip()})
                        st.session_state["pb_chat"].append({"role": "assistant", "content": "aplicado"})
                        st.rerun()
                    else:
                        st.error("La respuesta fue incompleta. Intenta de nuevo.")

    # ─────────────────────────────────────────────────────────────────────────
    # VISTA: playbook aprobado y guardado
    # ─────────────────────────────────────────────────────────────────────────
    elif playbook_aprobado:
        pb_html = playbook_file.read_text(encoding="utf-8")

        st.markdown("""
        <div style="background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;padding:10px 14px;margin-bottom:12px;font-size:12px;color:#166534;">
          ✅ Playbook aprobado y guardado en <b>06_PLAYBOOK_SDR/playbook_sdr.html</b>
        </div>
        """, unsafe_allow_html=True)

        dl1, dl2, dl3 = st.columns(3)
        with dl1:
            st.download_button("⬇️ Descargar HTML",
                data=pb_html.encode("utf-8"),
                file_name=f"Playbook_Comercial_{nombre.replace(' ','_')}.html",
                mime="text/html", use_container_width=True, type="primary")
        with dl2:
            if st.button("🔄 Crear nueva versión", use_container_width=True, key="pb_nueva"):
                playbook_file.unlink(missing_ok=True)
                datos_file.unlink(missing_ok=True)
                st.session_state.pop("pb_chat", None)
                st.session_state.pop("pb_netlify_url", None)
                st.rerun()
        with dl3:
            netlify_token = os.getenv("NETLIFY_TOKEN", "").strip()
            if not netlify_token:
                st.markdown('<div style="font-size:11px;color:#94a3b8;padding-top:8px;">Agrega NETLIFY_TOKEN en .env para publicar link</div>', unsafe_allow_html=True)
            else:
                if st.button("🌐 Publicar link para SDR", use_container_width=True, key="pb_netlify"):
                    # Recuperar site_id previo si existe
                    site_id = None
                    if datos_file.exists():
                        try:
                            meta = __import__('json').loads(datos_file.read_text(encoding="utf-8"))
                            site_id = meta.get("netlify_site_id")
                        except Exception:
                            pass
                    with st.spinner("Publicando en Netlify..."):
                        try:
                            result = _netlify_deploy(pb_html, nombre, netlify_token, site_id)
                            st.session_state["pb_netlify_url"] = result["url"]
                            # Guardar site_id para actualizaciones futuras
                            if datos_file.exists():
                                try:
                                    import json as _j
                                    meta = _j.loads(datos_file.read_text(encoding="utf-8"))
                                    meta["netlify_site_id"] = result["site_id"]
                                    meta["netlify_url"] = result["url"]
                                    datos_file.write_text(_j.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
                                except Exception:
                                    pass
                        except Exception as e:
                            st.error(f"Error Netlify: {e}")

        # Mostrar link Netlify si existe
        netlify_url = st.session_state.get("pb_netlify_url") or (
            __import__('json').loads(datos_file.read_text(encoding="utf-8")).get("netlify_url", "")
            if datos_file.exists() else ""
        )
        if netlify_url:
            st.markdown(f"""
            <div style="background:#f0fdf4;border:1.5px solid #22c55e;border-radius:10px;padding:14px 18px;margin:12px 0;">
              <div style="font-size:11px;font-weight:700;color:#15803d;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:6px;">
                🌐 Link público del playbook
              </div>
              <a href="{netlify_url}" target="_blank" style="font-size:14px;font-weight:600;color:#1a56db;word-break:break-all;">{netlify_url}</a>
              <div style="font-size:11px;color:#64748b;margin-top:6px;">
                Copia y pega este link para enviarlo a los SDR. Siempre abre la versión más reciente.
              </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.components.v1.html(pb_html, height=640, scrolling=True)


# ─────────────────────────────────────────────
# TAB: FIRMA
# ─────────────────────────────────────────────
def tab_firma(path: Path, est: dict):
    from modules.firma import PLANTILLAS
    nombre_cliente = est.get("nombre_cliente", "")
    st.markdown("#### ✍️ Generador de Firma de Email")
    _folder_badge("01_ADMIN_CLIENTE/firma_email.html  ·  02_BRANDING_Y_ACTIVOS/firma_email.html")

    firma_existente = path / "01_ADMIN_CLIENTE/firma_email.html"
    tiene_firma = firma_existente.exists() and firma_existente.stat().st_size > 100

    # ── Selector de plantilla (ancho completo arriba) ─────────────────────────
    st.markdown("##### 🎨 Elige una plantilla")
    tpl_cols = st.columns(len(PLANTILLAS))
    nombres_tpl = list(PLANTILLAS.keys())
    tpl_iconos  = ["📋", "⬛", "🌈", "✂️", "👤"]
    tpl_sel_key = "firma_tpl_sel"
    if tpl_sel_key not in st.session_state:
        st.session_state[tpl_sel_key] = nombres_tpl[0]

    for i, (col_t, nombre_t, ico) in enumerate(zip(tpl_cols, nombres_tpl, tpl_iconos)):
        with col_t:
            activa = st.session_state[tpl_sel_key] == nombre_t
            bg = "#1a56db" if activa else "#f1f5f9"
            tc = "#fff"   if activa else "#374151"
            st.markdown(
                f'<div style="background:{bg};color:{tc};border-radius:8px;padding:8px 6px;'
                f'text-align:center;font-size:12px;font-weight:{"700" if activa else "500"};'
                f'cursor:pointer;border:2px solid {"#1a56db" if activa else "#e2e8f0"};">'
                f'{ico}<br>{nombre_t}</div>',
                unsafe_allow_html=True
            )
            if st.button("Seleccionar", key=f"tpl_btn_{i}", use_container_width=True):
                st.session_state[tpl_sel_key] = nombre_t
                st.rerun()

    plantilla_activa = st.session_state[tpl_sel_key]
    st.markdown("---")

    # ── Campos (izquierda) — se leen antes de las columnas para poder
    #    generar la vista previa en vivo en la columna derecha ───────────────
    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown("##### 👤 Datos personales")
        f_nombre   = st.text_input("Nombre completo *", placeholder="Ej: María González", value=est.get("nombre_prospector", ""))
        f_cargo    = st.text_input("Cargo *", placeholder="Ej: Business Development Manager", value=est.get("cargo_prospector", ""))
        f_empresa  = st.text_input("Empresa", placeholder="Ej: Conprospección", value=nombre_cliente)
        f_correo   = st.text_input("Email", placeholder="maria@empresa.com", value=est.get("correo", ""))

        c_tel, c_mov = st.columns(2)
        with c_tel:
            f_telefono = st.text_input("Teléfono oficina", placeholder="+56 2 1234 5678", value=est.get("telefono", ""))
        with c_mov:
            f_movil    = st.text_input("Móvil / WhatsApp", placeholder="+56 9 8765 4321", value=est.get("movil", ""))

        f_web      = st.text_input("Sitio web", placeholder="empresa.com", value=est.get("sitio_web", ""))
        f_linkedin = st.text_input("LinkedIn", placeholder="linkedin.com/in/...", value=est.get("linkedin", ""))
        f_twitter  = st.text_input("Twitter / X (opcional)", placeholder="twitter.com/...", value=est.get("twitter", ""))
        f_dir      = st.text_input("Dirección (opcional)", placeholder="Av. Providencia 123, Santiago", value=est.get("direccion_firma", ""))

        st.markdown("##### 🎨 Colores")
        cc1, cc2 = st.columns(2)
        with cc1:
            f_color1 = st.color_picker("Color principal", value=est.get("color_marca", "#1a56db"))
        with cc2:
            f_color2 = st.color_picker("Color secundario", value=est.get("color_secundario", "#64748b"))

        st.markdown("##### 🖼️ Logo")
        logos = listar_logos(path)
        logo_path = None

        logo_upload = st.file_uploader("Sube el logo de la empresa", type=["png", "jpg", "jpeg", "svg", "webp"], key="logo_upload_firma")
        if logo_upload:
            res = guardar_archivo(path, logo_upload.name, logo_upload.getvalue())
            logo_path = res["guardado_en"]
            st.success("✅ Logo cargado")

        if not logo_path and logos:
            opciones = ["Sin logo"] + [l["nombre"] for l in logos]
            logo_sel = st.selectbox("O selecciona logo existente", opciones, key="logo_sel_firma")
            if logo_sel != "Sin logo":
                logo_path = next((l["path"] for l in logos if l["nombre"] == logo_sel), None)

        if logo_path:
            b64 = imagen_b64(logo_path)
            if b64:
                st.markdown(f'<img src="{b64}" style="max-height:45px;max-width:180px;border:1px solid #e2e8f0;border-radius:6px;padding:5px;margin-top:4px;">', unsafe_allow_html=True)

    # ── Vista previa en VIVO (se recalcula en cada interacción) ───────────────
    datos_preview = {
        "nombre": f_nombre or "Tu Nombre", "cargo": f_cargo or "Tu Cargo",
        "empresa": f_empresa, "email": f_correo, "telefono": f_telefono,
        "movil": f_movil, "web": f_web, "linkedin": f_linkedin,
        "twitter": f_twitter, "direccion": f_dir,
        "color_marca": f_color1, "color_secundario": f_color2,
        "nombre_cliente": nombre_cliente,
        "nombre_prospector": f_nombre or "Tu Nombre", "cargo_prospector": f_cargo or "Tu Cargo",
    }
    preview_html = generar_firma_html(datos_preview, logo_path, plantilla_activa)

    with col2:
        st.markdown("##### 👁️ Vista previa en vivo")
        st.markdown('<div style="border:1.5px solid #e2e8f0;border-radius:10px;padding:20px;background:#ffffff;min-height:160px;">', unsafe_allow_html=True)
        st.components.v1.html(preview_html, height=240, scrolling=False)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Botón Guardar firma (arriba en col2, sin bajar) ───────────────
        if st.button("💾 Guardar firma", type="primary", use_container_width=True, key="guardar_firma_top"):
            if not f_nombre or not f_cargo:
                st.error("El nombre y cargo son obligatorios.")
            else:
                firma_final = generar_firma_html(datos_preview, logo_path, plantilla_activa)
                (path / "01_ADMIN_CLIENTE/firma_email.html").write_text(firma_final, encoding="utf-8")
                (path / "02_BRANDING_Y_ACTIVOS/firma_email.html").write_text(firma_final, encoding="utf-8")
                st.session_state.firma_html = firma_final
                st.session_state.pop("firma_png", None)
                est_actual = cargar_estado(path)
                est_actual.update({
                    "nombre_prospector": f_nombre, "cargo_prospector": f_cargo,
                    "correo": f_correo, "telefono": f_telefono, "movil": f_movil,
                    "linkedin": f_linkedin, "twitter": f_twitter,
                    "direccion_firma": f_dir, "color_marca": f_color1,
                    "color_secundario": f_color2, "estado_firma": "generado",
                })
                guardar_estado(path, est_actual)
                st.success("✅ Firma guardada")

        firma_guardada = st.session_state.get("firma_html") or (
            firma_existente.read_text(encoding="utf-8") if tiene_firma else None
        )

        if firma_guardada:
            st.markdown("---")
            dl1, dl2, dl3 = st.columns(3)
            with dl1:
                st.download_button(
                    "⬇️ HTML",
                    data=firma_guardada.encode("utf-8"),
                    file_name=f"firma_{nombre_cliente.lower().replace(' ', '_')}.html",
                    mime="text/html", use_container_width=True,
                )
            with dl2:
                if st.button("🖼️ Generar PNG", use_container_width=True, key="btn_png_firma"):
                    try:
                        import tempfile, os
                        from html2image import Html2Image
                        with tempfile.TemporaryDirectory() as tmpdir:
                            hti = Html2Image(output_path=tmpdir,
                                             custom_flags=["--no-sandbox","--disable-gpu","--hide-scrollbars"])
                            hti.screenshot(html_str=firma_guardada, save_as="firma.png", size=(620, 280))
                            png_bytes = open(os.path.join(tmpdir, "firma.png"), "rb").read()
                        st.session_state["firma_png"] = png_bytes
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error PNG: {e}")
                if st.session_state.get("firma_png"):
                    st.download_button(
                        "⬇️ Guardar PNG",
                        data=st.session_state["firma_png"],
                        file_name=f"firma_{nombre_cliente.lower().replace(' ', '_')}.png",
                        mime="image/png", use_container_width=True, key="dl_png_firma",
                    )
            with dl3:
                if st.button("📋 Código HTML", use_container_width=True):
                    st.code(firma_guardada, language="html")

            st.markdown("---")
            st.caption("Pega este código en tu cliente de email (Gmail, Outlook, etc.):")
            st.text_area("", value=firma_guardada, height=130, key="firma_code_view")


# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# ─────────────────────────────────────────────
# TAB: GHL SETUP
# ─────────────────────────────────────────────
def _ghl_request(method: str, endpoint: str, token: str, data: dict = None) -> dict:
    import requests
    base = "https://services.leadconnectorhq.com"
    headers = {
        "Authorization": f"Bearer {token}",
        "Version": "2021-07-28",
        "Content-Type": "application/json",
    }
    url = f"{base}{endpoint}"

    r = requests.request(method, url, headers=headers, json=data, timeout=20)
    if r.status_code not in (200, 201):
        raise RuntimeError(f"GHL {r.status_code}: {r.text[:300]}")
    return r.json()


def _ghl_build_email_prompt(tipo: str, est: dict) -> str:
    nombre   = est.get("nombre_cliente", "")
    web      = est.get("sitio_web", "")
    desc     = est.get("resumen_servicio", "") or est.get("descripcion_empresa", "")
    propuesta= est.get("propuesta_valor", "")
    problema = est.get("problema_que_resuelve", "")
    cargos   = est.get("icp_macro_cargos", "")
    paises   = est.get("pais_objetivo", "")

    contexto = f"""Cliente: {nombre} | Web: {web} | Países: {paises}
Servicio: {desc}
Propuesta de valor: {propuesta}
Problema que resuelve: {problema}
Cargos objetivo: {cargos}"""

    merge_fields = """MERGE FIELDS disponibles en GHL (úsalos donde corresponda):
- {{contact.first_name}} → nombre del prospecto
- {{contact.full_name}} → nombre completo del prospecto
- {{contact.company_name}} → empresa del prospecto
- {{appointment.start_time}} → fecha y hora de la reunión
- {{appointment.title}} → título de la reunión
- {{user.name}} → nombre del SDR que agendó"""

    tipos = {
        "no_responde": {
            "asunto": f"Seguimiento — reunión con {nombre}",
            "instruccion": f"""Escribe un email de seguimiento cuando el prospecto NO ha respondido después de que el SDR lo contactó.
Tono: directo, sin presionar, profesional. Máximo 4 líneas de texto.
Menciona brevemente qué hace {nombre} y el valor que puede aportar.
Incluye una llamada a la acción clara (¿te acomoda reagendar?).
NO uses frases de relleno como "espero que estés bien".""",
        },
        "info_adicional": {
            "asunto": f"Información sobre {nombre} — lo que pediste",
            "instruccion": f"""Escribe un email que se envía cuando el prospecto pide más información sobre {nombre}.
Tono: consultivo, útil, sin sonar a folleto.
Resume en 3-4 bullets lo más relevante del servicio.
Invita a tener una llamada corta para resolver sus dudas específicas.
Termina con los datos de contacto (usa merge fields).""",
        },
        "derivar": {
            "asunto": f"Tu reunión con {nombre} — próximo paso",
            "instruccion": f"""Escribe un email que se envía cuando el prospecto es derivado/referido a una reunión de cierre o demo con {nombre}.
Tono: cálido, profesional, que genere confianza.
Confirma que recibirá una llamada / reunión con el equipo de {nombre}.
Menciona brevemente qué van a conversar y qué valor va a obtener el prospecto.
Deja claro el siguiente paso concreto.""",
        },
    }

    info = tipos[tipo]
    return f"""Eres un experto en copywriting B2B para ventas outbound.
Genera un email profesional para automatización en Go High Level.

TIPO: {tipo.replace('_', ' ').title()}
ASUNTO SUGERIDO: {info['asunto']}

INSTRUCCIÓN:
{info['instruccion']}

CONTEXTO DEL CLIENTE:
{contexto}

{merge_fields}

FORMATO DE RESPUESTA — devuelve EXACTAMENTE este JSON (sin markdown):
{{
  "asunto": "asunto del email",
  "cuerpo": "cuerpo completo del email en HTML simple (usa <p>, <b>, <br> solo si es necesario, sin CSS inline)",
  "notas": "una línea con el momento ideal para enviar este email en la automatización"
}}"""


def tab_ghl(path: Path, est: dict):
    st.markdown("#### 🔗 GHL Setup")
    _folder_badge("07_BASE_DATOS/ghl_config.json  ·  05_MENSAJERIA_COMERCIAL/ghl_emails/")

    nombre = est.get("nombre_cliente", "")
    client = get_claude_client()

    db_dir = path / "07_BASE_DATOS"
    db_dir.mkdir(parents=True, exist_ok=True)
    ghl_file = db_dir / "ghl_config.json"
    emails_dir = path / "05_MENSAJERIA_COMERCIAL" / "ghl_emails"
    emails_dir.mkdir(parents=True, exist_ok=True)

    ghl_cfg = {}
    if ghl_file.exists():
        try:
            ghl_cfg = json.loads(ghl_file.read_text(encoding="utf-8"))
        except Exception:
            ghl_cfg = {}

    # ── Credenciales ───────────────────────────────────────────────────────
    st.markdown("##### Credenciales de la subcuenta")
    st.markdown("""
    <div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;padding:10px 14px;font-size:12px;color:#0369a1;margin-bottom:10px;">
      <b>API Key:</b> Settings → Integrations → API Key (en la subcuenta)<br>
      <b>Location ID:</b> Settings → Business Info → Location ID
    </div>""", unsafe_allow_html=True)

    gc1, gc2 = st.columns(2)
    with gc1:
        api_key = st.text_input("API Key", value=ghl_cfg.get("api_key", ""),
                                type="password", key="ghl_api_key", placeholder="pit-...")
    with gc2:
        location_id = st.text_input("Location ID", value=ghl_cfg.get("location_id", ""),
                                    key="ghl_location_id", placeholder="abc123XYZ...")

    nombre_anterior = st.text_input(
        "Nombre del cliente anterior en el snapshot",
        value=ghl_cfg.get("nombre_anterior", ""),
        key="ghl_nombre_anterior",
        placeholder="Ej: BambuTech, Cliente Demo…",
        help="Se reemplazará por el nombre del cliente actual en los email templates")

    if st.button("💾 Guardar credenciales", key="ghl_save_creds"):
        ghl_cfg.update({"api_key": api_key, "location_id": location_id, "nombre_anterior": nombre_anterior})
        ghl_file.write_text(json.dumps(ghl_cfg, ensure_ascii=False, indent=2), encoding="utf-8")
        st.toast("✅ Credenciales guardadas", icon="🔑")
        st.rerun()

    # ── SDRs asignados ─────────────────────────────────────────────
    sdrs = ghl_cfg.get("sdrs", [])
    if sdrs:
        st.markdown("##### 👥 SDRs asignados a esta cuenta")
        sdr_cols = st.columns(len(sdrs))
        for i, sdr in enumerate(sdrs):
            with sdr_cols[i]:
                st.markdown(f"""<div style="background:#f0f9ff;border:1px solid #bae6fd;border-radius:8px;
                    padding:10px 12px;font-size:12px;color:#0369a1;text-align:center;">
                  <div style="font-weight:700;font-size:13px;color:#0f172a;">{sdr['nombre']}</div>
                  <div style="color:#64748b;margin-top:3px;">{sdr['email']}</div>
                  <div style="color:#64748b;">{sdr.get('telefono','')}</div>
                  <div style="font-size:10px;color:#94a3b8;margin-top:4px;font-family:monospace;">ID: {sdr['ghl_user_id']}</div>
                </div>""", unsafe_allow_html=True)
    else:
        st.info("No hay SDRs configurados. Edita directamente `07_BASE_DATOS/ghl_config.json` para agregarlos.")

    tiene_creds = bool(api_key and location_id)
    st.markdown("---")

    # ── SECCIÓN 1: Emails automáticos generados por Claude ─────────────────
    st.markdown("##### 1. Emails automáticos de seguimiento")
    st.markdown("""
    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px 14px;font-size:12px;color:#475569;margin-bottom:10px;">
      Claude genera los 3 emails de seguimiento usando los datos del cliente.
      Puedes revisar y editar antes de subir a GHL.<br>
      Usan merge fields de GHL: <code>{{contact.first_name}}</code>, <code>{{appointment.start_time}}</code>, etc.
    </div>""", unsafe_allow_html=True)

    TIPOS_EMAIL = {
        "no_responde":   ("📭 No responde",      "Seguimiento cuando el prospecto no contestó"),
        "info_adicional":("📎 Info adicional",   "Cuando el prospecto pide más información"),
        "derivar":       ("🤝 Derivar / Cierre", "Cuando se pasa a reunión de cierre o demo"),
    }

    if not client:
        st.warning("Configura ANTHROPIC_API_KEY en .env para generar los emails.")
    else:
        cols_e = st.columns(3)
        for (tipo, (label, desc)), col in zip(TIPOS_EMAIL.items(), cols_e):
            with col:
                archivo = emails_dir / f"{tipo}.json"
                tiene = archivo.exists()
                st.markdown(
                    f'<div style="background:{"#f0fdf4" if tiene else "#f8fafc"};border:1.5px solid '
                    f'{"#22c55e" if tiene else "#e2e8f0"};border-radius:10px;padding:12px;text-align:center;'
                    f'margin-bottom:8px;">'
                    f'<div style="font-size:16px;">{"✅" if tiene else "⬜"}</div>'
                    f'<div style="font-size:13px;font-weight:700;margin:4px 0;">{label}</div>'
                    f'<div style="font-size:11px;color:#64748b;">{desc}</div></div>',
                    unsafe_allow_html=True)
                if st.button(f"{'🔄 Regenerar' if tiene else '✨ Generar'}", key=f"ghl_gen_{tipo}", use_container_width=True):
                    prompt = _ghl_build_email_prompt(tipo, est)
                    with st.spinner(f"Generando {label}..."):
                        raw = llamar_claude(client,
                            [{"role": "user", "content": f"Genera el email de tipo {tipo} para {nombre}."}],
                            prompt, max_tokens=1500)
                    try:
                        raw = raw.strip()
                        if raw.startswith("```"):
                            raw = raw.split("```")[1]
                            if raw.startswith("json"): raw = raw[4:]
                        data = json.loads(raw.strip())
                        data["tipo"] = tipo
                        data["cliente"] = nombre
                        data["fecha_generacion"] = datetime.now().isoformat()
                        archivo.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
                        # Log en base de datos
                        log_file = db_dir / "ghl_emails_log.json"
                        log = {}
                        if log_file.exists():
                            try: log = json.loads(log_file.read_text(encoding="utf-8"))
                            except: log = {}
                        log[tipo] = {"asunto": data.get("asunto",""), "notas": data.get("notas",""),
                                     "fecha": data["fecha_generacion"], "cliente": nombre}
                        log_file.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
                        st.success(f"✅ Guardado en ghl_emails/{tipo}.json y 07_BASE_DATOS/ghl_emails_log.json")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error: {e}")
                        st.code(raw[:300])

        # Mostrar y editar los emails generados
        generados = [t for t in TIPOS_EMAIL if (emails_dir / f"{t}.json").exists()]
        if generados:
            st.markdown("<br>", unsafe_allow_html=True)
            sel_tipo = st.selectbox("Ver / editar email:",
                                    options=generados,
                                    format_func=lambda t: TIPOS_EMAIL[t][0],
                                    key="ghl_sel_email")
            data = json.loads((emails_dir / f"{sel_tipo}.json").read_text(encoding="utf-8"))

            ea1, ea2 = st.columns([1, 1])
            with ea1:
                nuevo_asunto = st.text_input("Asunto", value=data.get("asunto", ""), key="ghl_asunto")
            with ea2:
                st.markdown(f'<div style="font-size:11px;color:#64748b;padding-top:28px;">💡 {data.get("notas","")}</div>', unsafe_allow_html=True)

            nuevo_cuerpo = st.text_area("Cuerpo del email (HTML simple)",
                                        value=data.get("cuerpo", ""),
                                        height=200, key="ghl_cuerpo")

            eb1, eb2, eb3 = st.columns(3)
            with eb1:
                if st.button("💾 Guardar cambios", key="ghl_save_email"):
                    data.update({"asunto": nuevo_asunto, "cuerpo": nuevo_cuerpo})
                    (emails_dir / f"{sel_tipo}.json").write_text(
                        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
                    st.success("✅ Guardado.")
                    st.rerun()
            with eb2:
                st.download_button("⬇️ Descargar JSON",
                                   data=json.dumps(data, ensure_ascii=False, indent=2).encode(),
                                   file_name=f"ghl_email_{sel_tipo}.json",
                                   mime="application/json", key="ghl_dl_email")
            with eb3:
                if tiene_creds:
                    if st.button("🚀 Subir a GHL", key="ghl_upload_email", type="primary"):
                        try:
                            _ghl_request("POST", f"/locations/{location_id}/templates",
                                         api_key, {
                                             "name": f"{nombre} — {TIPOS_EMAIL[sel_tipo][0]}",
                                             "type": "email",
                                             "subject": nuevo_asunto,
                                             "body": nuevo_cuerpo,
                                         })
                            st.success("✅ Template subido a GHL.")
                        except Exception as e:
                            st.error(str(e))
                else:
                    st.markdown('<div style="font-size:11px;color:#94a3b8;padding-top:8px;">Agrega credenciales para subir a GHL</div>', unsafe_allow_html=True)

    # ── SECCIÓN 2: Calendarios ─────────────────────────────────────────────
    st.markdown("---")
    st.markdown("##### 2. Calendarios")

    if not tiene_creds:
        st.markdown('<div style="color:#94a3b8;font-size:13px;">Agrega las credenciales arriba para gestionar calendarios.</div>', unsafe_allow_html=True)
    else:
        if st.button("🔄 Leer calendarios desde GHL", key="ghl_fetch_cal"):
            try:
                resp = _ghl_request("GET", f"/calendars/?locationId={location_id}", api_key)
                st.session_state["ghl_calendarios"] = resp.get("calendars", [])
                st.success(f"✅ {len(st.session_state['ghl_calendarios'])} calendarios.")
            except Exception as e:
                st.error(f"Error: {e}")

        cals = st.session_state.get("ghl_calendarios", [])
        if cals:
            for cal in cals:
                cal_id   = cal.get("id", "")
                cal_name = cal.get("name", "")
                sugerido = cal_name.replace(nombre_anterior, nombre) if nombre_anterior and nombre_anterior in cal_name else cal_name
                cc1, cc2, cc3 = st.columns([3, 4, 1])
                with cc1:
                    st.markdown(f'<div style="font-size:12px;color:#64748b;padding-top:10px;">Actual: <b>{cal_name}</b></div>', unsafe_allow_html=True)
                with cc2:
                    nuevo_nombre = st.text_input("", value=sugerido, key=f"ghl_cal_{cal_id}",
                                                 label_visibility="collapsed")
                with cc3:
                    if st.button("✓", key=f"ghl_cal_upd_{cal_id}", help="Actualizar en GHL"):
                        try:
                            _ghl_request("PUT", f"/calendars/{cal_id}", api_key, {"name": nuevo_nombre})
                            st.success(f"✅ '{nuevo_nombre}'")
                        except Exception as e:
                            st.error(str(e))


# TAB: COMERCIAL
# ─────────────────────────────────────────────
def tab_comercial(path: Path, est: dict):
    import datetime as dt

    st.markdown("#### 💼 Datos Comerciales")
    _folder_badge("07_BASE_DATOS/comercial.json")

    nombre = est.get("nombre_cliente", "")
    color  = est.get("color_marca", "#1a56db")

    db_dir = path / "07_BASE_DATOS"
    db_dir.mkdir(parents=True, exist_ok=True)
    comercial_file = db_dir / "comercial.json"

    com = {}
    if comercial_file.exists():
        try:
            com = json.loads(comercial_file.read_text(encoding="utf-8"))
        except Exception:
            com = {}

    # ── Contrato ───────────────────────────────────────────────────────────
    st.markdown("##### Contrato")
    r1c1, r1c2, r1c3, r1c4 = st.columns(4)
    with r1c1:
        moneda = st.selectbox("Moneda", ["CLP", "USD"],
                              index=0 if com.get("moneda", "CLP") == "CLP" else 1,
                              key="com_moneda")
    with r1c2:
        paso = 10_000 if moneda == "CLP" else 100
        monto_fijo = st.number_input(f"Fijo mensual ({moneda})",
                                     min_value=0, step=paso,
                                     value=int(com.get("monto_fijo", 0)),
                                     key="com_monto")
    with r1c3:
        meses = st.number_input("Meses de contrato", min_value=1, max_value=24,
                                value=int(com.get("meses_contrato", 5)),
                                key="com_meses")
    with r1c4:
        semanas_setup = st.number_input("Semanas de setup", min_value=1, max_value=8,
                                        value=int(com.get("semanas_setup", 3)),
                                        key="com_setup")

    r2c1, r2c2 = st.columns(2)
    with r2c1:
        fecha_str = com.get("fecha_inicio_contrato", dt.date.today().isoformat())
        try:
            fecha_default = dt.date.fromisoformat(fecha_str)
        except Exception:
            fecha_default = dt.date.today()
        fecha_inicio = st.date_input("Fecha de inicio del contrato", value=fecha_default, key="com_fecha")
    with r2c2:
        notas = st.text_area("Notas del contrato",
                             value=com.get("notas", ""),
                             height=68,
                             placeholder="Ej: pago día 1 de cada mes, incluye 4 reuniones/mes en el fijo…",
                             key="com_notas")

    # ── Reuniones ──────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("##### Reuniones")
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        reu_garantizadas = st.number_input("Reuniones garantizadas (total contrato)",
                                           min_value=0,
                                           value=int(com.get("reuniones_garantizadas", 0)),
                                           key="com_reu_gar")
    with m2:
        reu_incluidas_mes = st.number_input("Incluidas en el fijo (por mes)",
                                            min_value=0,
                                            value=int(com.get("reuniones_incluidas_fijo", 0)),
                                            help="Reuniones que ya están cubiertas por el monto fijo mensual",
                                            key="com_reu_incl")
    with m3:
        meta_mes = st.number_input("Meta de reuniones por mes",
                                   min_value=0,
                                   value=int(com.get("meta_reuniones_mes", 0)),
                                   key="com_meta_mes")
    with m4:
        paso_reu = 1_000 if moneda == "CLP" else 5
        costo_reu = st.number_input(f"Costo por reunión adicional ({moneda})",
                                    min_value=0, step=paso_reu,
                                    value=int(com.get("costo_reunion", 0)),
                                    key="com_costo_reu")

    # ── Cálculos automáticos ───────────────────────────────────────────────
    fecha_fin_setup        = fecha_inicio + dt.timedelta(weeks=int(semanas_setup))
    fecha_inicio_prosp     = fecha_fin_setup
    fecha_fin_contrato     = fecha_inicio_prosp + dt.timedelta(days=int(meses) * 30)
    semanas_efectivas      = int(meses) * 4
    meta_semana            = round(meta_mes / 4.33, 1) if meta_mes else 0
    ingreso_fijo_total     = monto_fijo * int(meses)
    reu_incluidas_total    = reu_incluidas_mes * int(meses)
    reu_adicionales_total  = max(0, reu_garantizadas - reu_incluidas_total)
    ingreso_variable       = reu_adicionales_total * costo_reu
    ingreso_total          = ingreso_fijo_total + ingreso_variable
    ticket_por_reunion     = round(ingreso_total / reu_garantizadas, 0) if reu_garantizadas else 0

    # ── Timeline visual ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("##### Proyección del contrato")

    def _fecha_box(label, fecha, bg, tc, desc=""):
        return (f'<div style="text-align:center;background:{bg};border-radius:10px;'
                f'padding:12px 10px;min-width:120px;">'
                f'<div style="font-size:10px;font-weight:700;color:{tc};text-transform:uppercase;'
                f'letter-spacing:0.4px;">{label}</div>'
                f'<div style="font-size:15px;font-weight:800;color:{tc};margin:4px 0;">'
                f'{fecha.strftime("%d/%m/%Y")}</div>'
                f'{"<div style=font-size:10px;color:" + tc + ";opacity:0.7;>" + desc + "</div>" if desc else ""}'
                f'</div>')

    arrow = '<div style="font-size:22px;color:#cbd5e1;display:flex;align-items:center;padding:0 6px;">→</div>'
    st.markdown(f"""
    <div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;
                padding:18px 20px;display:flex;align-items:center;gap:4px;flex-wrap:wrap;">
      {_fecha_box("Inicio contrato", fecha_inicio, "#dbeafe", "#1e40af", "Firma")}
      {arrow}
      {_fecha_box(f"Fin setup ({semanas_setup} sem.)", fecha_fin_setup, "#fef3c7", "#92400e", "Onboarding")}
      {arrow}
      {_fecha_box("Prospección efectiva", fecha_inicio_prosp, "#d1fae5", "#065f46", "Empieza SDR")}
      {arrow}
      {_fecha_box(f"Fin contrato ({meses} m.)", fecha_fin_contrato, "#f3e8ff", "#6b21a8", f"{semanas_efectivas} sem. efectivas")}
    </div>
    """, unsafe_allow_html=True)

    # ── Métricas resumen ───────────────────────────────────────────────────
    st.markdown("<br>", unsafe_allow_html=True)
    fmt = lambda v: f"${v:,.0f}"

    ka, kb, kc, kd, ke = st.columns(5)
    ka.metric("Ingreso fijo total", f"{fmt(ingreso_fijo_total)} {moneda}")
    kb.metric("Ingreso variable", f"{fmt(ingreso_variable)} {moneda}",
              help=f"{reu_adicionales_total} reuniones adicionales × {fmt(costo_reu)} {moneda}")
    kc.metric("Ingreso total estimado", f"{fmt(ingreso_total)} {moneda}")
    kd.metric("Meta semanal", f"{meta_semana} reu/sem")
    ke.metric("Ticket promedio/reunión", f"{fmt(ticket_por_reunion)} {moneda}" if ticket_por_reunion else "—")

    # ── Tabla mes a mes ────────────────────────────────────────────────────
    if meta_mes > 0 and meses > 0:
        st.markdown("---")
        st.markdown("##### Proyección mes a mes")

        rows = []
        for i in range(int(meses)):
            mes_num = i + 1
            fecha_mes = fecha_inicio_prosp + dt.timedelta(days=i * 30)
            en_setup = fecha_mes < fecha_fin_setup
            reu_mes_esperadas = 0 if en_setup else meta_mes
            reu_adicionales_mes = max(0, reu_mes_esperadas - reu_incluidas_mes)
            ingreso_mes = monto_fijo + (reu_adicionales_mes * costo_reu)
            rows.append({
                "Mes": mes_num,
                "Período": fecha_mes.strftime("%b %Y"),
                "Fase": "Setup" if en_setup else "Prospección",
                f"Reuniones objetivo": reu_mes_esperadas,
                f"Incluidas en fijo": reu_incluidas_mes if not en_setup else 0,
                "Adicionales": reu_adicionales_mes,
                f"Ingreso estimado ({moneda})": f"{ingreso_mes:,.0f}",
            })

        import pandas as pd
        df = pd.DataFrame(rows)
        st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Guardar ────────────────────────────────────────────────────────────
    st.markdown("---")
    if st.button("💾 Guardar datos comerciales", type="primary", key="com_guardar"):
        data = {
            "cliente":                   nombre,
            "moneda":                    moneda,
            "monto_fijo":                monto_fijo,
            "meses_contrato":            int(meses),
            "semanas_setup":             int(semanas_setup),
            "fecha_inicio_contrato":     fecha_inicio.isoformat(),
            "fecha_fin_setup":           fecha_fin_setup.isoformat(),
            "fecha_inicio_prospeccion":  fecha_inicio_prosp.isoformat(),
            "fecha_fin_contrato":        fecha_fin_contrato.isoformat(),
            "reuniones_garantizadas":    reu_garantizadas,
            "reuniones_incluidas_fijo":  reu_incluidas_mes,
            "meta_reuniones_mes":        meta_mes,
            "meta_reuniones_semana":     meta_semana,
            "costo_reunion":             costo_reu,
            "ingreso_fijo_total":        ingreso_fijo_total,
            "ingreso_variable_estimado": ingreso_variable,
            "ingreso_total_estimado":    ingreso_total,
            "ticket_promedio_reunion":   ticket_por_reunion,
            "semanas_efectivas":         semanas_efectivas,
            "notas":                     notas,
        }
        comercial_file.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        # Sincronizar campos clave al estado del cliente
        actualizar_campo(path, "fecha_inicio_contrato",    fecha_inicio.isoformat())
        actualizar_campo(path, "fecha_inicio_prospeccion", fecha_inicio_prosp.isoformat())
        actualizar_campo(path, "fecha_fin_contrato",       fecha_fin_contrato.isoformat())
        actualizar_campo(path, "monto_fijo",               monto_fijo)
        actualizar_campo(path, "moneda",                   moneda)
        actualizar_campo(path, "reuniones_garantizadas",   reu_garantizadas)
        st.success("✅ Guardado en 07_BASE_DATOS/comercial.json y sincronizado al estado del cliente.")
        st.rerun()

    # Mostrar JSON guardado
    if comercial_file.exists():
        with st.expander("🗂️ Ver JSON guardado (para Supabase)", expanded=False):
            st.json(json.loads(comercial_file.read_text(encoding="utf-8")))
            st.download_button("⬇️ Descargar comercial.json",
                               data=comercial_file.read_bytes(),
                               file_name=f"comercial_{nombre.replace(' ','_')}.json",
                               mime="application/json", key="dl_com")


# TAB: ARCHIVOS GENERADOS
# ─────────────────────────────────────────────
def tab_archivos_generados(path: Path, est: dict):
    st.markdown("#### 📄 Archivos Generados")
    st.caption("Todos los archivos base del cliente. Descárgalos o ábrelos directamente en la carpeta.")

    if st.button("📂 Abrir carpeta del cliente"):
        import subprocess
        subprocess.Popen(f'explorer "{path}"')

    st.markdown(f"**Ubicación:** `{path}`")
    st.markdown("---")

    secciones = {
        "🗂️ Admin": "01_ADMIN_CLIENTE",
        "🔍 Análisis": "03_ANALISIS_CLIENTE",
        "🎯 ICP": "04_ICP_ESTRATEGIA",
        "📨 Mensajería": "05_MENSAJERIA_COMERCIAL",
        "📋 Playbook": "06_PLAYBOOK_SDR",
        "🔎 Búsquedas": "07_APOLLO_Y_BUSQUEDAS",
        "💼 Brief cliente": "13_BRIEF_CLIENTE_INTERACTIVO",
        "🗄️ Supabase": "14_SUPABASE_METABASE",
    }

    for titulo, carpeta in secciones.items():
        p = path / carpeta
        if p.exists():
            archivos = [f for f in sorted(p.iterdir()) if f.is_file()]
            if archivos:
                with st.expander(f"{titulo} ({len(archivos)})", expanded=False):
                    cols = st.columns(3)
                    for i, f in enumerate(archivos):
                        with cols[i % 3]:
                            ext = f.suffix.lower()
                            ico = {"md": "📝", "json": "🔷", "html": "🌐", "sql": "🗄️"}.get(ext[1:], "📄")
                            try:
                                contenido = f.read_text(encoding="utf-8")
                                st.download_button(
                                    f"{ico} {f.name}",
                                    data=contenido.encode("utf-8"),
                                    file_name=f.name,
                                    key=f"dl2_{carpeta}_{f.name}",
                                    use_container_width=True,
                                )
                            except:
                                st.caption(f"{ico} {f.name}")


# ─────────────────────────────────────────────
# TAB: ESTADO Y DATOS
# ─────────────────────────────────────────────
def tab_estado_datos(path: Path, est: dict):
    st.markdown("#### 🗄️ Estado del cliente")

    nombre = est.get("nombre_cliente", "")
    color  = est.get("color_marca", "#1a56db")

    # ── 1. PROGRESO POR MÓDULO ─────────────────────────────────────────────
    modulos = [
        ("Archivos",    "estado_archivos",   "📤"),
        ("Análisis web","estado_analisis",   "🌐"),
        ("ICP",         "estado_icp",        "🎯"),
        ("Bases",       "estado_apollo",     "📊"),
        ("Mensajería",  "estado_mensajeria", "✉️"),
        ("Playbook",    "estado_playbook",   "📋"),
        ("Firma",       "estado_firma",      "✍️"),
        ("Branding",    "estado_branding",   "🎨"),
    ]
    st.markdown("##### Módulos")
    cols = st.columns(4)
    for i, (label, campo, ico) in enumerate(modulos):
        val = est.get(campo, "pendiente")
        ok  = val not in ("pendiente", "", None)
        bg  = "#d1fae5" if ok else "#f8fafc"
        tc  = "#065f46" if ok else "#94a3b8"
        brd = "#10b981" if ok else "#e2e8f0"
        em  = "✅" if ok else "⏳"
        with cols[i % 4]:
            st.markdown(
                f'<div style="background:{bg};border:1.5px solid {brd};border-radius:8px;'
                f'padding:10px 8px;text-align:center;font-size:11px;margin:3px 0;">'
                f'<div style="font-size:18px;">{em}</div>'
                f'<div style="font-weight:700;color:{tc};margin-top:3px;">{ico} {label}</div>'
                f'<div style="color:{tc};font-size:10px;margin-top:2px;">{val}</div>'
                f'</div>',
                unsafe_allow_html=True)

    # ── 2. RESUMEN ICP ──────────────────────────────────────────────────────
    icp_macro_cargos = est.get("icp_macro_cargos", "")
    icp_cargos       = est.get("icp_cargos", "")
    icp_industrias   = est.get("icp_industrias", "")
    icp_micro        = est.get("icp_micro_industrias", "")
    icp_tamano       = est.get("icp_tamano_empresa", "")
    icp_paises       = est.get("icp_paises_foco", est.get("pais_objetivo", ""))
    icp_descarte     = est.get("icp_criterios_descarte", "")
    icp_prioridad    = est.get("icp_criterios_prioridad", "")

    tiene_icp = any([icp_industrias, icp_cargos, icp_macro_cargos])

    st.markdown("---")
    st.markdown("##### ICP consolidado")

    if not tiene_icp:
        st.markdown('<div style="color:#94a3b8;font-size:13px;">ICP aún no definido. Completa el tab ICP primero.</div>', unsafe_allow_html=True)
    else:
        icp1, icp2 = st.columns(2)
        with icp1:
            def _icp_chip(label, valor, bg, tc):
                if not valor: return
                items = [v.strip() for v in valor.replace("\n", ",").split(",") if v.strip()]
                chips = "".join(
                    f'<span style="display:inline-block;background:{bg};color:{tc};font-size:11px;'
                    f'font-weight:600;padding:3px 10px;border-radius:20px;margin:2px;">{it}</span>'
                    for it in items
                )
                st.markdown(
                    f'<div style="margin-bottom:10px;">'
                    f'<div style="font-size:11px;font-weight:700;color:#64748b;text-transform:uppercase;'
                    f'letter-spacing:0.5px;margin-bottom:4px;">{label}</div>{chips}</div>',
                    unsafe_allow_html=True)

            _icp_chip("Macro cargos", icp_macro_cargos, "#dbeafe", "#1e40af")
            _icp_chip("Cargos específicos", icp_cargos, "#eff6ff", "#3b82f6")
            _icp_chip("Países", icp_paises, "#f0fdf4", "#15803d")
            if icp_tamano:
                st.markdown(f'<div style="font-size:12px;color:#64748b;margin-top:4px;">📏 <b>Tamaño:</b> {icp_tamano}</div>', unsafe_allow_html=True)

        with icp2:
            _icp_chip("Macro industrias", icp_industrias, "#f0fdf4", "#15803d")
            _icp_chip("Micro nichos", icp_micro, "#fdf4ff", "#7e22ce")
            if icp_prioridad:
                st.markdown(
                    f'<div style="background:#fef3c7;border-left:3px solid #f59e0b;border-radius:6px;'
                    f'padding:8px 12px;margin-top:6px;font-size:12px;color:#92400e;">'
                    f'<b>Priorizar:</b> {icp_prioridad}</div>', unsafe_allow_html=True)
            if icp_descarte:
                st.markdown(
                    f'<div style="background:#fee2e2;border-left:3px solid #ef4444;border-radius:6px;'
                    f'padding:8px 12px;margin-top:6px;font-size:12px;color:#991b1b;">'
                    f'<b>Descartar:</b> {icp_descarte}</div>', unsafe_allow_html=True)

    # ── 3. FOCOS ESTRATÉGICOS (top 3 segmentos cruzados) ───────────────────
    st.markdown("---")
    st.markdown("##### Focos de prospección")

    # Construir focos cruzando cargos × industrias
    cargos_list = [c.strip() for c in icp_macro_cargos.replace("\n",",").split(",") if c.strip()][:3]
    ind_list    = [i.strip() for i in icp_industrias.replace("\n",",").split(",") if i.strip()][:3]
    paises_list = [p.strip() for p in icp_paises.replace("\n",",").split(",") if p.strip()]

    if cargos_list and ind_list:
        focos = []
        for idx in range(max(len(cargos_list), len(ind_list))):
            cargo = cargos_list[idx] if idx < len(cargos_list) else cargos_list[-1]
            ind   = ind_list[idx] if idx < len(ind_list) else ind_list[-1]
            pais  = paises_list[0] if paises_list else ""
            focos.append({"n": idx+1, "cargo": cargo, "industria": ind, "pais": pais})
        focos = focos[:3]

        fc_cols = st.columns(len(focos))
        iconos_foco = ["🥇", "🥈", "🥉"]
        for fc, col_fc in zip(focos, fc_cols):
            with col_fc:
                st.markdown(
                    f'<div style="background:#fff;border:2px solid {color};border-radius:10px;'
                    f'padding:14px 12px;text-align:center;">'
                    f'<div style="font-size:20px;">{iconos_foco[fc["n"]-1]}</div>'
                    f'<div style="font-size:11px;font-weight:700;color:{color};text-transform:uppercase;'
                    f'letter-spacing:0.5px;margin:4px 0;">Foco {fc["n"]}</div>'
                    f'<div style="font-size:13px;font-weight:700;color:#0f172a;">{fc["cargo"]}</div>'
                    f'<div style="font-size:12px;color:#64748b;margin-top:2px;">{fc["industria"]}</div>'
                    f'{"<div style=font-size:11px;color:#94a3b8;margin-top:2px;>" + fc["pais"] + "</div>" if fc["pais"] else ""}'
                    f'</div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div style="color:#94a3b8;font-size:13px;">Define el ICP para ver los focos automáticamente.</div>', unsafe_allow_html=True)

    # ── 4. QUÉ FALTA PEDIRLE AL CLIENTE ────────────────────────────────────
    st.markdown("---")
    st.markdown("##### Información pendiente del cliente")

    faltantes = []
    if not est.get("resumen_servicio") and not est.get("descripcion_empresa"):
        faltantes.append(("Descripción del servicio / empresa", "Necesario para análisis, mensajería y playbook"))
    if not est.get("propuesta_valor"):
        faltantes.append(("Propuesta de valor", "Qué lo diferencia de la competencia"))
    if not est.get("problema_que_resuelve"):
        faltantes.append(("Problema principal que resuelve", "Base para mensajería y playbook"))
    if not est.get("ticket_promedio"):
        faltantes.append(("Ticket promedio / precio", "Para filtrar ICP y armar propuesta de valor"))
    logos = listar_logos(path)
    if not logos:
        faltantes.append(("Logo en alta resolución", "Para firma, playbook y materiales"))
    if not icp_industrias:
        faltantes.append(("Industrias objetivo", "Necesario para armar bases y segmentación"))
    if not icp_cargos:
        faltantes.append(("Cargos específicos a prospectar", "Necesario para armar bases"))
    if not est.get("objetivo_comercial"):
        faltantes.append(("Objetivo comercial (ej: 20 reuniones en 3 meses)", "Para alinear expectativas"))

    if faltantes:
        for label, razon in faltantes:
            st.markdown(
                f'<div style="background:#fef9f0;border-left:3px solid #f59e0b;border-radius:6px;'
                f'padding:8px 14px;margin:4px 0;font-size:12px;">'
                f'<b style="color:#92400e;">⚠️ {label}</b> — '
                f'<span style="color:#78350f;">{razon}</span></div>',
                unsafe_allow_html=True)
    else:
        st.markdown('<div style="background:#f0fdf4;border-left:3px solid #22c55e;border-radius:6px;padding:10px 14px;font-size:12px;color:#15803d;"><b>✅ Información completa</b> — No hay campos críticos pendientes.</div>', unsafe_allow_html=True)

    # ── 5. PENDIENTES Y PRÓXIMOS PASOS ─────────────────────────────────────
    pendientes = est.get("pendientes", [])
    proximos   = est.get("proximos_pasos", [])
    if pendientes or proximos:
        st.markdown("---")
        pp1, pp2 = st.columns(2)
        with pp1:
            if pendientes:
                st.markdown("##### ⏳ Pendientes")
                for p in pendientes:
                    st.markdown(f'<div style="font-size:12px;color:#374151;padding:3px 0;">• {p}</div>', unsafe_allow_html=True)
        with pp2:
            if proximos:
                st.markdown("##### ▶️ Próximos pasos")
                for i, p in enumerate(proximos, 1):
                    st.markdown(f'<div style="font-size:12px;color:#374151;padding:3px 0;"><b>{i}.</b> {p}</div>', unsafe_allow_html=True)

    # ── 6. DATOS RAW (colapsable) ───────────────────────────────────────────
    st.markdown("---")
    with st.expander("🗂️ Ver datos estructurados (JSON)", expanded=False):
        t1, t2 = st.tabs(["estado_cliente.json", "playbook_log.json"])
        with t1:
            ruta = path / "estado_cliente.json"
            if ruta.exists():
                data = json.loads(ruta.read_text(encoding="utf-8"))
                st.json(data)
                st.download_button("⬇️ Descargar", data=json.dumps(data, ensure_ascii=False, indent=2).encode(),
                                   file_name="estado_cliente.json", mime="application/json", key="dl_est")
        with t2:
            ruta = path / "07_BASE_DATOS/playbook_log.json"
            if ruta.exists():
                data = json.loads(ruta.read_text(encoding="utf-8"))
                st.json(data)
                st.download_button("⬇️ Descargar", data=json.dumps(data, ensure_ascii=False, indent=2).encode(),
                                   file_name="playbook_log.json", mime="application/json", key="dl_pb")


# ─────────────────────────────────────────────
# TAB: CHAT
# ─────────────────────────────────────────────
def _chat_system_prompt(est: dict, path: Path) -> str:
    nombre    = est.get("nombre_cliente", "")
    web       = est.get("sitio_web", "")
    paises    = est.get("pais_objetivo", "")
    objetivo  = est.get("objetivo_comercial", "")
    resumen   = est.get("resumen_servicio", "") or est.get("descripcion_empresa", "")
    propuesta = est.get("propuesta_valor", "")
    problema  = est.get("problema_que_resuelve", "")
    diferenc  = est.get("diferenciacion", "")
    ticket    = est.get("ticket_promedio", "")
    icp_cargos= est.get("icp_cargos", "") or est.get("icp_macro_cargos", "")
    icp_ind   = est.get("icp_industrias", "")
    icp_tam   = est.get("icp_tamano_empresa", "")
    icp_pais  = est.get("icp_paises_foco", paises)
    icp_desc  = est.get("icp_criterios_descarte", "")
    icp_prio  = est.get("icp_criterios_prioridad", "")
    analisis  = ""
    analisis_file = path / "03_ANALISIS_CLIENTE" / "analisis_web.md"
    if analisis_file.exists():
        try: analisis = analisis_file.read_text(encoding="utf-8", errors="ignore")[:4000]
        except: pass

    return f"""Eres un experto senior en prospección B2B y SDR outbound para la agencia Conprospección.
Tienes todo el contexto del cliente {nombre} ya cargado. Responde preguntas directas y accionables.

CLIENTE: {nombre} | {web} | Países: {paises}
Objetivo: {objetivo}

ANÁLISIS:
- Servicio: {resumen or '(no completado)'}
- Propuesta de valor: {propuesta or '(no completado)'}
- Problema que resuelve: {problema or '(no completado)'}
- Diferenciación: {diferenc or '(no completado)'}
- Ticket promedio: {ticket or '(no especificado)'}

ICP DEFINIDO:
- Cargos objetivo: {icp_cargos or '(pendiente)'}
- Industrias: {icp_ind or '(pendiente)'}
- Tamaño empresa: {icp_tam or '(pendiente)'}
- Países foco: {icp_pais or '(pendiente)'}
- Priorizar: {icp_prio or '(pendiente)'}
- Descartar: {icp_desc or '(pendiente)'}
{f"ANÁLISIS WEB:{chr(10)}{analisis}" if analisis else ""}

INSTRUCCIONES:
- Responde siempre en español
- Sé directo y accionable — nada genérico
- Cuando des listas para Apollo, usa formato separado por comas listo para copiar
- Si te piden ajustar el ICP, razona brevemente y da la versión ajustada
- Mantén consistencia con el ICP y cliente ya definido"""


def tab_chat(path: Path, est: dict):
    st.markdown("#### 💬 Chat Contextual")

    client = get_claude_client()
    if not client:
        st.warning("Configura ANTHROPIC_API_KEY en .env para usar el chat.")
        return

    nombre = est.get("nombre_cliente", "")
    hist_key = f"chat_hist_{path.name}"
    if hist_key not in st.session_state:
        st.session_state[hist_key] = []

    def _enviar_chat(msg_texto: str):
        if not msg_texto.strip(): return
        st.session_state[hist_key].append({"role": "user", "content": msg_texto.strip()})
        sistema = _chat_system_prompt(est, path)
        with st.spinner("Pensando..."):
            resp = llamar_claude(client, st.session_state[hist_key], sistema, max_tokens=3000)
        st.session_state[hist_key].append({"role": "assistant", "content": resp})
        st.rerun()

    # Chips rápidos — auto-envían
    sugeridas = [
        f"Dame 20 empresas reales que encajen con el ICP en {est.get('pais_objetivo','Chile')}, con nombre, dominio, industria y cargo a contactar",
        "Dame los cargos específicos separados por coma para Apollo",
        "Dame las industrias separadas por coma para Apollo",
        "Dame los dominios de 15 empresas objetivo separados por coma",
        "¿Qué información falta pedirle al cliente?",
        "Dame criterios Tier A, B y C para priorizar leads",
        "¿Qué debería ir a WhatsApp vs email en la secuencia?",
        "Consolida en máximo 4 combinaciones cargo+industria para Snov.io",
    ]
    st.markdown('<div style="font-size:12px;color:#64748b;margin-bottom:6px;">Acciones rápidas:</div>', unsafe_allow_html=True)
    cols_s = st.columns(4)
    for i, preg in enumerate(sugeridas):
        label = preg[:35] + "…" if len(preg) > 35 else preg
        with cols_s[i % 4]:
            if st.button(label, key=f"chat_chip_{i}", use_container_width=True):
                _enviar_chat(preg)

    st.markdown("---")

    # Historial
    hist = st.session_state[hist_key]
    if hist:
        for msg in hist:
            if msg["role"] == "user":
                st.markdown(
                    f'<div style="background:#dbeafe;border-radius:10px 10px 0 10px;padding:10px 14px;'
                    f'margin:6px 0 2px auto;max-width:85%;font-size:13px;color:#1e3a5f;">'
                    f'<b>Tú:</b> {msg["content"]}</div>', unsafe_allow_html=True)
            else:
                st.markdown(
                    f'<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px 10px 10px 0;'
                    f'padding:10px 14px;margin:2px 0 6px;font-size:13px;color:#1e293b;white-space:pre-wrap;">'
                    f'{msg["content"]}</div>', unsafe_allow_html=True)
        if st.button("🗑️ Limpiar", key="chat_clear"):
            st.session_state[hist_key] = []
            st.rerun()
    else:
        st.markdown('<div style="background:#f8fafc;border:2px dashed #e2e8f0;border-radius:12px;'
                    'padding:20px;text-align:center;color:#94a3b8;font-size:13px;">'
                    '💬 Usa un botón rápido o escribe / habla tu pregunta abajo</div>', unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Voz del servidor
    cv1, cv2 = st.columns([2, 1])
    with cv2:
        dur_chat = st.selectbox("Duración", [10, 20, 30, 45], index=1, key="chat_dur",
                                format_func=lambda x: f"{x} seg")
    with cv1:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🎤 Grabar y enviar", key="chat_grabar", use_container_width=True):
            with st.spinner(f"🔴 Grabando {dur_chat} seg — habla ahora..."):
                texto_voz = grabar_y_transcribir(duracion=dur_chat)
            if texto_voz.startswith("ERROR"):
                st.error(f"No se pudo transcribir: {texto_voz}")
            else:
                st.info(f"🎤 Reconocido: *{texto_voz}*")
                _enviar_chat(texto_voz)

    # Input texto
    col1, col2 = st.columns([5, 1])
    with col1:
        pregunta = st.text_area("", placeholder=f"O escribe tu pregunta sobre {nombre}...",
                                height=80, key="chat_input", label_visibility="collapsed")
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("Enviar →", key="chat_send", type="primary", use_container_width=True):
            _enviar_chat(pregunta)


# ─────────────────────────────────────────────
# TAB: ROADMAP
# ─────────────────────────────────────────────
def tab_roadmap(path: Path, est: dict):
    st.markdown("#### 🗺️ Integraciones Futuras")

    items = [
        ("🌐 Análisis web automático", "Scraping + Claude API para analizar la web sin copiar/pegar", "Alta", "#fee2e2"),
        ("🎯 Apollo API", "Búsqueda y exportación de contactos según ICP aprobado", "Alta", "#fee2e2"),
        ("📧 Snov.io", "Exportar bases calificadas a campañas de email", "Media", "#fef3c7"),
        ("🗄️ Supabase Sync", "Sincronización automática hacia Supabase — tablas: clientes, reuniones, sdrs, comisiones", "Alta", "#fee2e2"),
        ("📊 Metabase", "Dashboards de reuniones, progreso, actividad SDR y comisiones", "Alta", "#fee2e2"),
        ("💬 WhatsApp Campaign Builder", "Generador de campañas con segmentos y plantillas", "Media", "#fef3c7"),
        ("✅ Validación de reuniones", "Portal para que el cliente marque válida/no válida/comercial y actualice GHL", "Alta", "#fee2e2"),
        ("👩‍💻 Dashboard SDR", "Métricas de actividad, reuniones por semana y cálculo de comisiones", "Alta", "#fee2e2"),
        ("📂 Clasificación masiva de bases", "Filtrar 20,000+ contactos según ICP automáticamente", "Media", "#fef3c7"),
        ("🎙️ tl;dv / grabación reuniones", "Transcripción y calificación automática de reuniones agendadas", "Media", "#fef3c7"),
    ]

    for nombre, desc, prio, color in items:
        st.markdown(f"""
        <div style="background:{color};border-radius:10px;padding:12px 16px;margin:6px 0;display:flex;align-items:flex-start;gap:12px;">
          <div style="flex:1;">
            <b style="font-size:14px;">{nombre}</b>
            <span style="background:white;color:#374151;font-size:10px;font-weight:700;padding:1px 8px;border-radius:999px;margin-left:8px;">{prio.upper()}</span>
            <div style="font-size:12px;color:#4b5563;margin-top:3px;">{desc}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    sidebar()

    page = st.session_state.page
    if page == "inicio":
        page_inicio()
    elif page == "nuevo_cliente":
        page_nuevo_cliente()
    elif page == "cliente":
        cid = st.session_state.get("cliente_id")
        if cid:
            page_cliente(cid)
        else:
            ir("inicio")
            st.rerun()
    else:
        page_inicio()


main()
